# --- IMPORTS ---
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import os
import re
import base64
from datetime import datetime
import time
import json
import tempfile
import io
import zipfile
import requests
import warnings
from scipy import stats
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib.dates as mdates
from pandas.api.types import is_numeric_dtype, is_datetime64_any_dtype
from pandasql import sqldf
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics import renderPDF
try:
    import torch
    from transformers import AutoTokenizer, AutoModelForCausalLM
    TORCH_AVAILABLE = True
except ImportError:
    torch = None
    AutoTokenizer = None
    AutoModelForCausalLM = None
    TORCH_AVAILABLE = False
import matplotlib
matplotlib.use('Agg')  # Utiliser le backend non interactif pour éviter les problèmes de thread
import google.generativeai as genai
import openai
from typing import Optional, Dict, Any
from dotenv import load_dotenv

# --- Moteur d'analyse intelligente (nouveau module dédié) ---
from smart_analysis import (
    profile_dataframe,
    generate_smart_insights,
    render_smart_dashboard,
    build_executive_summary,
    build_survey_context,
)
from report_builder import build_word_report
import ai_interpreter
from pathlib import Path
try:
    from ccc_analysis_tool import FIPAAnalysisPipeline, AnalysisStage, AIProvider
    PIPELINE_DISPONIBLE = True
except Exception as e:
    PIPELINE_DISPONIBLE = False
    print(f"Pipeline avancé non chargé : {e}")

# Charger les variables d'environnement
load_dotenv()

# Gemini : préférer les variantes "flash" (rapides / quota) plutôt que "pro"
DEFAULT_GEMINI_MODEL = "gemini-1.5-flash"
DEFAULT_OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")


def get_effective_gemini_model_name() -> str:
    """Modèle Gemini effectif : sélection sidebar > GEMINI_MODEL dans .env > défaut flash."""
    try:
        for key in ("gemini_model_new", "gemini_model_legacy"):
            v = st.session_state.get(key)
            if v:
                return str(v)
    except Exception:
        pass
    return os.getenv("GEMINI_MODEL", DEFAULT_GEMINI_MODEL)


def _looks_like_gemini_quota_error(err: Exception | str) -> bool:
    s = str(err or "").lower()
    return (
        "429" in s
        and ("quota" in s or "rate limit" in s or "rate-limits" in s or "exceeded" in s)
    )


def _has_openai_key() -> bool:
    return bool(os.environ.get("OPENAI_API_KEY") or getattr(openai, "api_key", None))


def _norm_id_basic(s: str) -> str:
    """Normalise un identifiant (accents/ponctuation) pour matcher des colonnes."""
    import re as _re
    import unicodedata as _ud

    s = str(s or "").strip().strip('"').strip()
    s = _ud.normalize("NFKD", s)
    s = "".join(ch for ch in s if not _ud.combining(ch))
    s = s.lower()
    s = _re.sub(r"[^a-z0-9_]+", "_", s)
    s = _re.sub(r"_+", "_", s).strip("_")
    return s


def extract_columns_from_sql(sql: str, df_columns: list[str]) -> list[str]:
    """
    Extrait les colonnes référencées dans une requête SQL (principalement via "colonne").
    Retourne des noms de colonnes tels qu'ils existent dans df_columns (originaux).
    """
    import re as _re

    if not sql:
        return []

    norm_map = {_norm_id_basic(c): c for c in (df_columns or [])}
    used: list[str] = []

    # 1) Identifiants entre guillemets doubles
    for m in _re.finditer(r'"([^"]+)"', sql):
        inner = m.group(1)
        n = _norm_id_basic(inner)
        if n in norm_map:
            col = norm_map[n]
            if col not in used:
                used.append(col)

    # 2) Identifiants non quotés (fallback) - prudence: mots SQL
    sql_keywords = {
        "select","from","where","group","by","order","limit","as","and","or","in","is","null",
        "count","sum","avg","min","max","distinct","having","join","left","right","inner","outer","on",
        "case","when","then","else","end","desc","asc","like","between",
        "data",
    }
    for tok in _re.findall(r"\b[a-zA-Z_][a-zA-Z0-9_]*\b", sql):
        t = tok.lower()
        if t in sql_keywords:
            continue
        n = _norm_id_basic(tok)
        if n in norm_map:
            col = norm_map[n]
            if col not in used:
                used.append(col)

    return used


# Configuration des modèles IA
class AIModelHandler:
    def __init__(self, provider: str = "gemini"):
        """
        Initialise le gestionnaire de modèles d'IA
        
        Args:
            provider: "gemini", "openai", ou "local" - le fournisseur à utiliser par défaut
        """
        self.provider = provider
        self._setup_providers()
    
    def _setup_providers(self):
        """Configure les clés API pour les différents fournisseurs"""
        # Configuration de Gemini
        if 'GOOGLE_API_KEY' in os.environ:
            genai.configure(api_key=os.environ['GOOGLE_API_KEY'])
        
        # Configuration d'OpenAI
        if 'OPENAI_API_KEY' in os.environ:
            openai.api_key = os.environ['OPENAI_API_KEY']
    
    def generate_text(
        self,
        prompt: str,
        provider: Optional[str] = None,
        model: str = None,
        **kwargs
    ) -> str:
        """
        Génère du texte en utilisant le fournisseur spécifié
        
        Args:
            prompt: Le texte d'entrée
            provider: "gemini", "openai", ou "local" (utilise self.provider si None)
            model: Le modèle à utiliser (par défaut dépend du fournisseur)
            **kwargs: Arguments supplémentaires pour la génération
            
        Returns:
            Le texte généré
        """
        provider = provider or self.provider
        provider = provider.lower()
        
        if provider == "gemini":
            return self._generate_with_gemini(prompt, model, **kwargs)
        elif provider == "openai":
            return self._generate_with_openai(prompt, model, **kwargs)
        elif provider == "local":
            return self._generate_with_local_model(prompt, **kwargs)
        else:
            raise ValueError(f"Fournisseur non supporté: {provider}")
    
    def _generate_with_gemini(
        self,
        prompt: str,
        model: Optional[str] = None,
        **kwargs
    ) -> str:
        """Génère du texte avec Google Gemini"""
        try:
            model = model or get_effective_gemini_model_name()
            model_obj = genai.GenerativeModel(model)
            response = model_obj.generate_content(prompt, **kwargs)
            return response.text
        except Exception as e:
            return f"Erreur Gemini: {str(e)}"
    
    def _generate_with_openai(
        self,
        prompt: str,
        model: str = DEFAULT_OPENAI_MODEL,
        temperature: float = 0.7,
        max_tokens: int = 1000,
        **kwargs
    ) -> str:
        """Génère du texte avec OpenAI"""
        try:
            api_key = os.environ.get("OPENAI_API_KEY") or getattr(openai, "api_key", None)
            if not api_key:
                return "Clé API OpenAI non configurée"

            # SDK OpenAI >= 1.x / 2.x
            try:
                from openai import OpenAI  # type: ignore

                client = OpenAI(api_key=api_key)
                resp = client.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": prompt}],
                    temperature=temperature,
                    max_tokens=max_tokens,
                )
                return (resp.choices[0].message.content or "").strip()
            except Exception:
                # Fallback anciens SDK (si présent)
                if model.startswith("gpt-") and hasattr(openai, "ChatCompletion"):
                    response = openai.ChatCompletion.create(
                        model=model,
                        messages=[{"role": "user", "content": prompt}],
                        temperature=temperature,
                        max_tokens=max_tokens,
                        **kwargs,
                    )
                    return response.choices[0].message.content
                if hasattr(openai, "Completion"):
                    response = openai.Completion.create(
                        model=model,
                        prompt=prompt,
                        temperature=temperature,
                        max_tokens=max_tokens,
                        **kwargs,
                    )
                    return response.choices[0].text

                raise
        except Exception as e:
            return f"Erreur OpenAI: {str(e)}"
    
    def _generate_with_local_model(
        self,
        prompt: str,
        **kwargs
    ) -> str:
        """Génère du texte avec le modèle local"""
        try:
            global local_gemma_model, local_gemma_tokenizer
            
            if local_gemma_model is None or local_gemma_tokenizer is None:
                load_local_gemma()
            
            if local_gemma_model is None:
                return "Modèle local non disponible"
            
            inputs = local_gemma_tokenizer.encode(prompt, return_tensors="pt")
            outputs = local_gemma_model.generate(
                inputs,
                max_length=inputs.shape[1] + 200,
                num_return_sequences=1,
                temperature=0.7,
                do_sample=True
            )
            response = local_gemma_tokenizer.decode(outputs[0], skip_special_tokens=True)
            return response[len(prompt):].strip()
        except Exception as e:
            return f"Erreur modèle local: {str(e)}"

# Instance globale du gestionnaire IA
ai_handler = None

def get_ai_handler():
    """Retourne l'instance du gestionnaire IA"""
    global ai_handler
    if ai_handler is None:
        # Par défaut, essayer Gemini, puis OpenAI, puis local
        if 'GOOGLE_API_KEY' in os.environ:
            ai_handler = AIModelHandler(provider="gemini")
        elif 'OPENAI_API_KEY' in os.environ:
            ai_handler = AIModelHandler(provider="openai")
        else:
            ai_handler = AIModelHandler(provider="local")
    return ai_handler

# Configuration du modèle local Gemma (dossier Hugging Face ou snapshot local)
LOCAL_GEMMA_PATH = os.getenv("LOCAL_GEMMA_PATH", "")
local_gemma_loaded = False
local_gemma_model = None
local_gemma_tokenizer = None
local_gemma_last_error = None


@st.cache_resource(show_spinner=False)
def _load_local_gemma_cached(model_path: str):
    """
    Charge et met en cache le modèle/tokenizer pour éviter le rechargement à chaque rerun Streamlit.
    Note: on force un chargement CPU float32 par défaut pour maximiser la stabilité (évite crashs/AV sur float16 CPU).
    """
    if not TORCH_AVAILABLE or not model_path:
        return None, None
    tok = AutoTokenizer.from_pretrained(model_path)

    use_cuda = bool(getattr(torch, "cuda", None) and torch.cuda.is_available())
    model_kwargs: dict[str, Any] = {"low_cpu_mem_usage": True}
    if use_cuda:
        model_kwargs["torch_dtype"] = torch.float16
        model_kwargs["device_map"] = "auto"
    else:
        model_kwargs["torch_dtype"] = torch.float32

    mdl = AutoModelForCausalLM.from_pretrained(model_path, **model_kwargs)
    if not use_cuda:
        mdl = mdl.to("cpu")
    mdl.eval()
    return mdl, tok

def load_local_gemma():
    """Charge le modèle Gemma local si disponible."""
    global local_gemma_loaded, local_gemma_model, local_gemma_tokenizer, local_gemma_last_error
    
    if local_gemma_loaded:
        return local_gemma_model, local_gemma_tokenizer

    if not TORCH_AVAILABLE:
        local_gemma_last_error = "torch/transformers non installés (optionnel sur le cloud)"
        local_gemma_loaded = True
        return None, None

    if not LOCAL_GEMMA_PATH:
        local_gemma_last_error = "LOCAL_GEMMA_PATH non configuré"
        local_gemma_loaded = True
        return None, None
        
    try:
        local_gemma_last_error = None
        if os.path.exists(LOCAL_GEMMA_PATH):
            st.sidebar.info("Chargement du modèle Gemma local...")
            # Détecter les modèles multimodaux (Gemma3n). On les autorise (sur demande),
            # mais on affiche un avertissement car le chargement peut être très lourd.
            try:
                cfg_path = os.path.join(LOCAL_GEMMA_PATH, "config.json")
                if os.path.exists(cfg_path):
                    with open(cfg_path, "r", encoding="utf-8") as f:
                        cfg = json.load(f)
                    archs = cfg.get("architectures") or []
                    if any("Gemma3n" in str(a) for a in archs):
                        st.sidebar.warning(
                            "Modèle Gemma3n (multimodal) détecté. Chargement possible mais lourd; "
                            "sur CPU cela peut être très lent et consommer beaucoup de RAM."
                        )
            except Exception:
                pass
            # Chargement réel via cache Streamlit (évite rechargements et réduit les risques de déconnexion)
            with st.sidebar.status("Chargement du modèle en cours (peut prendre plusieurs minutes)...", expanded=False):
                local_gemma_model, local_gemma_tokenizer = _load_local_gemma_cached(LOCAL_GEMMA_PATH)
            local_gemma_loaded = True
            st.sidebar.success("Modèle Gemma chargé avec succès!")
        else:
            st.sidebar.warning(f"Le modèle Gemma n'a pas été trouvé dans {LOCAL_GEMMA_PATH}")
    except Exception as e:
        local_gemma_last_error = str(e)
        st.sidebar.error(f"Erreur lors du chargement du modèle Gemma: {str(e)}")
        
    return local_gemma_model, local_gemma_tokenizer


def _local_gemma_ready() -> bool:
    """True si le chemin Gemma existe et le modèle + tokenizer sont chargés."""
    global local_gemma_model, local_gemma_tokenizer
    if not os.path.exists(LOCAL_GEMMA_PATH):
        return False
    if local_gemma_model is not None and local_gemma_tokenizer is not None:
        return True
    load_local_gemma()
    return local_gemma_model is not None and local_gemma_tokenizer is not None


# --- FONCTIONS UTILITAIRES POUR LES VISUALISATIONS ---
def safe_plotly_call(func_name, data_obj, **kwargs):
    """
    Wrapper sécurisé pour les appels à Plotly Express avec gestion des erreurs et fallback Matplotlib.
    
    Args:
        func_name (str): Nom de la fonction Plotly Express à appeler (ex: 'bar', 'line', 'scatter')
        data_obj: Données à utiliser (DataFrame ou autre objet compatible)
        **kwargs: Arguments à passer à la fonction Plotly Express
        
    Returns:
        Figure Plotly ou Matplotlib, ou message d'erreur sous forme de chaîne
    """
    # Récupérer la fonction plotly si px est disponible, sinon None
    try:
        import plotly.express as px
        plotly_func = getattr(px, func_name, None)
    except Exception:
        plotly_func = None
    
    # Si la fonction plotly n'est pas disponible, essayer Matplotlib
    if not plotly_func:
        return _fallback_to_matplotlib(func_name, data_obj, **kwargs)
    
    # Préparation des données pour Plotly
    processed_data = data_obj
    if isinstance(data_obj, pd.DataFrame):
        # Gestion des MultiIndex
        if isinstance(data_obj.index, pd.MultiIndex):
            processed_data = data_obj.reset_index()
        # Nettoyage des noms de colonnes
        if isinstance(processed_data.columns, pd.MultiIndex):
            processed_data = processed_data.copy()
            processed_data.columns = ['_'.join(map(str, col)).strip() for col in processed_data.columns.values]
    
    # Nettoyage des arguments
    processed_kwargs = kwargs.copy()
    
    # Vérification des colonnes pour les DataFrames
    if isinstance(processed_data, pd.DataFrame):
        available_cols = processed_data.columns.tolist()
        
        # Vérifier et corriger les noms de colonnes dans les arguments
        for arg_name, arg_value in list(processed_kwargs.items()):
            if isinstance(arg_value, str) and arg_value not in available_cols:
                # Essayer de trouver une correspondance insensible à la casse
                matching_cols = [col for col in available_cols if col.lower() == arg_value.lower()]
                if matching_cols:
                    processed_kwargs[arg_name] = matching_cols[0]
    
    try:
        # Appel de la fonction Plotly Express - ne pas passer data_frame deux fois
        if 'data_frame' in processed_kwargs:
            del processed_kwargs['data_frame']
        fig = plotly_func(processed_data, **processed_kwargs)
        
        # Personnalisation supplémentaire du graphique
        if 'title' in processed_kwargs:
            fig.update_layout(
                title=processed_kwargs['title'],
                title_x=0.5,
                title_font=dict(size=16, family="Arial, sans-serif")
            )
            
        # Amélioration de la lisibilité
        fig.update_layout(
            margin=dict(l=40, r=40, t=60, b=40),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(family="Arial, sans-serif", size=12, color="#2c3e50")
        )
        
        # Personnalisation spécifique au type de graphique
        if func_name in ['bar', 'line', 'scatter']:
            fig.update_xaxes(showgrid=False, zeroline=False)
            fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='#f0f0f0', zeroline=False)
            
        return fig
        
    except Exception as e:
        error_msg = f"Erreur Plotly ({func_name}): {str(e)}"
        st.warning(f"Erreur avec Plotly, tentative avec Matplotlib... ({str(e)})")
        return _fallback_to_matplotlib(func_name, data_obj, **kwargs)

def _fallback_to_matplotlib(func_name, data_obj, **kwargs):
    """
    Fonction de secours utilisant Matplotlib quand Plotly échoue.
    """
    try:
        import matplotlib.pyplot as plt
        import seaborn as sns
        
        # Configuration du style
        try:
            plt.style.use('seaborn-v0_8')  # Style plus récent
        except:
            plt.style.use('seaborn')  # Fallback pour les versions plus anciennes
        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['Arial']
        plt.rcParams['figure.facecolor'] = 'white'  # Fond blanc pour le graphique
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Gestion des différents types de graphiques
        if func_name == 'bar':
            if isinstance(data_obj, pd.DataFrame) and 'x' in kwargs and 'y' in kwargs:
                x_col = kwargs['x']
                y_col = kwargs['y']
                sns.barplot(data=data_obj, x=x_col, y=y_col, ax=ax, palette='viridis')
                ax.set_xlabel(x_col, fontsize=12)
                ax.set_ylabel(y_col, fontsize=12)
                
                # Rotation des étiquettes si nécessaire
                if len(data_obj[x_col].unique()) > 5:
                    plt.xticks(rotation=45, ha='right')
                    
        elif func_name == 'histogram' and 'x' in kwargs:
            x_col = kwargs['x']
            bins = kwargs.get('nbins', 30)
            sns.histplot(data=data_obj, x=x_col, bins=bins, kde=True, ax=ax, color='#3498db')
            ax.set_xlabel(x_col, fontsize=12)
            ax.set_ylabel('Fréquence', fontsize=12)
            
        elif func_name == 'box' and 'y' in kwargs:
            y_col = kwargs['y']
            sns.boxplot(data=data_obj, y=y_col, ax=ax, color='#3498db')
            ax.set_ylabel(y_col, fontsize=12)
            
        elif func_name == 'scatter' and 'x' in kwargs and 'y' in kwargs:
            x_col = kwargs['x']
            y_col = kwargs['y']
            sns.scatterplot(data=data_obj, x=x_col, y=y_col, ax=ax, alpha=0.6, color='#3498db')
            ax.set_xlabel(x_col, fontsize=12)
            ax.set_ylabel(y_col, fontsize=12)
            
        # Personnalisation du graphique
        if 'title' in kwargs:
            ax.set_title(kwargs['title'], fontsize=14, pad=20)
            
        # Amélioration de la lisibilité
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        
        plt.tight_layout()
        return fig
        
    except Exception as e:
        return f"Erreur lors de la génération du graphique avec Matplotlib: {str(e)}"

# --- UTILITAIRES D'AFFICHAGE SÉCURISÉ (Arrow-compatible) ---
def ensure_arrow_compatible(df: pd.DataFrame) -> pd.DataFrame:
    try:
        if df is None or not isinstance(df, pd.DataFrame):
            return df
        safe = df.copy()
        for col in safe.columns:
            # Forcer les objets mixtes en chaînes pour éviter les erreurs Arrow (ex: "1 3")
            if pd.api.types.is_object_dtype(safe[col]):
                safe[col] = safe[col].astype(str)
        return safe
    except Exception:
        return df

def show_df(df: pd.DataFrame, **kwargs):
    # Wrapper centralisé pour l'affichage des DataFrames
    return st.dataframe(ensure_arrow_compatible(df), **kwargs)

# --- GÉNÉRATION PYTHON (basée uniquement sur les colonnes retenues) ---
def generate_python_script_with_ai(nl_request: str,
                                   df: pd.DataFrame,
                                   all_columns: list,
                                   numerical_cols: list,
                                   categorical_cols: list):
    """
    Génère un script Python basé STRICTEMENT sur les colonnes validées par l'utilisateur.
    Retourne (script, title). Aucun import n'est ajouté; l'environnement fournit df/pd/np.
    """
    try:
        request_lower = (nl_request or "").lower()

        # Colonnes par défaut uniquement parmi les retenues
        cat_col = categorical_cols[0] if categorical_cols else None
        num_col = numerical_cols[0] if numerical_cols else None

        # Cas 0: Demande d'insights/résumé global sur variables sélectionnées
        if any(k in request_lower for k in ["insight", "résumé", "resume", "overview", "vue globale", "synthèse", "synthese", "aperçu global"]):
            title = "Insights des variables sélectionnées"
            # Construire un script qui retourne un dictionnaire de résultats
            keys_blocks = []
            # Limiter le nombre pour lisibilité si trop de colonnes
            sel_cats = (categorical_cols or [])
            sel_nums = (numerical_cols or [])
            for c in sel_cats:
                keys_blocks.append(
f"'Répartition - {c} (table)': df['{c}'].value_counts(dropna=False).reset_index().rename(columns={{'index': '{c}', '{c}': 'Nombre'}})"
                )
                # Graphique barres pour la répartition (wrapper sécurisé)
                keys_blocks.append(
f"'Répartition - {c} (graph)': (safe_plotly_call('bar', df['{c}'].value_counts(dropna=False).reset_index().rename(columns={{'index': '{c}', '{c}': 'Nombre'}}), x='{c}', y='Nombre') if 'safe_plotly_call' in globals() else None) if '{c}' in df.columns else 'Colonne absente'"
                )
            for n in sel_nums:
                keys_blocks.append(
f"'Statistiques - {n} (table)': df['{n}'].describe().to_frame() if '{n}' in df.columns else df.head(0)"
                )
                # Histogramme pour la distribution numérique (wrapper sécurisé)
                keys_blocks.append(
f"'Distribution - {n} (graph)': (safe_plotly_call('histogram', df, x='{n}') if 'safe_plotly_call' in globals() else None) if '{n}' in df.columns else 'Colonne absente'"
                )
            # Si rien n'a été sélectionné, fallback minimal
            if not keys_blocks:
                script = "result = df.head(20)"
            else:
                dict_body = ",\n".join(keys_blocks)
                script = f"""
result = {{
{dict_body}
}}
                """
            return script, title

        # Cas 1: Groupby / comptage (mots-clés forts)
        if any(k in request_lower for k in ["group", "grouper", "par ", "répartition", "count par", "compter par", "value_counts"]):
            group_col = None
            for c in (categorical_cols or []):
                if c and c.lower() in request_lower:
                    group_col = c
                    break
            if not group_col:
                group_col = cat_col
            title = f"Comptage par {group_col}" if group_col else "Comptage des valeurs"
            script = f"""
result = None
if '{group_col}' in df.columns:
    result = (
        df['{group_col}']
          .value_counts(dropna=False)
          .reset_index()
          .rename(columns={{'index': '{group_col}', '{group_col}': 'Nombre'}})
    )
    # Ajouter un graphique barres via wrapper sécurisé
    try:
        fig = safe_plotly_call('bar', result, x='{group_col}', y='Nombre') if 'safe_plotly_call' in globals() else None
        if fig is not None:
            result = {'table': result, 'graph': fig}
    except Exception:
        pass
else:
    result = df.head(20)
            """
            return script, title

        # Cas 2: Statistiques descriptives (mots-clés forts)
        if any(k in request_lower for k in ["stat", "statistique", "describe", "moyenne", "écart-type", "distribution"]):
            target_num = None
            for c in (numerical_cols or []):
                if c and c.lower() in request_lower:
                    target_num = c
                    break
            if not target_num:
                target_num = num_col
            title = f"Statistiques descriptives de {target_num}" if target_num else "Statistiques descriptives"
            script = f"""
result = None
num_cols = [c for c in df.select_dtypes(include=['number']).columns]
if '{target_num}' in df.columns:
    result = df['{target_num}'].describe().to_frame()
elif num_cols:
    result = df[num_cols].describe()
else:
    result = df.head(20)
try:
    fig = (safe_plotly_call('histogram', df, x='{target_num}') if 'safe_plotly_call' in globals() and '{target_num}' in df.columns else None)
    if fig is not None:
        result = {'table': result, 'graph': fig}
except Exception:
    pass
            """
            return script, title

        # Cas 3: Tableau croisé simple (cat vs num) (mots-clés forts)
        if any(k in request_lower for k in ["moyenne par", "sum par", "agrég", "aggregate", "pivot", "tableau croisé"]):
            gcol = cat_col
            vcol = num_col
            title = f"Moyenne de {vcol} par {gcol}" if (gcol and vcol) else "Tableau croisé"
            agg_line = f"df.groupby('{gcol}')['{vcol}'].mean().reset_index(name='Moyenne')" if (gcol and vcol) else "df.head(20)"
            script = f"""
result = {agg_line}
try:
    fig = (safe_plotly_call('bar', result, x='{gcol}', y='Moyenne') if 'safe_plotly_call' in globals() and '{gcol}' in result.columns and 'Moyenne' in result.columns else None)
    if fig is not None:
        result = {'table': result, 'graph': fig}
except Exception:
    pass
            """
            return script, title

        # Heuristique par colonnes sélectionnées (si pas de mots-clés forts)
        if cat_col and not num_col:
            title = f"Comptage par {cat_col}"
            script = f"""
result = (
    df['{cat_col}']
      .value_counts(dropna=False)
      .reset_index()
      .rename(columns={{'index': '{cat_col}', '{cat_col}': 'Nombre'}})
)
            """
            return script, title

        if num_col and not cat_col:
            title = f"Statistiques descriptives de {num_col}"
            script = f"""
num_cols = [c for c in df.select_dtypes(include=['number']).columns]
result = df['{num_col}'].describe().to_frame() if '{num_col}' in df.columns else (df[num_cols].describe() if num_cols else df.head(20))
            """
            return script, title

        if cat_col and num_col:
            title = f"Moyenne de {num_col} par {cat_col}"
            script = f"""
result = df.groupby('{cat_col}')['{num_col}'].mean().reset_index(name='Moyenne')
            """
            return script, title

        # Fallback minimal
        title = "Aperçu des données"
        script = "result = df.head(20)"
        return script, title
    except Exception as e:
        return "result = df.head(20)", f"Analyse Python - Erreur gérée ({e})"

# --- FONCTIONS DE GESTION DE LA CONFIGURATION ---
def save_config(api_key, model_name):
    """Sauvegarde la configuration dans le fichier .env"""
    with open('.env', 'w') as f:
        f.write(f'GOOGLE_API_KEY={api_key}\n')
        f.write(f'GEMINI_MODEL={model_name}\n')

def load_config():
    """Charge la configuration depuis le fichier .env"""
    load_dotenv()
    return {
        'api_key': os.getenv("GOOGLE_API_KEY", ""),
        'model_name': os.getenv("GEMINI_MODEL", DEFAULT_GEMINI_MODEL)
    }

# --- CONFIGURATION GLOBALE ---
# Charger la configuration
global_config = load_config()
google_api_key = global_config['api_key']

# Configuration de l'API Google Gemini
if google_api_key:
    try:
        genai.configure(api_key=google_api_key)
    except Exception as e:
        st.error(f"Erreur lors de la configuration de l'API Google Gemini: {e}")
        google_api_key = None

def get_available_models():
    """Récupère la liste des modèles disponibles"""
    try:
        models = genai.list_models()
        available_models = []
        for model in models:
            if 'generateContent' in model.supported_generation_methods:
                available_models.append(model.name.split('/')[-1])  # Ne garder que le nom du modèle
        return sorted(list(set(available_models)))  # Éviter les doublons et trier
    except Exception as e:
        st.warning(f"Impossible de récupérer la liste des modèles: {e}")
        # Retourner une liste de modèles par défaut
        return [
            'gemini-1.5-flash',
            'gemini-1.5-flash-8b',
            'gemini-2.0-flash',
            'gemini-2.0-flash-lite',
            'gemini-pro',
            'gemini-1.5-pro',
            'gemini-1.0-pro',
        ]

# --- CONFIGURATION DE L'INTERFACE UTILISATEUR ---
# Ajouter une section de configuration pour les modèles IA
with st.sidebar.expander("🤖 Configuration IA", expanded=True):
    st.markdown("### Configuration des Modèles IA")
    
    # Sélection du fournisseur
    ai_provider = st.selectbox(
        "Fournisseur IA",
        options=["Gemini", "OpenAI", "Modèle local"],
        index=0,
        help="Choisissez le fournisseur IA à utiliser pour la génération de code",
        key="ai_provider_select"
    )
    
    # Configuration Gemini
    if ai_provider == "Gemini":
        st.markdown("#### 🔧 Configuration Gemini")
        gemini_api_key = st.text_input(
            "Clé API Google Gemini",
            value=os.environ.get('GOOGLE_API_KEY', ''),
            type="password",
            help="Obtenez votre clé API sur https://aistudio.google.com/app/apikey",
            key="gemini_api_key_new"
        )
        
        if gemini_api_key:
            try:
                genai.configure(api_key=gemini_api_key)
                available_models = get_available_models()
                _gm_idx = 0
                if available_models:
                    if DEFAULT_GEMINI_MODEL in available_models:
                        _gm_idx = available_models.index(DEFAULT_GEMINI_MODEL)
                    else:
                        _flash_idxs = [i for i, m in enumerate(available_models) if "flash" in m.lower()]
                        _gm_idx = _flash_idxs[0] if _flash_idxs else 0
                gemini_model = st.selectbox(
                    "Modèle Gemini",
                    options=available_models,
                    index=_gm_idx,
                    key="gemini_model_new"
                )
                # Mettre à jour la variable d'environnement
                os.environ['GOOGLE_API_KEY'] = gemini_api_key
            except Exception as e:
                st.error(f"Erreur de connexion à Gemini: {e}")
    
    # Configuration OpenAI
    elif ai_provider == "OpenAI":
        st.markdown("#### 🔧 Configuration OpenAI")
        openai_api_key = st.text_input(
            "Clé API OpenAI",
            value=os.environ.get('OPENAI_API_KEY', ''),
            type="password",
            help="Obtenez votre clé API sur https://platform.openai.com/api-keys",
            key="openai_api_key_new"
        )
        
        if openai_api_key:
            openai_model = st.selectbox(
                "Modèle OpenAI",
                options=["gpt-3.5-turbo", "gpt-4", "gpt-4-turbo-preview"],
                index=0,
                key="openai_model_new"
            )
            # Mettre à jour la clé API
            os.environ['OPENAI_API_KEY'] = openai_api_key
            openai.api_key = openai_api_key
    
    # Configuration modèle local
    elif ai_provider == "Modèle local":
        st.markdown("#### 🔧 Configuration Modèle Local")
        st.info("Le modèle local Gemma sera utilisé si disponible")
        
        # Bouton pour charger le modèle local
        if st.button("Charger le modèle Gemma local", key="load_local_model"):
            with st.spinner("Chargement du modèle local..."):
                load_local_gemma()
    
    # Afficher le statut actuel
    st.markdown("#### 📊 Statut Actuel")
    current_ai = get_ai_handler()
    st.success(f"Fournisseur actuel: {current_ai.provider.upper()}")
    
    # Bouton pour tester la connexion
    if st.button("Tester la connexion IA", key="test_ai_connection"):
        with st.spinner("Test de connexion..."):
            try:
                test_response = current_ai.generate_text("Test simple", provider=current_ai.provider.lower())
                if "Erreur" not in test_response:
                    st.success("✅ Connexion réussie!")
                    st.text(test_response[:100] + "..." if len(test_response) > 100 else test_response)
                else:
                    st.error("❌ Échec de connexion")
            except Exception as e:
                st.error(f"❌ Erreur: {str(e)}")

# Ancienne configuration Gemini (conservée pour compatibilité)
with st.sidebar.expander("⚙️ Configuration API Gemini (Legacy)", expanded=False):
    st.markdown("### Configuration de l'API Gemini")
    
    # Champ pour la clé API
    user_api_key = st.text_input(
        "Clé API Google Gemini",
        value=google_api_key or "",
        type="password",
        help="Obtenez votre clé API sur https://aistudio.google.com/app/apikey",
        key="gemini_api_key_legacy"
    )
    
    # Charger les modèles disponibles si une clé est fournie
    available_models = []
    if user_api_key:
        try:
            genai.configure(api_key=user_api_key)
            available_models = get_available_models()
        except Exception as e:
            st.error(f"Erreur de connexion à l'API: {e}")
    
    # Sélecteur de modèle
    selected_model = st.selectbox(
        "Modèle Gemini",
        options=available_models,
        index=available_models.index(global_config['model_name']) if global_config['model_name'] in available_models and available_models else 0,
        disabled=not available_models,
        key="gemini_model_legacy"
    )
    
    # Bouton de sauvegarde
    if st.button("Enregistrer la configuration", key="save_legacy_config"):
        if user_api_key.strip() and selected_model:
            save_config(user_api_key, selected_model)
            st.success("Configuration enregistrée avec succès !")
            st.experimental_rerun()
        else:
            st.error("Veuillez fournir une clé API valide et sélectionner un modèle")

# Mettre à jour la configuration globale avec les valeurs de l'utilisateur
global_config = load_config()
google_api_key = global_config['api_key']

# --- DÉFINITION DES URLS DES COLLECTES KOBO ---
# Assurez-vous que cette section est présente et non commentée
urls = {
    "Collecte 1: Diagnostic rapide des coopératives": "https://eu.kobotoolbox.org/api/v2/assets/aX8mpWRZaVBomEs3jZ5ULR/export-settings/esp68CyMYKKrHVSPDhwGc2X/data.xlsx",
    "Collecte 2: Unité de démonstration/application": "https://eu.kobotoolbox.org/api/v2/assets/aqKsjwyNuGzbwxWHkUeRjj/export-settings/escMDPrVnDBqMdzxBW8Cn3C/data.xlsx",
    "Collecte 3: Diagnostic des coopératives (PDA 4)": "https://eu.kobotoolbox.org/api/v2/assets/adHBEPncaoH7ShGzCaRZoo/export-settings/esnTMAAYFxPGqwCcwuPpKyA/data.xlsx",
    "Collecte 4: Formulaire d'Identification de la coopérative/Groupement ou centre de formation/d'incubation/ Financement ENDEV/ GIZ": "https://eu.kobotoolbox.org/api/v2/assets/aD66rdQKVqb8TKKfvziUca/export-settings/esnjXxfNZUPqfwhhrNd3gwW/data.xlsx",
    "Collecte 5: Diagnostic des exploitations individuelles des producteurs (PADIAP & CCR-B)": "https://eu.kobotoolbox.org/api/v2/assets/a6in5RdHQR5wZbL6wwy3RQ/export-settings/eso26F8NvjuTN8mbW4dZrQD/data.xlsx",
    "Collecte 6: Dagnostic des exploitations et coopératives (Zone SUD_ CCR-B & RIKOLTO/PARSAD": "https://eu.kobotoolbox.org/api/v2/assets/a2qBRE2k5xXtUVRsUGsarq/export-settings/es4EJrPjbWrxCvdDTLjZazb/data.xlsx",
    "Collecte 7: Point des bénéficiaires enrolés sur la plateforme Kobo dans le cadre de la convention PADAAM & IFRIZ-Bénin (campagne 2025)": "https://eu.kobotoolbox.org/api/v2/assets/aXWgudnb6ATH3nCNsyYfui/export-settings/es9AcQMg5d7vX5frABwV2Lp/data.xlsx",
    "Collecte 8: Point des données CEP dans le cadre de la convention PADAAM & CCR-Bénin (campagne 2025)": "https://eu.kobotoolbox.org/api/v2/assets/acBKznjtZd4NhquvGrXre9/export-settings/esJE4btAyRDkbL9Ng7KCNWi/data.xlsx",
    "Collecte 9: Données des Champs Ecole Paysan (CEP) installée par le CCR-B grâce à l'appui de Delta Mono / Campagne 2025": "https://eu.kobotoolbox.org/api/v2/assets/aEgTKMhkNBEjpf4pfU9wAd/export-settings/esQpgaQinZy4ixZJh2tirn2/data.xlsx",
    "Collecte 10: Suivi individuel des producteurs accompagnés sur Delta Mono": "https://eu.kobotoolbox.org/api/v2/assets/aphJjNZk84MxAQYoi3eqiD/export-settings/es3P5jpfnH7ctP8rvyUTNQ3/data.xlsx",
    "Collecte 1_Ferme SAIN: Questionnaire pour recueillir la perception des acteurs concernant la gestion des emballages de pesticides chimiques de synthèse dans la vallée du Niger": "https://eu.kobotoolbox.org/api/v2/assets/aSHDKg7wpTzynyw5KUKtHG/export-settings/esrv2pyAVVKSBTMSiDG8BFW/data.xlsx",
    "Collecte 2_Ferme SAIN: Questionnaire OPA concernant la gestion des emballages de pesticides chimiques de synthèser": "https://eu.kobotoolbox.org/api/v2/assets/aUds5KDh7wEtco2ru6Rvu4/export-settings/esH3N3FkKBnHJQWQog9Vb4q/data.xlsx",
    "Collecte 11: Consitution de la base de données des producteurs accompagnés sur PRIMA, campagne 2025": "https://eu.kobotoolbox.org/api/v2/assets/aTPCQC66D9AbpBHuUHVpHz/export-settings/esVvJJSUYsBgbpLGg25PfzW/data.xlsx",
    "Collecte 12: Enquête de satisfaction des bénéficiaires du projet Delta Mono avec l'accompagnement du CCR-B et Enabel": "https://eu.kobotoolbox.org/api/v2/assets/asn84D8iA8NJ9NVVAgfUWM/export-settings/esejFfBVAsVbwxbuMgxv3Dk/data.xlsx",
    "Collecte 13: Collecte de données sur DYNAMIQUE DES SYSTEMES AGROALIMENTAIRES TERRITORIALISES (SAT) AU BENIN : CAS DE LA FILIERE RIZ DANS LA BASSE VALLEE DE L'OUEME": "https://eu.kobotoolbox.org/api/v2/assets/aB9HbmdsJHdimQp2UooUoS/export-settings/esG4p4uHu48e3Y6eFYpEqH2/data.xlsx",
    "Collecte 14: Collecte générale pour le suivi des indicateurs de la convention CCR-B & RIKOLTO/ Projet PARSAD D'EnabelBenin": "https://eu.kobotoolbox.org/api/v2/assets/aZXXS77TK2u6iCyzgQVhPB/export-settings/esCoK33r5v8nEt7omWmeZsv/data.xlsx",
    "Collecte 15: Formulaire de collecte de données sur les CEP_ PDA4 _campagne 2025": "https://eu.kobotoolbox.org/api/v2/assets/aAYhthxB4DCSgH47ZD2kWk/export-settings/esTnnwytg3gdgaBkoergAfU/data.xlsx",
    "Collecte 16: Point des données CEP; Zone SUD, campagne 2025": "https://eu.kobotoolbox.org/api/v2/assets/awEKERo4GkxN4GMzUFhyov/export-settings/esThEAfzHPzm8YoqaevXZA4/data.xlsx",
    "Collecte 17: Formulaire UPE – Étude des Besoins/ Financement ENDEV/ GIZ (2e Collecte_ Septembre 2025)": "https://eu.kobotoolbox.org/api/v2/assets/aRvpuwmCjZJuj8KXEk8dV5/export-settings/esDnpta82Ax8BsLmy6rb6xt/data.xlsx",
    "Collecte 18: Collecte de données sur les carrés de rendements sous le projet PADIAP, campagne 2025": "https://eu.kobotoolbox.org/api/v2/assets/avnKKrg8v9CaD3rmvM2uvV/export-settings/esDue93XsnMRWXm644NwZxZ/data.xlsx",
    "Collecte 19: Evaluation TAPE_2025": "https://eu.kobotoolbox.org/api/v2/assets/a7ULmLBAwLFuRoEN6htuaN/export-settings/esvjmXCkLDBCWcn8ZuWZ3JL/data.xlsx",
    "Collecte 20: Collecte des données pour la mise à jour des indicateurs du cadre de résultat du projet PADIAP avec CCR-B": "https://eu.kobotoolbox.org/api/v2/assets/aLCq2C8iWZkGt6hBWPUvyA/export-settings/esgWEDvWDKE2sG8nz3N6j2r/data.xlsx",
    "Collecte 21: Collecte de données sur la baseline du projet FIPA mise en oeuvre avec RIKOLTO et CCR-B, juin 2026": "https://kf.kobotoolbox.org/private-media/AnonymousUser/exports/arSNAHn25xfwT34PZoo7iE/Enqu%C3%AAte_Baseline_FIPA__CCR-B__Rikolto_B%C3%A9nin_-_all_versions_-_default_-_2026-07-20-16-36-07.xlsx",
}

# --- DÉFINITION DES CHEMINS DE LOGOS ---
# Assurez-vous que ces fichiers existent dans le même répertoire que le script ou ajustez les chemins
logo_base_path = "." # Chemin relatif au script
logo_paths = {
    "delta": os.path.join(logo_base_path, "PP Delta Mono.png"),
    "aagac": os.path.join(logo_base_path, "logo_aagac.png"),
    "ifriz": os.path.join(logo_base_path, "PP IFRIZ.png"),
    "ccrb": os.path.join(logo_base_path, "PP CCRB.png"),
    "ccrb_padiap": os.path.join(logo_base_path, "PP CCRB & PADIAP.png"),
}

# --- DÉFINITION DES FONCTIONS UTILS POUR L'IA ---
safe_px_funcs = [
    'bar', 'line', 'scatter', 'histogram', 'box', 'pie', 'area',
    'scatter_3d', 'line_3d', 'surface', 'contour', 'density_heatmap',
    'density_contour', 'violin', 'strip', 'boxen', '더라구요', 'candlestick',
    'ohlc', 'timeline', 'treemap', 'sunburst', 'icicle', 'funnel',
    'funnelarea', 'indicator', 'map', 'scatter_geo', 'scatter_mapbox',
    'choropleth', 'choropleth_mapbox', 'density_mapbox', 'line_geo',
    'line_mapbox', 'mapbox', 'scatter_mapbox', 'bar_polar', 'line_polar',
    'scatter_polar', 'pie_polar', 'sunburst_polar', 'icicle_polar',
    'scatter_matrix', 'parallel_coordinates', 'parallel_categories'
]

# --- FONCTION DE TÉLÉCHARGEMENT DE TOUTES LES COLLECTES ---
def download_selected_collections_excel(selected_collections):
    """
    Télécharge les collectes sélectionnées de KoboToolbox dans un fichier Excel unique
    avec chaque collecte dans une feuille séparée.
    
    Args:
        selected_collections (list): Liste des noms des collectes à télécharger
    """
    if not selected_collections:
        st.warning("⚠️ Veuillez sélectionner au moins une collecte à télécharger.")
        return None, None
        
    try:
        # Créer un nouveau classeur Excel
        wb = Workbook()
        
        # Supprimer la feuille par défaut
        wb.remove(wb.active)
        
        # Dictionnaire pour stocker les données de chaque collecte
        collections_data = {}
        failed_collections = []
        
        # Afficher un indicateur de progression
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        total_collections = len(selected_collections)
        
        for i, collection_name in enumerate(selected_collections):
            if collection_name not in urls:
                failed_collections.append(f"{collection_name} (URL non trouvée)")
                continue
                
            url = urls[collection_name]
            
            try:
                status_text.text(f"📥 Chargement de : {collection_name}")
                progress_bar.progress((i + 1) / total_collections)
                
                # Charger les données de cette collecte
                df = load_data(url)
                
                if df is not None and not df.empty:
                    # Nettoyer le nom de la feuille (Excel limite à 31 caractères)
                    sheet_name = collection_name[:31].replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('[', '_').replace(']', '_')
                    
                    # Créer une nouvelle feuille
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # Ajouter les données du DataFrame à la feuille
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    
                    # Ajuster la largeur des colonnes
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Limiter à 50 caractères
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    collections_data[collection_name] = {
                        'rows': len(df),
                        'columns': len(df.columns),
                        'sheet_name': sheet_name
                    }
                else:
                    failed_collections.append(f"{collection_name} (Données vides)")
                    
            except Exception as e:
                failed_collections.append(f"{collection_name} (Erreur: {str(e)})")
                continue
        
        # Créer une feuille de résumé si on a des données
        if collections_data:
            ws_summary = wb.create_sheet(title="Résumé", index=0)
            
            # En-têtes du résumé
            ws_summary.append(["Nom de la Collecte", "Nom de la Feuille", "Nombre de Lignes", "Nombre de Colonnes"])
            
            # Données du résumé
            for collection_name, data in collections_data.items():
                ws_summary.append([
                    collection_name,
                    data['sheet_name'],
                    data['rows'],
                    data['columns']
                ])
            
            # Ajuster la largeur des colonnes du résumé
            for column in ws_summary.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_summary.column_dimensions[column_letter].width = adjusted_width
            
            # Style pour les en-têtes du résumé
            from openpyxl.styles import Font
            for cell in ws_summary[1]:
                cell.font = Font(bold=True)
        
        # Sauvegarder le fichier dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Générer le nom du fichier
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Collectes_Selectionnees_{timestamp}.xlsx"
        
        # Afficher un résumé
        st.success(f"✅ Fichier Excel généré avec succès !")
        st.info(f"📊 **Résumé :** {len(collections_data)} collectes chargées, {len(failed_collections)} échecs")
        
        if collections_data:
            st.markdown("**📋 Collectes chargées :**")
            for collection_name, data in collections_data.items():
                st.markdown(f"• **{collection_name}** : {data['rows']} lignes, {data['columns']} colonnes")
        
        if failed_collections:
            st.warning("**⚠️ Collectes en échec :**")
            for failed in failed_collections:
                st.markdown(f"• {failed}")
        
        return buffer.getvalue(), filename
        
    except Exception as e:
        st.error(f"❌ Erreur lors de la génération du fichier Excel : {str(e)}")
        return None, None

def download_all_collections_excel():
    """
    Télécharge toutes les collectes KoboToolbox dans un fichier Excel unique
    avec chaque collecte dans une feuille séparée.
    """
    return download_selected_collections_excel(list(urls.keys()))
    try:
        # Créer un nouveau classeur Excel
        wb = Workbook()
        
        # Supprimer la feuille par défaut
        wb.remove(wb.active)
        
        # Dictionnaire pour stocker les données de chaque collecte
        collections_data = {}
        failed_collections = []
        
        # Afficher un indicateur de progression
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        total_collections = len(urls)
        
        for i, (collection_name, url) in enumerate(urls.items()):
            try:
                status_text.text(f"📥 Chargement de : {collection_name}")
                progress_bar.progress((i + 1) / total_collections)
                
                # Charger les données de cette collecte
                df = load_data(url)
                
                if df is not None and not df.empty:
                    # Nettoyer le nom de la feuille (Excel limite à 31 caractères)
                    sheet_name = collection_name[:31].replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('[', '_').replace(']', '_')
                    
                    # Créer une nouvelle feuille
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # Ajouter les données du DataFrame à la feuille
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    
                    # Ajuster la largeur des colonnes
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Limiter à 50 caractères
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    collections_data[collection_name] = {
                        'rows': len(df),
                        'columns': len(df.columns),
                        'sheet_name': sheet_name
                    }
                else:
                    failed_collections.append(collection_name)
                    
            except Exception as e:
                failed_collections.append(f"{collection_name} (Erreur: {str(e)})")
                continue
        
        # Créer une feuille de résumé
        if collections_data:
            ws_summary = wb.create_sheet(title="Résumé", index=0)
            
            # En-têtes du résumé
            ws_summary.append(["Nom de la Collecte", "Nom de la Feuille", "Nombre de Lignes", "Nombre de Colonnes"])
            
            # Données du résumé
            for collection_name, data in collections_data.items():
                ws_summary.append([
                    collection_name,
                    data['sheet_name'],
                    data['rows'],
                    data['columns']
                ])
            
            # Ajuster la largeur des colonnes du résumé
            for column in ws_summary.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_summary.column_dimensions[column_letter].width = adjusted_width
            
            # Style pour les en-têtes du résumé
            from openpyxl.styles import Font
            for cell in ws_summary[1]:
                cell.font = Font(bold=True)
        
        # Sauvegarder le fichier dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Générer le nom du fichier
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Toutes_Collectes_CCRB_{timestamp}.xlsx"
        
        # Afficher un résumé
        st.success(f"✅ Fichier Excel généré avec succès !")
        st.info(f"📊 **Résumé :** {len(collections_data)} collectes chargées, {len(failed_collections)} échecs")
        
        if collections_data:
            st.markdown("**📋 Collectes chargées :**")
            for collection_name, data in collections_data.items():
                st.markdown(f"• **{collection_name}** : {data['rows']} lignes, {data['columns']} colonnes")
        
        if failed_collections:
            st.warning("**⚠️ Collectes en échec :**")
            for failed in failed_collections:
                st.markdown(f"• {failed}")
        
        return buffer.getvalue(), filename
        
    except Exception as e:
        st.error(f"❌ Erreur lors de la génération du fichier Excel : {str(e)}")
        return None, None

# --- FONCTIONS UTILITAIRES ---
def generer_titre_automatique(variables, prefixe="Analyse de"):
    """Génère un titre basé sur une liste de variables."""
    if not variables or all(not v for v in variables): # Vérifie si la liste est vide ou contient uniquement des chaînes vides
        return f" (aucune variable sélectionnée)"

    # Filtrer les variables vides avant de joindre
    valid_variables = [v for v in variables if v and v.strip()]
    if not valid_variables:
        return f" (aucune variable valide sélectionnée)"

    titre = f" " + ", ".join(valid_variables)
    return titre

def sanitize_column_name_for_sql(col_name, existing_sanitized_names):
    """
    Nettoie un nom de colonne pour être utilisé dans une requête SQL.
    Évite les doublons et assure un format valide.
    """
    # Remplacer les caractères non alphanumériques par des underscores
    sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', str(col_name))
    # Remplacer les underscores multiples par un seul
    sanitized = re.sub(r'__+', '_', sanitized)
    # Supprimer les underscores au début ou à la fin
    sanitized = sanitized.strip('_')

    # Assurer que le nom commence par une lettre ou un underscore si nécessaire (SQL standard)
    if not sanitized or sanitized[0].isdigit():
        sanitized = 'col_' + sanitized

    # Convertir en minuscules pour la cohérence SQL
    sanitized = sanitized.lower()

    # Gérer les doublons
    original_sanitized = sanitized
    counter = 1
    while sanitized in existing_sanitized_names:
        sanitized = f"{original_sanitized}_{counter}"
        counter += 1

    if sanitized: # Ajouter seulement si le nom nettoyé n'est pas vide
        existing_sanitized_names.add(sanitized)
    else: # Si le nettoyage a résulté en une chaîne vide, utiliser un nom par défaut
        sanitized = f"col_{len(existing_sanitized_names)}"
        existing_sanitized_names.add(sanitized)

    return sanitized

def get_insight_card(analysis):
    """
    Génère les informations (icone, couleur, titre, résumé) pour une carte d'aperçu
    basée sur le type d'analyse et son résultat.
    """
    params = analysis.get('executed_params', {})
    titre = params.get('title')
    result = analysis.get('result')
    analysis_type = analysis.get('type')

    # Générer un titre par défaut si non fourni ou vide
    if not titre or not titre.strip():
        if analysis_type == 'aggregated_table':
            titre = generer_titre_automatique(params.get('group_by_columns', []), prefixe="Tableau agrégé de")
        elif analysis_type == 'descriptive_stats':
            titre = generer_titre_automatique(params.get('selected_columns', []), prefixe="Statistiques descriptives de")
        elif analysis_type == 'graph':
            # Essayer de construire un titre plus informatif pour les graphiques
            chart_type = params.get('chart_type', 'Graphique')
            x_col = params.get('x_column') or params.get('names_column')
            y_col = params.get('y_column') or params.get('values_column')

            if x_col and y_col: titre = f"Graphique : {x_col} vs {y_col}"
            elif x_col and chart_type == 'histogram': titre = f"Histogramme de {x_col}"
            elif x_col and chart_type == 'box': titre = f"Boîte à moustaches de {x_col}"
            elif x_col and chart_type == 'pie': titre = f"Répartition de {x_col}"
            else: titre = f"Graphique {chart_type}"

        elif analysis_type == 'sql_query': titre = params.get('title') or "Résultat : Requête SQL"
        elif analysis_type == 'python_script': titre = params.get('title') or "Résultat : Script Python"
        else: titre = "Analyse"

    # Initialiser les variables de retour
    icon, color, resume = '🔎', '#888888', "Résultat non formaté."

    # Formater le résumé basé sur le type d'analyse et le résultat
    if result is not None:
        if analysis_type == 'aggregated_table' and isinstance(result, pd.DataFrame):
            icon = '📊'; color = '#4F8DFD'
            if not result.empty:
                # Insight général concis au lieu des insights détaillés
                total_rows = result.shape[0]
                total_cols = result.shape[1]
                
                # Identifier les colonnes numériques pour un résumé simple
                numeric_cols = result.select_dtypes(include=[np.number]).columns.tolist()
                if numeric_cols:
                    # Prendre la première colonne numérique pour un exemple
                    first_numeric = numeric_cols[0]
                    if first_numeric in result.columns:
                        total_sum = result[first_numeric].sum()
                        max_val = result[first_numeric].max()
                        avg_val = result[first_numeric].mean()
                        resume = f"{total_rows} lignes, {total_cols} colonnes. <strong>💡 Insight général :</strong> Données agrégées avec {len(numeric_cols)} colonne(s) numérique(s). Valeur totale principale : {total_sum:.0f} (moyenne: {avg_val:.1f}, max: {max_val:.0f})"
                    else:
                        resume = f"{total_rows} lignes, {total_cols} colonnes. <strong>💡 Insight général :</strong> Tableau d'agrégation des données collectées."
                else:
                    resume = f"{total_rows} lignes, {total_cols} colonnes. <strong>💡 Insight général :</strong> Tableau d'agrégation des données collectées."
            else: resume = "Le tableau agrégé est vide."

        elif analysis_type == 'descriptive_stats' and isinstance(result, pd.DataFrame):
            icon = '📈'; color = '#00B86B'
            if not result.empty and 'mean' in result.columns:
                means = result.loc['mean']
                if not means.empty:
                    means_str_parts = [f"<span class='badge'>{idx}: {val:.2f}</span>" for idx, val in means.head(3).items()]
                    resume = f"Moyenne(s) : {', '.join(means_str_parts)}"
                else: resume = "Aucune statistique calculée."
            elif not result.empty:
                means = result.iloc[0]
                if not pd.isna(means): resume = f"Statistique clé : <span class='badge'>{means:.2f}</span>"
                else: resume = "Aucune statistique calculée."
            else: resume = "Le tableau de statistiques est vide."

        elif analysis_type == 'graph':
            icon = '📉'; color = '#FFB200'
            x = params.get('x_column') or params.get('names_column')
            y = params.get('y_column') or params.get('values_column')
            resume = f"Type: <span class='badge'>{params.get('chart_type', 'Inconnu')}</span><br>"
            if x: resume += f"X: <b>{x}</b> "
            if y: resume += f"| Y: <b>{y}</b>"
            if not x and not y: resume = "Graphique généré."

        elif analysis_type == 'sql_query' and isinstance(result, pd.DataFrame):
            icon = '🗃️'; color = '#7C3AED'
            if not result.empty:
                numerical_cols_result = result.select_dtypes(include=np.number).columns.tolist()
                if numerical_cols_result:
                    summary_parts = []
                    for col in numerical_cols_result:
                        try:
                            col_sum = result[col].sum()
                            summary_parts.append(f"<b>{col}</b>: <span class='badge'>{col_sum:.2f}</span>")
                        except Exception: pass
                    if summary_parts: resume = f"{result.shape[0]} lignes. Somme(s) : {', '.join(summary_parts)}"
                    else: resume = f"{result.shape[0]} lignes, {result.shape[1]} colonnes."
                else: resume = f"{result.shape[0]} lignes, {result.shape[1]} colonnes."
            else: resume = "La requête SQL n'a retourné aucun résultat."

        elif analysis_type == 'python_script':
            icon = '🐍'; color = '#FF6B35'
            if hasattr(result, 'to_plotly_json'): resume = f"Graphique interactif généré par Python"
            elif isinstance(result, pd.DataFrame): resume = f"Tableau {result.shape[0]}×{result.shape[1]} généré par Python"
            elif hasattr(result, 'figure'): resume = f"Graphique Matplotlib généré par Python"
            else: resume = f"Résultat Python : {str(result)[:50]}..."

        # Gestion des analyses intelligentes
        elif analysis_type.startswith('intelligent_'):
            if 'demographic' in analysis_type:
                icon = '👥'; color = '#4ECDC4'
                resume = f"Analyse démographique intelligente - {params.get('variable', 'Variable')}"
            elif 'satisfaction' in analysis_type:
                icon = '😊'; color = '#FFE66D'
                resume = f"Analyse de satisfaction intelligente - {params.get('variable', 'Variable')}"
            elif 'geographic' in analysis_type:
                icon = '🗺️'; color = '#95E1D3'
                resume = f"Analyse géographique intelligente - {params.get('variable', 'Variable')}"
            elif 'correlation' in analysis_type:
                icon = '🔗'; color = '#F38181'
                resume = f"Analyse des corrélations intelligente - {len(params.get('variables', []))} variables"
            elif 'general' in analysis_type:
                icon = '📈'; color = '#A8E6CF'
                resume = f"Analyse générale intelligente - {params.get('variable', 'Variable')}"
            elif 'overview' in analysis_type:
                icon = '📊'; color = '#74B9FF'
                resume = f"Vue d'ensemble - {params.get('variable', 'Variable')}"
            else:
                icon = '🧠'; color = '#6C5CE7'
                resume = f"Analyse intelligente - {params.get('title', 'Variable')}"

    return icon, color, titre, resume

# --- GESTION DES LOGOS ---
logo_base_path = "." # Répertoire courant par défaut
logo_paths = {
    "delta": os.path.join(logo_base_path, "PP Delta Mono.png"),
    "aagac": os.path.join(logo_base_path, "logo_aagac.png"),
    "ifriz": os.path.join(logo_base_path, "PP IFRIZ.png"),
    "ccrb": os.path.join(logo_base_path, "PP CCRB.png"),
    "rikolto": os.path.join(logo_base_path, "logo_rikolto.png"),
    "ccrb_padiap": os.path.join(logo_base_path, "PP CCRB & PADIAP.png"),
}

def load_logo_base64(logo_path):
    """Charge un logo depuis un chemin et le retourne en base64."""
    if os.path.exists(logo_path):
        try:
            with open(logo_path, "rb") as f: return base64.b64encode(f.read()).decode('utf-8')
        except Exception as e:
            st.sidebar.warning(f"Erreur chargement logo '{os.path.basename(logo_path)}': {str(e)}")
            return ""
    else: return ""

def get_logo_and_org_info(selected_header_key):
    """Détermine le logo, nom d'organisation et largeur."""
    org_name = "CCR-B"; logo_key = "ccrb"; logo_width = 180
    if not selected_header_key: return logo_paths.get(logo_key, ""), org_name, logo_width

    if 'Collecte 10' in selected_header_key or 'Collecte 12' in selected_header_key or 'Collecte 19' in selected_header_key: # Delta Mono
        logo_key = "delta"; org_name = "CCR-B & Delta Mono"; logo_width = 250
    elif 'Collecte 1_Ferme SAIN' in selected_header_key or 'Collecte 2_Ferme SAIN' in selected_header_key: # Ferme SAIN / AAGAC
        logo_key = "aagac"; org_name = "AAGAC & Partenaires"; logo_width = 150
    elif 'Collecte 5' in selected_header_key or 'Collecte 18' in selected_header_key or 'Collecte 20' in selected_header_key: # PADIAP
        logo_key = "ccrb_padiap"; org_name = "CCR-B & PADIAP"; logo_width = 200
    elif 'Collecte 13' in selected_header_key: # SAT - Systèmes Agroalimentaires Territorialisés
        logo_key = "ccrb"; org_name = "CCR-B & Systèmes Agroalimentaires Territorialisés (SAT)"; logo_width = 200
    elif 'Collecte 14' in selected_header_key: # RIKOLTO/PARSAD - EnabelBenin
        logo_key = "ccrb"; org_name = "CCR-B & RIKOLTO/PARSAD - EnabelBenin"; logo_width = 200
    elif 'Collecte 21' in selected_header_key: # FIPA Baseline
        logo_key = "ccrb"; org_name = "CCR-B & RIKOLTO - Baseline FIPA"; logo_width = 200
    elif 'PADAAM' in selected_header_key: # IFRIZ-Bénin pour les collectes PADAAM
        logo_key = "ifriz"; org_name = "IFRIZ-Bénin"; logo_width = 180
    elif any(f'Collecte ' in selected_header_key for i in [1, 5]): # CCRB Générique
        logo_key = "ccrb"; org_name = "CCR-B"
    elif any(f'Collecte ' in selected_header_key for i in [7, 8]): # IFRIZ-Bénin (ancienne logique)
        logo_key = "ifriz"; org_name = "IFRIZ-Bénin"; logo_width = 180

    return logo_paths.get(logo_key, ""), org_name, logo_width

# --- FONCTION DE GÉNÉRATION DE RAPPORT HTML ---
def find_first_matching_column(df, keywords):
    """Trouve la première colonne dans le dataframe qui correspond à une liste de mots-clés."""
    for col in df.columns:
        for keyword in keywords:
            if keyword in col.lower():
                return col
    return None

def safe_string_contains(series, value, case=False, na=False):
    """
    Fonction sécurisée pour rechercher une valeur dans une série pandas.
    Gère automatiquement les types de données (string vs numérique).
    """
    try:
        if series.dtype == 'object':
            # Pour les colonnes de type string/object
            return series.astype(str).str.contains(value, case=case, na=na)
        else:
            # Pour les colonnes numériques, chercher des valeurs positives
            positive_numeric_values = [1, 2, 3, 4, 5]
            return series.isin(positive_numeric_values)
    except Exception as e:
        # En cas d'erreur, retourner une série de False
        st.warning(f"Erreur lors de la recherche dans la colonne: {e}")
        return pd.Series([False] * len(series), index=series.index)

def generate_table_insights(df, analysis_params):
    """
    Génère des insights automatiques pour un tableau agrégé
    """
    if not isinstance(df, pd.DataFrame) or df.empty:
        return "Aucun insight disponible pour ce tableau."
    
    insights = []
    
    try:
        # Analyser la structure du tableau
        total_rows = len(df)
        total_cols = len(df.columns)
        
        # Identifier les colonnes clés
        group_by_cols = analysis_params.get('group_by', [])
        value_cols = analysis_params.get('columns', [])
        
        # Insight 1: Vue d'ensemble
        insights.append(f"📊 **Vue d'ensemble** : {total_rows} lignes, {total_cols} colonnes")
        
        # Insight 2: Analyse des groupes
        if group_by_cols:
            for col in group_by_cols:
                if col in df.columns:
                    unique_values = df[col].nunique()
                    if unique_values <= 10:  # Si peu de valeurs uniques, les lister
                        top_values = df[col].value_counts().head(3)
                        top_str = ", ".join([f"{k} ({v})" for k, v in top_values.items()])
                        insights.append(f"🏷️ **{col}** : {unique_values} valeurs uniques. Top 3 : {top_str}")
                    else:
                        insights.append(f"🏷️ **{col}** : {unique_values} valeurs uniques")
        
        # Insight 3: Analyse des valeurs numériques
        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        if numeric_cols:
            for col in numeric_cols:
                if col in df.columns:
                    total = df[col].sum()
                    max_val = df[col].max()
                    min_val = df[col].min()
                    avg_val = df[col].mean()
                    
                    # Identifier les records
                    max_row = df.loc[df[col] == max_val]
                    min_row = df.loc[df[col] == min_val]
                    
                    insights.append(f"🔢 **{col}** : Total={total:.0f}, Moyenne={avg_val:.1f}, Min={min_val:.0f}, Max={max_val:.0f}")
                    
                    # Insight sur les records
                    if not max_row.empty:
                        record_info = []
                        for group_col in group_by_cols:
                            if group_col in max_row.columns:
                                record_info.append(f"{group_col}={max_row.iloc[0][group_col]}")
                        if record_info:
                            insights.append(f"🏆 **Record {col}** : {' | '.join(record_info)} = {max_val:.0f}")
        
        # Insight 4: Patterns et distributions
        if len(df) > 1:
            # Identifier les patterns de distribution
            if numeric_cols:
                for col in numeric_cols:
                    if col in df.columns:
                        # Calculer la variance pour identifier la dispersion
                        variance = df[col].var()
                        if variance > 0:
                            cv = (df[col].std() / df[col].mean()) * 100 if df[col].mean() != 0 else 0
                            if cv > 50:
                                insights.append(f"📈 **{col}** : Forte dispersion (CV={cv:.1f}%)")
                            elif cv > 20:
                                insights.append(f"📈 **{col}** : Dispersion modérée (CV={cv:.1f}%)")
                            else:
                                insights.append(f"📈 **{col}** : Faible dispersion (CV={cv:.1f}%)")
        
        # Insight 5: Analyse comparative si plusieurs groupes
        if len(group_by_cols) >= 2:
            insights.append(f"🔍 **Analyse comparative** : {len(group_by_cols)} dimensions d'analyse")
        
        # Insight 6: Résumé des données manquantes
        missing_data = df.isnull().sum().sum()
        if missing_data > 0:
            insights.append(f"⚠️ **Données manquantes** : {missing_data} valeurs manquantes détectées")
        
        # Insight 7: Recommandations
        if total_rows > 50:
            insights.append("💡 **Recommandation** : Volume de données élevé, considérez des visualisations pour une meilleure lisibilité")
        elif total_rows < 5:
            insights.append("💡 **Recommandation** : Volume de données faible, les tendances peuvent ne pas être représentatives")
        
        return "\n\n".join(insights)
        
    except Exception as e:
        return f"Erreur lors de la génération des insights : {str(e)}"

def generate_report_html_content():
    """Génère le contenu HTML du rapport complet."""
    try:
        # Initialiser la clé API au début de la fonction
        current_api_key = st.session_state.get('user_api_key') or google_api_key

        # Initialiser les styles ReportLab pour le PDF
        try:
            from reportlab.lib.styles import getSampleStyleSheet
            styles = getSampleStyleSheet()
        except ImportError:
            styles = None

        selected_header_key = st.session_state.get('current_collection_header', '')
        logo_path, org_name, logo_width = get_logo_and_org_info(selected_header_key)
        
        if any(f'Collecte {i}' in selected_header_key for i in [14, 15, 16]):
            logo_ccrb_b64 = load_logo_base64(logo_paths.get("ccrb", ""))
            logo_rikolto_b64 = load_logo_base64(logo_paths.get("rikolto", ""))
            
            logo_html = f"""<div style='display:flex; justify-content:space-between; align-items:center; margin-bottom:10px;'>
            <img src='data:image/png;base64,{logo_ccrb_b64}' alt='Logo CCR-B' width='150' style='margin-bottom:8px;'/>
            <h2 style='margin:0; color:#007bff; text-align:center;'>CCR-B & RIKOLTO</h2>
            <img src='data:image/png;base64,{logo_rikolto_b64}' alt='Logo Rikolto' width='150' style='margin-bottom:8px;'/>
            </div>"""
        else:
            logo_b64 = load_logo_base64(logo_path)
            logo_html = f"""<div style='text-align:center; margin-bottom:10px;'>
            <img src='data:image/png;base64,{logo_b64}' alt='Logo {org_name}' width='{logo_width}' style='margin-bottom:8px;'/>
            <h2 style='margin:0; color:#007bff;'>{org_name}</h2>
            </div>""" if logo_b64 else f"""<div style='text-align:center; margin-bottom:10px;'>
            <span style='color:red;'>Logo non disponible</span>
            <h2 style='margin:0; color:#007bff;'>{org_name}</h2>
            </div>"""

        report_title_from_config = st.session_state.get('report_title', st.session_state.get('current_collection_header', 'Collecte'))
        html_content = f"""
        <!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>Rapport d'Analyse - {report_title_from_config}</title>
        <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap" rel="stylesheet">
        <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
        <script>window.PlotlyConfig = {{MathJaxConfig: 'local'}}</script>
        <script src="https://cdnjs.cloudflare.com/ajax/libs/iframe-resizer/4.3.2/iframeResizer.contentWindow.min.js"></script>
        <style>
            body {{ font-family: 'Inter', sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 20px; background-color: #f4f6f8; }}
            .container {{ max-width: 1000px; margin: 20px auto; background: #fff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 20px rgba(0,0,0,0.05); }}
            h1, h2, h3, h4 {{ color: #2c3e50; margin-top: 1.5em; margin-bottom: 0.5em; }} h1 {{ text-align: center; color: #007bff; }}
            h2 {{ border-bottom: 2px solid #eee; padding-bottom: 5px; margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; page-break-inside: auto; }}
            th, td {{ border: 1px solid #ddd; padding: 10px; text-align: left; page-break-inside: avoid; }}
            th {{ background-color: #f9f9f9; }}
            thead {{ display: table-header-group; }}
            tbody {{ display: table-row-group; }}
            tr {{ page-break-inside: avoid; page-break-after: auto; }}
            .insight-cards-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 1.5rem; margin-top: 1.5rem; margin-bottom: 2rem; }}
            .insight-card {{ background: #fff; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.08); padding: 1.2rem 1.5rem; border-left: 8px solid #4F8DFD; display: flex; flex-direction: column; }}
            .insight-card .icon {{ font-size: 2.2rem; margin-bottom: 0.2rem; line-height: 1; }}
            .insight-card .title {{ font-weight: 600; font-size: 1.1rem; margin-bottom: 0.3rem; color: #2c3e50; }}
            .insight-card .resume {{ font-size: 1rem; color: #555; }}
            .insight-card .badge {{ display: inline-block; background: #e9ecef; color: #495057; border-radius: 6px; padding: 0.1rem 0.6rem; font-size: 0.9em; margin: 0 0.2em 0.2em 0; white-space: nowrap; }}
            .chat-history {{ background: #f0f2f5; border-radius: 8px; padding: 15px; margin-top: 20px; }}
            .chat-message {{ margin-bottom: 15px; padding: 10px 15px; border-radius: 10px; }}
            .chat-message.user {{ background-color: #e0f2f7; text-align: right; margin-left: 20%; }}
            .chat-message.assistant {{ background-color: #f7f7f7; text-align: left; margin-right: 20%; }}
            .chat-message strong {{ font-weight: 600; }}
            .footer {{ text-align: center; margin-top: 50px; padding-top: 20px; border-top: 1px solid #eee; color: #777; font-size: 0.9em; }}
            @media print {{
                table {{ page-break-inside: auto; }}
                tr {{ page-break-inside: avoid; page-break-after: auto; }}
                thead {{ display: table-header-group; }}
                tbody {{ display: table-row-group; }}
            }}
        </style></head><body><div class="container">{logo_html}<h1>Rapport d'Analyse des Collectes KoboToolbox {f"- {st.session_state.get('current_collection_header')}" if st.session_state.get('current_collection_header') else ""}</h1><p>Généré le: {datetime.now().strftime("%d/%m/%Y à %H:%M:%S")}</p><h2>Résumé des Insights</h2><div class="insight-cards-grid">
        """

        total_forms_submitted = 0
        df_current = st.session_state.get('df_current')
        if df_current is not None and not df_current.empty:
            potential_id_cols = ['_index', 'index', 'ID', 'id', 'formhub/uuid']
            id_col_found = None
            for col in potential_id_cols:
                if col in df_current.columns and df_current[col].nunique() > 0:
                    id_col_found = col; break
            total_forms_submitted = df_current[id_col_found].nunique() if id_col_found else len(df_current)
            html_content += f"<div class='insight-card' style='border-left-color:#28a745;'><div class='icon'>📊</div><div class='title'>Formulaires Soumis</div><div class='resume'>Total: <strong>{total_forms_submitted}</strong></div></div>"

        analyses = st.session_state.get('analyses', [])
        if analyses:
            for analysis in analyses:
                if analysis.get('result') is not None:
                    icon, color, titre, resume = get_insight_card(analysis)
                    html_content += f"<div class='insight-card' style='border-left-color:{color};'><div class='icon'>{icon}</div><div class='title'>{titre}</div><div class='resume'>{resume}</div></div>"
        else:
            html_content += "<p style='grid-column: 1 / -1; text-align: center; color: #888;'>Aucune analyse n'a été effectuée pour générer des insights.</p>"

        html_content += "</div><h2>Détails des Analyses</h2>"

        if analyses:
            for i, analysis in enumerate(analyses):
                if analysis.get('result') is not None:
                    params = analysis.get('executed_params', {})
                    _, _, titre, _ = get_insight_card(analysis)

                    html_content += f"<h3>{i+1}. {titre} ({analysis['type'].replace('_', ' ').title()})</h3>"
                    ai_interpretation = ""
                    # Vérifier s'il y a une interprétation modifiée par l'utilisateur
                    modified_interpretation = analysis.get('modified_interpretation', '')
                    
                    if modified_interpretation:
                        # Utiliser l'interprétation modifiée par l'utilisateur
                        ai_interpretation = f"<div style='background-color: #e6f7ff; border-left: 5px solid #66b3ff; padding: 10px 15px; margin-top: 15px; border-radius: 5px; font-size: 0.95em;'><b>Interprétation IA (modifiée) :</b> {modified_interpretation}</div>"
                    elif current_api_key and (analysis['type'] == 'aggregated_table' or analysis['type'] == 'descriptive_stats') and analysis['result'] is not None:
                        # Initialiser ai_interpretation à une chaîne vide par défaut
                        ai_interpretation = ""
                        
                        try:
                            # Essayer de configurer l'API avec la clé fournie
                            genai.configure(api_key=current_api_key)
                            
                            # Vérifier d'abord si le modèle est disponible
                            try:
                                available_models = genai.list_models()
                                model_name = 'gemini-1.5-flash'  # Modèle plus stable
                                
                                # Vérifier si le modèle est disponible
                                model_available = any(model.name.endswith(model_name) for model in available_models)
                                
                                if model_available:
                                    model_chat_report = genai.GenerativeModel(model_name)
                                    interpretation_prompt = ""
                                    
                                    if analysis['type'] == 'aggregated_table' and isinstance(analysis['result'], pd.DataFrame):
                                        # Convertir les colonnes de rendement de kg/ha en tonnes/ha (diviser par 1000)
                                        result_df = analysis['result'].copy()
                                        rendement_cols = [col for col in result_df.columns if 'rendement' in col.lower() or 'yield' in col.lower()]
                                        for col in rendement_cols:
                                            if pd.api.types.is_numeric_dtype(result_df[col]):
                                                result_df[col] = result_df[col] / 1000  # Conversion kg/ha en tonnes/ha
                                                
                                        df_summary_head = result_df.head(5).to_html(index=False, classes='dataframe', border=0)
                                        
                                        # Mettre à jour le résultat dans l'analyse pour qu'il soit utilisé plus tard
                                        analysis['result'] = result_df
                                        
                                        interpretation_prompt = f"""Je dispose d'un tableau de données agrégées. Voici les premières lignes:

{df_summary_head}

Paramètres utilisés : {params}.
Interprète brièvement ce tableau de manière concise et professionnelle. Commence directement par l'interprétation."""
                                    elif analysis['type'] == 'descriptive_stats' and isinstance(analysis['result'], pd.DataFrame):
                                        df_summary_stats = analysis['result'].to_html(index=True, classes='dataframe', border=0)
                                        interpretation_prompt = f"""J'ai les statistiques descriptives suivantes pour une ou plusieurs colonnes :

{df_summary_stats}

Analyse des colonnes: {params.get('selected_columns', [])}.
Interprète ces statistiques de manière concise. Commence directement par l'interprétation."""

                                    if interpretation_prompt:
                                        try:
                                            response_report_ia = model_chat_report.generate_content(interpretation_prompt)
                                            if response_report_ia and hasattr(response_report_ia, 'text') and response_report_ia.text.strip():
                                                ai_interpretation = f"<div style='background-color: #e6f7ff; border-left: 5px solid #66b3ff; padding: 10px 15px; margin-top: 15px; border-radius: 5px; font-size: 0.95em;'><b>Interprétation IA :</b> {response_report_ia.text.strip()}</div>"
                                        except Exception:
                                            # En cas d'erreur lors de la génération du contenu, on ignore silencieusement
                                            pass
                            except Exception:
                                # En cas d'erreur lors de la vérification des modèles, on ignore silencieusement
                                pass
                        except Exception:
                            # En cas d'erreur de configuration de l'API, on ignore silencieusement
                            pass

                    if isinstance(analysis['result'], dict):
                        html_content += ai_interpretation
                        html_content += "<div style='margin: 20px 0;'>"
                        for key, value in analysis['result'].items():
                            element_title = key.replace('fig', 'Graphique ').replace('fig1', 'Graphique 1').replace('fig2', 'Graphique 2').replace('table', 'Tableau').replace('stats', 'Statistiques')
                            html_content += f"<h4 style='color: #007bff; margin-top: 30px; margin-bottom: 15px;'>📊 {element_title}</h4>"
                            if hasattr(value, 'to_plotly_json'):
                                # Mettre à jour les libellés des axes si nécessaire
                                if 'yaxis' in value.layout and 'title' in value.layout.yaxis and 'rendement' in str(value.layout.yaxis.title.text).lower():
                                    value.update_layout(yaxis_title=f"{value.layout.yaxis.title.text} (tonnes/ha)")
                                
                                # Générer le HTML pour le graphique avec une hauteur fixe
                                plot_div = f"""
                                <div id='plotly-chart-{i}-{key}' style='width:100%; height:500px; margin: 20px 0; border: 1px solid #e0e0e0; border-radius: 8px; overflow: hidden;'></div>
                                <script>
                                    document.addEventListener('DOMContentLoaded', function() {{
                                        var figure = {0};
                                        Plotly.newPlot('plotly-chart-{1}-{2}', 
                                            figure.data, 
                                            figure.layout || {{}}, 
                                            {{responsive: true, displayModeBar: true}}
                                        );
                                        
                                        // Redimensionner le graphique quand la fenêtre change de taille
                                        window.addEventListener('resize', function() {{
                                            Plotly.Plots.resize('plotly-chart-{1}-{2}');
                                        }});
                                    }});
                                </script>
                                """.format(value.to_json(), i, key)
                                html_content += plot_div
                            elif isinstance(value, pd.DataFrame):
                                df_html_table = value.to_html(classes='dataframe', border=0, index=False)
                                html_content += df_html_table
                            else:
                                html_content += f"<pre style='background: #f8f9fa; padding: 15px; border-radius: 5px;'>{value}</pre>"
                        html_content += "</div>"
                    elif isinstance(analysis['result'], pd.DataFrame):
                        html_content += ai_interpretation
                        df_html_table = analysis['result'].to_html(classes='dataframe', border=0, index=False)
                        html_content += df_html_table
                    elif hasattr(analysis['result'], 'to_plotly_json'):
                        html_content += ai_interpretation
                        # Mettre à jour les libellés des axes si nécessaire pour le graphique principal
                        if hasattr(analysis['result'], 'layout') and 'yaxis' in analysis['result'].layout and 'title' in analysis['result'].layout.yaxis:
                            if 'rendement' in str(analysis['result'].layout.yaxis.title.text).lower():
                                analysis['result'].update_layout(yaxis_title=f"{analysis['result'].layout.yaxis.title.text} (tonnes/ha)")
                        
                        # Créer un conteneur pour le graphique avec une hauteur fixe
                        plot_div = """
                        <div id='plotly-main-chart-{0}' style='width:100%; height:500px; margin: 20px 0; border: 1px solid #e0e0e0; border-radius: 8px; overflow: hidden;'></div>
                        <script>
                            document.addEventListener('DOMContentLoaded', function() {{
                                var figure = {1};
                                Plotly.newPlot('plotly-main-chart-{0}', 
                                    figure.data, 
                                    figure.layout || {{}}, 
                                    {{responsive: true, displayModeBar: true}}
                                );
                                
                                // Redimensionner le graphique quand la fenêtre change de taille
                                window.addEventListener('resize', function() {{
                                    Plotly.Plots.resize('plotly-main-chart-{0}');
                                }});
                            }});
                        </script>
                        """.format(i, analysis['result'].to_json())
                        html_content += plot_div
                    else:
                        html_content += ai_interpretation
                        html_content += f"<pre style='background: #f8f9fa; padding: 15px; border-radius: 5px;'>{analysis['result']}</pre>"
        else:
            html_content += "<p>Aucune analyse détaillée à afficher.</p>"

        # Ajout de l'historique du chat IA s'il existe
        gemini_chat_history = st.session_state.get('gemini_chat_history', [])
        if current_api_key and gemini_chat_history:
            html_content += "<h2>Historique du Chat IA (Résumé)</h2>"
            html_content += "<div class='chat-history'>"
            # Résumer les N derniers messages pour le PDF
            num_messages_to_include = 5
            for message in gemini_chat_history[-(num_messages_to_include):]: # Prendre les N derniers
                role = message.get("role", "user")
                content = message.get("parts", [""])[0][:200] # Tronquer le contenu
                role_class = "user" if role == "user" else "assistant"
                html_content += f"<div class='chat-message {role_class}'><strong>{role.capitalize()}:</strong> {content}...</div>"
            if len(gemini_chat_history) > num_messages_to_include:
                html_content += "<p style='font-style:italic;'>... et d'autres messages.</p>"
            html_content += "</div>"
        elif current_api_key: # Si l'API est configurée mais pas d'historique
            html_content += "<p>Aucun historique de chat IA.</p>"
        else: # Si l'API n'est pas configurée
            html_content += "<p>Le chat IA est désactivé (clé API manquante).</p>"

        # Ajout de l'analyse intelligente si elle a été effectuée et demandée pour le rapport
        if st.session_state.get('include_intelligent_analysis_in_reports', True) and st.session_state.get('intelligent_analysis') and df_current is not None:
            html_content += "<h2>🧠 Analyse Intelligente</h2>"
            html_content += "<p>Cette section présente les résultats de l'analyse intelligente basée sur la signification des variables.</p>"
            
            analysis_config = st.session_state.intelligent_analysis
            html_content += f"<h3>🎯 Objectif d'analyse</h3>"
            html_content += f"<p><strong>Objectif :</strong> {analysis_config.get('objective', 'Non spécifié')}</p>"
            html_content += f"<p><strong>Contexte détecté :</strong> {analysis_config.get('context', 'Non spécifié')}</p>"
            html_content += f"<p><strong>Variables sélectionnées :</strong> {', '.join(analysis_config.get('variables', []))}</p>"
            
            # Ajouter les résultats de l'analyse intelligente
            html_content += "<h3>📊 Résultats de l'analyse intelligente</h3>"
            html_content += "<p>Les analyses détaillées sont disponibles dans l'interface principale.</p>"

        # Ajout de l'analyse complète automatique si elle a été effectuée ou demandée pour le rapport
        if st.session_state.get('include_complete_analysis_in_reports', False) and df_current is not None:
            html_content += "<h2>Analyse Complète Automatique</h2><p>Cette section présente une analyse complète et automatique des données collectées.</p>"

            # --- ANALYSE GÉOGRAPHIQUE ---
            html_content += "<h3>🗺️ ANALYSE GÉOGRAPHIQUE</h3>"

            # Détection dynamique des colonnes géographiques
            geo_cols_present = []
            commune_columns = [col for col in df_current.columns if 'commune' in col.lower() or 'ville' in col.lower() or 'localité' in col.lower()]
            dept_columns = [col for col in df_current.columns if 'département' in col.lower() or 'province' in col.lower() or 'région' in col.lower()]

            if dept_columns:
                geo_cols_present.append(dept_columns[0])
            if commune_columns:
                geo_cols_present.append(commune_columns[0])

            # Analyse par département/région
            if dept_columns:
                dept_col = dept_columns[0]
                dept_counts = df_current[dept_col].value_counts()
                try:
                    fig_dept = px.pie(values=dept_counts.values, names=dept_counts.index, title=f"Répartition par {dept_col}", color_discrete_sequence=px.colors.qualitative.Set3)
                    fig_dept.update_traces(textposition='inside', textinfo='percent+label')
                    html_content += f"<h4>📊 Répartition par {dept_col}</h4>"
                    html_content += fig_dept.to_html(full_html=False, include_plotlyjs='inline', config={'responsive': True})
                    html_content += f"<p><strong>Description :</strong> Répartition géographique des répondants par {dept_col}.</p>"
                except Exception as e:
                    html_content += f"<p>Erreur graphique : {e}</p>"

            # Analyse par commune/ville
            if commune_columns:
                commune_col = commune_columns[0]
                commune_counts = df_current[commune_col].value_counts().reset_index()
                commune_counts.columns = [commune_col, "Nombre de répondants"]
                try:
                    fig_commune = px.bar(commune_counts, x=commune_col, y="Nombre de répondants", title=f"Nombre de répondants par {commune_col}", color="Nombre de répondants", color_continuous_scale="Viridis")
                    fig_commune.update_layout(xaxis_tickangle=-45)
                    html_content += f"<h4>📊 Répartition par {commune_col}</h4>"
                    html_content += fig_commune.to_html(full_html=False, include_plotlyjs='inline', config={'responsive': True})
                    html_content += f"<p><strong>Description :</strong> Nombre de répondants par {commune_col}.</p>"

                    # Tableau des communes
                    html_content += f"<h4>📋 Tableau : Répartition par {commune_col}</h4>"
                    commune_table_report = commune_counts.copy()
                    commune_table_report["Pourcentage"] = (commune_table_report["Nombre de répondants"] / commune_table_report["Nombre de répondants"].sum() * 100).round(1)
                    commune_table_report["Pourcentage"] = commune_table_report["Pourcentage"].astype(str) + "%"
                    html_content += commune_table_report.to_html(classes='dataframe', border=0, index=False)
                except Exception as e:
                    html_content += f"<p>Erreur : {e}</p>"

            # Si aucune colonne géographique n'est trouvée
            if not geo_cols_present:
                html_content += "<p><em>Aucune donnée géographique détectée dans cette collecte.</em></p>"

            # --- ANALYSE DÉMOGRAPHIQUE ---
            html_content += "<h3>👥 ANALYSE DÉMOGRAPHIQUE</h3>"
            sexe_col = find_first_matching_column(df_current, ['sexe', 'genre'])
            age_col = find_first_matching_column(df_current, ['âge', 'age'])

            if sexe_col:
                sexe_counts = df_current[sexe_col].value_counts()
                try:
                    fig_sexe = px.pie(values=sexe_counts.values, names=sexe_counts.index, title="Répartition par sexe", color_discrete_sequence=['#FF6B6B', '#4ECDC4'])
                    fig_sexe.update_traces(textposition='inside', textinfo='percent+label')
                    html_content += "<h4>📊 Répartition par sexe</h4>"
                    html_content += fig_sexe.to_html(full_html=False, include_plotlyjs='inline', config={'responsive': True})
                    html_content += "<p><strong>Description :</strong> Répartition des répondants par sexe.</p>"
                except Exception as e:
                    html_content += f"<p>Erreur graphique sexe: {e}</p>"

            if age_col:
                age_counts = df_current[age_col].value_counts()
                try:
                    fig_age = px.bar(x=age_counts.index, y=age_counts.values, title="Répartition par tranche d'âge", color=age_counts.values, color_continuous_scale="Plasma")
                    fig_age.update_layout(xaxis_title="Tranche d'âge", yaxis_title="Nombre de répondants")
                    html_content += "<h4>📊 Répartition par âge</h4>"
                    html_content += fig_age.to_html(full_html=False, include_plotlyjs='inline', config={'responsive': True})
                    html_content += "<p><strong>Description :</strong> Répartition des répondants par tranche d'âge.</p>"
                except Exception as e:
                    html_content += f"<p>Erreur graphique âge: {e}</p>"

            # --- ANALYSE DE SATISFACTION ---
            satisfaction_col = find_first_matching_column(df_current, ['satisfaction', 'satisfait', 'évaluation', 'appréciation', 'note'])
            if satisfaction_col and satisfaction_col in df_current.columns:
                html_content += "<h3>😊 ANALYSE DE SATISFACTION</h3>"
                satisfaction_counts = df_current[satisfaction_col].value_counts()
                try:
                    fig_satisfaction = px.pie(values=satisfaction_counts.values, names=satisfaction_counts.index, title="Répartition de la satisfaction globale", color_discrete_sequence=['#FF6B6B', '#FFE66D', '#4ECDC4'])
                    fig_satisfaction.update_traces(textposition='inside', textinfo='percent+label')
                    html_content += "<h4>📊 Satisfaction globale</h4>"
                    html_content += fig_satisfaction.to_html(full_html=False, include_plotlyjs='inline', config={'responsive': True})
                    html_content += "<p><strong>Description :</strong> Répartition de la satisfaction globale.</p>"
                except Exception as e:
                    html_content += f"<p>Erreur satisfaction: {e}</p>"
            # Si aucune variable de satisfaction n'est trouvée, la section n'est pas ajoutée

            # --- RÉSUMÉ STATISTIQUE GÉNÉRAL ---
            html_content += "<h3>📊 RÉSUMÉ STATISTIQUE</h3>"

            # Calcul dynamique des métriques géographiques
            commune_columns = [col for col in df_current.columns if 'commune' in col.lower() or 'ville' in col.lower() or 'localité' in col.lower()]
            dept_columns = [col for col in df_current.columns if 'département' in col.lower() or 'province' in col.lower() or 'région' in col.lower()]

            communes_count = 0
            if commune_columns:
                communes_count = df_current[commune_columns[0]].nunique()
            elif dept_columns:
                # Si pas de colonne commune, utiliser le département comme approximation
                communes_count = df_current[dept_columns[0]].nunique()

            dept_count = 0
            if dept_columns:
                dept_count = df_current[dept_columns[0]].nunique()

            # Calcul dynamique du taux de satisfaction
            satisfaction_columns = [col for col in df_current.columns if any(keyword in col.lower() for keyword in ['satisfaction', 'satisfait', 'évaluation', 'appréciation', 'note'])]
            satisfaction_rate_str = "N/A"
            if satisfaction_columns and satisfaction_columns[0] in df_current.columns:
                try:
                    satisfaction_col = satisfaction_columns[0]
                    # Chercher les valeurs positives de satisfaction
                    positive_values = ['satisfait', 'très satisfait', 'excellent', 'bon', 'positif', 'oui']
                    satisfaction_count = 0
                    total_responses = len(df_current)

                    # Vérifier le type de données de la colonne de satisfaction
                    if df_current[satisfaction_col].dtype == 'object':
                        # Si c'est une colonne de type string/object
                        for value in positive_values:
                            satisfaction_count += (df_current[satisfaction_col].str.lower().str.contains(value, na=False)).sum()
                    else:
                        # Si c'est une colonne numérique, compter les valeurs positives
                        positive_numeric_values = [1, 2, 3, 4, 5]  # Valeurs numériques positives
                        satisfaction_count = df_current[satisfaction_col].isin(positive_numeric_values).sum()

                    if total_responses > 0:
                        satisfaction_rate = (satisfaction_count / total_responses) * 100
                        satisfaction_rate_str = f"{satisfaction_rate:.1f}%"
                except Exception as e:
                    # En cas d'erreur, utiliser une valeur par défaut
                    satisfaction_rate_str = "N/A"
                    st.warning(f"Erreur lors du calcul du taux de satisfaction: {e}")

            summary_data = {
                'Métrique': ['Total répondants', 'Nombre de communes', 'Nombre de départements', 'Taux de satisfaction'],
                'Valeur': [len(df_current), communes_count, dept_count, satisfaction_rate_str]
            }
            summary_df = pd.DataFrame(summary_data)
            html_content += summary_df.to_html(classes='dataframe', border=0, index=False)

        # Pied de page HTML
        html_content += "<div class='footer'>"
        html_content += "<p>Rapport généré par l'application CCR-B & IFRIZ DataViz.</p>"
        html_content += "<p>Contact: +229 01 96911346 | Portfolio: Sidoine YEBADOKPO</p>"
        html_content += "</div>"

        html_content += "</div></body></html>"
        return html_content

    except Exception as e:
        # Si une erreur survient pendant la génération HTML, retourner un message d'erreur clair
        st.error(f"Erreur critique lors de la génération du rapport HTML: {str(e)}")
        import traceback
        st.error(f"Détails de l'erreur: {traceback.format_exc()}")
        # Retourner un contenu HTML minimal pour indiquer l'erreur
        return f"""
        <html><body><h1>Erreur de Génération du Rapport</h1><p>Une erreur est survenue : {str(e)}. Veuillez vérifier les logs de l'application.</p></body></html>
        """

# --- FONCTION DE GÉNÉRATION PDF ---
def _generate_and_store_pdf(report_html=None):
    """
    Génère un rapport PDF à partir du contenu HTML en utilisant ReportLab.
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        import io

        # Créer un buffer pour stocker le PDF
        buffer = io.BytesIO()

        # Créer le document PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

        # Obtenir les styles par défaut
        styles = getSampleStyleSheet()

        # Créer une liste pour stocker les éléments du PDF
        elements = []

        # Titre du rapport
        elements.append(Paragraph("Rapport d'Analyse CCR-B & Enabel", styles['Title']))
        elements.append(Spacer(1, 12))

        # Informations de base
        elements.append(Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Contenu principal (simplifié pour le PDF)
        if report_html:
            # Extraire le texte du HTML pour le PDF (version simplifiée)
            elements.append(Paragraph("Résumé des analyses effectuées", styles['Heading1']))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph("Les analyses détaillées sont disponibles dans le rapport HTML.", styles['Normal']))
        else:
            elements.append(Paragraph("Aucun contenu HTML disponible pour la génération PDF.", styles['Normal']))

        # Ajouter des informations sur les données si disponibles
        df_current = st.session_state.get('df_current')
        if df_current is not None:
            elements.append(Spacer(1, 12))
            elements.append(Paragraph("Informations sur les données", styles['Heading2']))
            elements.append(Spacer(1, 6))

            # Créer un tableau de résumé
            summary_data = {
                'Métrique': ['Total répondants', 'Nombre de colonnes'],
                'Valeur': [len(df_current), len(df_current.columns)]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_data_for_table = [summary_df.columns.tolist()] + summary_df.values.tolist()
            reportlab_rows_summary = [[Paragraph(str(cell), styles['Normal']) for cell in row] for row in summary_data_for_table]
            table_summary = Table(reportlab_rows_summary, repeatRows=1)
            table_style_summary = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
            ]
            table_summary.setStyle(TableStyle(table_style_summary))
            elements.append(table_summary)
            elements.append(Spacer(1, 15))

            # Ajouter les analyses avec leurs interprétations modifiées
            analyses = st.session_state.get('analyses', [])
            if analyses:
                elements.append(Paragraph("Résultats des analyses", styles['Heading2']))
                elements.append(Spacer(1, 12))
                
                for i, analysis in enumerate(analyses):
                    if analysis.get('result') is not None:
                        # Titre de l'analyse
                        params = analysis.get('executed_params', {})
                        titre = params.get('title', f"Analyse {i+1}")
                        elements.append(Paragraph(f"{i+1}. {titre}", styles['Heading3']))
                        elements.append(Spacer(1, 6))
                        
                        # Interprétation modifiée si disponible
                        modified_interpretation = analysis.get('modified_interpretation', '')
                        if modified_interpretation:
                            elements.append(Paragraph(f"<b>Interprétation IA (modifiée) :</b> {modified_interpretation}", styles['Normal']))
                            elements.append(Spacer(1, 6))
                        
                        # Ajouter le tableau si c'est un DataFrame
                        if isinstance(analysis['result'], pd.DataFrame):
                            # Limiter le nombre de lignes pour le PDF
                            df_for_pdf = analysis['result'].head(50)  # Limiter à 50 lignes pour le PDF
                            
                            # Préparer les données pour ReportLab
                            table_data = [df_for_pdf.columns.tolist()] + df_for_pdf.values.tolist()
                            reportlab_table_data = [[Paragraph(str(cell), styles['Normal']) for cell in row] for row in table_data]
                            
                            # Créer le tableau avec répétition des en-têtes
                            pdf_table = Table(reportlab_table_data, repeatRows=1)
                            pdf_table_style = [
                                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                ('FONTSIZE', (0, 0), (-1, 0), 8),
                                ('FONTSIZE', (0, 1), (-1, -1), 7),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                            ]
                            pdf_table.setStyle(TableStyle(pdf_table_style))
                            elements.append(pdf_table)
                            
                            if len(analysis['result']) > 50:
                                elements.append(Paragraph(f"<i>Note : Affichage limité à 50 lignes sur {len(analysis['result'])} au total.</i>", styles['Normal']))
                            
                            elements.append(Spacer(1, 12))
                        
                        # Ajouter un saut de page après chaque analyse pour éviter la troncature
                        if i < len(analyses) - 1:
                            elements.append(PageBreak())

        # Pied de page
        elements.append(PageBreak())
        elements.append(Paragraph("Rapport généré par l'application CCR-B & IFRIZ DataViz.", styles['Normal']))
        elements.append(Paragraph("Contact: +229 01 96911346 | Portfolio: Sidoine YEBADOKPO", styles['Normal']))

        # Générer le PDF
        if not elements:
            st.error("❌ Erreur : Aucun contenu à intégrer dans le PDF.")
            return False

        with st.spinner("🔄 Génération du PDF en cours..."):
            doc.build(elements)

        # Stocker le PDF dans session_state
        pdf_data = buffer.getvalue()
        buffer.close()

        # Générer un nom de fichier
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_base = f"Rapport_CCRB_"

        st.session_state['pdf_report'] = pdf_data
        st.session_state['pdf_filename'] = filename_base + ".pdf"

        st.success("✅ PDF généré avec succès!")
        return True

    except Exception as e:
        st.error(f"Erreur lors de la génération du PDF: {str(e)}")
        import traceback
        st.error(f"Détails de l'erreur: {traceback.format_exc()}")
        return False

# --- FONCTIONS D'ANALYSE INTELLIGENTE ---
def analyze_variable_significance(df):
    """
    Analyse la signification des variables dans le dataframe et propose des analyses appropriées.
    """
    variable_analysis = {}
    
    for col in df.columns:
        col_analysis = {
            'name': col,
            'dtype': str(df[col].dtype),
            'nunique': df[col].nunique(),
            'missing_pct': (df[col].isnull().sum() / len(df)) * 100,
            'significance': 'unknown',
            'suggested_analyses': [],
            'sample_values': df[col].dropna().unique()[:5].tolist()
        }
        
        # Analyser la signification basée sur le nom de la colonne et les valeurs
        col_lower = col.lower()
        
        # Variables démographiques
        if any(keyword in col_lower for keyword in ['age', 'âge', 'sexe', 'genre', 'gender', 'sex']):
            col_analysis['significance'] = 'demographic'
            col_analysis['suggested_analyses'] = ['distribution', 'comparison', 'correlation']
        elif any(keyword in col_lower for keyword in ['commune', 'ville', 'village', 'region', 'département', 'zone']):
            col_analysis['significance'] = 'geographic'
            col_analysis['suggested_analyses'] = ['distribution', 'mapping', 'comparison']
        elif any(keyword in col_lower for keyword in ['satisfaction', 'satisfait', 'content', 'heureux', 'happy']):
            col_analysis['significance'] = 'satisfaction'
            col_analysis['suggested_analyses'] = ['distribution', 'trend', 'comparison']
        elif any(keyword in col_lower for keyword in ['revenu', 'salaire', 'income', 'argent', 'budget']):
            col_analysis['significance'] = 'economic'
            col_analysis['suggested_analyses'] = ['distribution', 'correlation', 'comparison']
        elif any(keyword in col_lower for keyword in ['éducation', 'education', 'niveau', 'diplôme', 'scolarité']):
            col_analysis['significance'] = 'education'
            col_analysis['suggested_analyses'] = ['distribution', 'correlation', 'comparison']
        elif any(keyword in col_lower for keyword in ['métier', 'profession', 'job', 'travail', 'activité']):
            col_analysis['significance'] = 'professional'
            col_analysis['suggested_analyses'] = ['distribution', 'comparison']
        elif any(keyword in col_lower for keyword in ['famille', 'ménage', 'household', 'membre']):
            col_analysis['significance'] = 'household'
            col_analysis['suggested_analyses'] = ['distribution', 'correlation']
        elif any(keyword in col_lower for keyword in ['santé', 'health', 'maladie', 'médical']):
            col_analysis['significance'] = 'health'
            col_analysis['suggested_analyses'] = ['distribution', 'correlation', 'comparison']
        elif any(keyword in col_lower for keyword in ['agriculture', 'culture', 'production', 'rendement']):
            col_analysis['significance'] = 'agricultural'
            col_analysis['suggested_analyses'] = ['distribution', 'correlation', 'trend']
        elif any(keyword in col_lower for keyword in ['coopérative', 'association', 'groupe', 'organisation']):
            col_analysis['significance'] = 'organizational'
            col_analysis['suggested_analyses'] = ['distribution', 'comparison']
        else:
            # Analyser basé sur les valeurs et le type
            if df[col].dtype in ['int64', 'float64']:
                if df[col].nunique() <= 10:
                    col_analysis['significance'] = 'categorical_numeric'
                    col_analysis['suggested_analyses'] = ['distribution', 'comparison']
                else:
                    col_analysis['significance'] = 'continuous_numeric'
                    col_analysis['suggested_analyses'] = ['distribution', 'correlation', 'trend']
            else:
                if df[col].nunique() <= 20:
                    col_analysis['significance'] = 'categorical'
                    col_analysis['suggested_analyses'] = ['distribution', 'comparison']
                else:
                    col_analysis['significance'] = 'text'
                    col_analysis['suggested_analyses'] = ['text_analysis']
        
        variable_analysis[col] = col_analysis
    
    return variable_analysis

def generate_intelligent_analysis_plan(df, objective):
    """
    Génère un plan d'analyse intelligent basé sur l'objectif et la signification des variables.
    """
    # Analyser la signification des variables
    var_analysis = analyze_variable_significance(df)
    
    # Détecter le contexte de l'analyse basé sur l'objectif
    objective_lower = objective.lower()
    
    if any(keyword in objective_lower for keyword in ['démographique', 'démographie', 'population', 'âge', 'sexe', 'genre']):
        context = "Analyse démographique"
        relevant_vars = [col for col, analysis in var_analysis.items() 
                        if analysis['significance'] in ['demographic', 'age', 'gender']]
    elif any(keyword in objective_lower for keyword in ['géographique', 'géographie', 'commune', 'région', 'zone', 'localité']):
        context = "Analyse géographique"
        relevant_vars = [col for col, analysis in var_analysis.items() 
                        if analysis['significance'] in ['geographic', 'location']]
    elif any(keyword in objective_lower for keyword in ['satisfaction', 'satisfait', 'content', 'évaluation']):
        context = "Étude de satisfaction"
        relevant_vars = [col for col, analysis in var_analysis.items() 
                        if analysis['significance'] in ['satisfaction', 'feedback']]
    elif any(keyword in objective_lower for keyword in ['corrélation', 'relation', 'lien', 'association', 'impact']):
        context = "Étude des corrélations"
        relevant_vars = [col for col, analysis in var_analysis.items() 
                        if analysis['significance'] in ['continuous_numeric', 'economic', 'demographic']]
    elif any(keyword in objective_lower for keyword in ['comparaison', 'comparer', 'différence', 'comparatif']):
        context = "Analyse comparative"
        relevant_vars = [col for col, analysis in var_analysis.items() 
                        if analysis['significance'] in ['categorical', 'demographic', 'geographic']]
    elif any(keyword in objective_lower for keyword in ['prédiction', 'modélisation', 'modèle', 'prédire', 'tendance']):
        context = "Prédiction ou modélisation"
        relevant_vars = [col for col, analysis in var_analysis.items() 
                        if analysis['significance'] in ['continuous_numeric', 'economic', 'demographic']]
    elif any(keyword in objective_lower for keyword in ['vue globale', 'vue d\'ensemble', 'aperçu', 'résumé', 'indicateurs', 'collecte', 'données']):
        context = "Vue d'ensemble des données"
        # Pour une vue globale, sélectionner les variables les plus importantes
        important_vars = [col for col, analysis in var_analysis.items() 
                         if analysis['significance'] != 'unknown' and analysis['significance'] != 'text']
        relevant_vars = important_vars[:10]  # Limiter à 10 variables pour la vue globale
    else:
        context = "Analyse générale"
        relevant_vars = list(df.columns)
    
    # Si pas assez de variables pertinentes, ajouter des variables importantes
    if len(relevant_vars) < 2:
        important_vars = [col for col, analysis in var_analysis.items() 
                         if analysis['significance'] != 'unknown' and col not in relevant_vars]
        relevant_vars.extend(important_vars[:3])
    
    # Déterminer le niveau d'analyse
    if any(keyword in objective_lower for keyword in ['simple', 'basique', 'descriptif']):
        level = "Descriptif (statistiques de base)"
    elif any(keyword in objective_lower for keyword in ['complet', 'détaillé', 'approfondi']):
        level = "Complet (tous les niveaux)"
    elif any(keyword in objective_lower for keyword in ['comparatif', 'comparaison']):
        level = "Comparatif (entre groupes)"
    elif any(keyword in objective_lower for keyword in ['corrélation', 'relation']):
        level = "Corrélationnel (relations entre variables)"
    else:
        level = "Complet (tous les niveaux)"
    
    return {
        'context': context,
        'variables': relevant_vars,
        'level': level,
        'variable_analysis': var_analysis,
        'objective': objective
    }

def generate_enhanced_categorical_analysis(df, column_name):
    """
    Génère une analyse catégorielle améliorée basée sur les statistiques détaillées des modalités.
    """
    st.markdown(f"### 📊 Analyse de répartition de `{column_name}`")
    
    # Compter les valeurs avec gestion des valeurs manquantes
    value_counts = df[column_name].value_counts(dropna=False)
    total_respondents = len(df)
    valid_responses = df[column_name].notna().sum()
    missing_responses = df[column_name].isna().sum()
    
    # Créer le tableau des statistiques détaillées des modalités
    st.subheader(f"📋 Statistiques détaillées des modalités : {column_name}")
    
    # Préparer les données pour le tableau
    stats_data = pd.DataFrame({
        'Modalité': value_counts.index,
        'Nombre': value_counts.values,
        'Pourcentage': (value_counts.values / valid_responses * 100).round(1) if valid_responses > 0 else 0
    })
    
    # Ajouter des colonnes statistiques
    stats_data['Rang'] = stats_data['Nombre'].rank(ascending=False, method='dense').astype(int)
    stats_data['Pourcentage_cumulé'] = stats_data['Pourcentage'].cumsum().round(1)
    stats_data['Pourcentage_str'] = stats_data['Pourcentage'].astype(str) + '%'
    stats_data['Pourcentage_cumulé_str'] = stats_data['Pourcentage_cumulé'].astype(str) + '%'
    
    # Réorganiser les colonnes pour l'affichage
    display_data = stats_data[['Rang', 'Modalité', 'Nombre', 'Pourcentage_str', 'Pourcentage_cumulé_str']]
    display_data.columns = ['Rang', 'Modalité', 'Nombre', 'Pourcentage', 'Pourcentage cumulé']
    
    # Afficher le tableau des statistiques détaillées
    st.dataframe(display_data, use_container_width=True)
    
    # Graphique en camembert basé sur les statistiques des modalités
    if valid_responses > 0:
        # Utiliser les données du tableau pour créer le graphique
        pie_data = stats_data.copy()
        
        # Créer le graphique en camembert basé sur les statistiques
        fig = px.pie(
            data_frame=pie_data,
            values='Nombre',  # Utiliser le nombre comme base
            names='Modalité',
            title=f"Répartition de {column_name} (basée sur les statistiques des modalités)",
            color_discrete_sequence=px.colors.qualitative.Set3,
            hover_data=['Nombre', 'Pourcentage']
        )
        
        # Personnaliser l'affichage pour montrer les statistiques
        fig.update_traces(
            textposition='inside',
            textinfo='percent+label',
            hovertemplate="<b>%{label}</b><br>" +
                          "Nombre: %{customdata[0]}<br>" +
                          "Pourcentage: %{customdata[1]:.1f}%<br>" +
                          "<extra></extra>"
        )
        
        st.plotly_chart(fig, use_container_width=True)

        # Métriques clés basées sur les statistiques
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total répondants", total_respondents)
        col2.metric("Réponses valides", valid_responses)
        col3.metric("Modalités uniques", stats_data.shape[0])
        col4.metric("Modalité principale", f"{stats_data.iloc[0]['Pourcentage']:.1f}%")
        
        # Résumé basé sur les statistiques des modalités
        st.markdown(f"**💡 Résumé basé sur les statistiques des modalités :**")
        st.markdown(f"- **Modalité dominante :** '{stats_data.iloc[0]['Modalité']}' avec {stats_data.iloc[0]['Nombre']} réponses ({stats_data.iloc[0]['Pourcentage']:.1f}%)")
        st.markdown(f"- **Diversité :** {stats_data.shape[0]} modalités différentes")
        if missing_responses > 0:
            missing_pct = (missing_responses / total_respondents * 100)
            st.markdown(f"- **Données manquantes :** {missing_responses} réponses ({missing_pct:.1f}%)")
        
        # Analyse de la distribution basée sur les statistiques
        if stats_data.shape[0] > 1:
            # Calculer l'indice de diversité basé sur les pourcentages du tableau
            proportions = stats_data['Pourcentage'] / 100
            hhi = np.sum(proportions ** 2)
            diversity_index = 1 - hhi
            
            st.markdown(f"- **Indice de diversité :** {diversity_index:.3f} (0 = concentration maximale, 1 = distribution uniforme)")
            
            if diversity_index < 0.3:
                st.markdown("  - *Distribution très concentrée*")
            elif diversity_index < 0.6:
                st.markdown("  - *Distribution modérément diversifiée*")
            else:
                st.markdown("  - *Distribution très diversifiée*")
    
    else:
        st.warning("Aucune donnée valide disponible pour cette variable.")

def generate_demographic_analysis(df, variables):
    """Génère une analyse démographique complète."""
    # Vérifier d'abord s'il y a des variables démographiques dans la dataframe
    demographic_vars = [var for var in variables if var in df.columns and any(keyword in var.lower() for keyword in ['sexe', 'genre', 'âge', 'age', 'démographique', 'population'])]
    
    if not demographic_vars:
        st.info("ℹ️ Aucune variable démographique détectée dans cette collecte.")
        return
    
    st.markdown("**📊 Profil démographique de l'échantillon**")

    # Statistiques descriptives par variable démographique
    for var in demographic_vars:
        st.markdown(f"#### ")
        # Utiliser la nouvelle fonction d'analyse catégorielle améliorée
        generate_enhanced_categorical_analysis(df, var)

def generate_satisfaction_analysis(df, variables):
    """Génère une analyse de satisfaction."""
    # Vérifier d'abord s'il y a des variables de satisfaction dans la dataframe
    satisfaction_vars = [var for var in variables if var in df.columns and any(keyword in var.lower() for keyword in ['satisfaction', 'satisfait', 'évaluation', 'appréciation', 'note'])]
    
    if not satisfaction_vars:
        st.info("ℹ️ Aucune variable de satisfaction détectée dans cette collecte.")
        return
    
    st.markdown("**😊 Analyse de la satisfaction des bénéficiaires**")

    for var in satisfaction_vars:
            st.markdown(f"#### ")

            # Compter les niveaux de satisfaction
            satisfaction_counts = df[var].value_counts()

            # Graphique en barres pour la satisfaction
            # Créer un DataFrame pour le graphique
            bar_data = pd.DataFrame({
                'Niveau de satisfaction': satisfaction_counts.index,
                'Nombre de répondants': satisfaction_counts.values
            })

            fig = px.bar(
                data_frame=bar_data,
                x='Niveau de satisfaction',
                y='Nombre de répondants',
                title=f"Niveaux de satisfaction - "
            )
            st.plotly_chart(fig, use_container_width=True)

            # Calcul du taux de satisfaction global
            positive_keywords = ['satisfait', 'content', 'très satisfait', 'excellent', 'bon']
            total_responses = len(df[var].dropna())
            positive_responses = 0

            for keyword in positive_keywords:
                # Vérifier le type de données avant d'utiliser .str
                if df[var].dtype == 'object':
                    positive_responses += len(df[df[var].astype(str).str.contains(keyword, case=False, na=False)])
                else:
                    # Pour les colonnes numériques, chercher des valeurs positives
                    positive_numeric_values = [1, 2, 3, 4, 5]
                    positive_responses += len(df[df[var].isin(positive_numeric_values)])

            satisfaction_rate = (positive_responses / total_responses * 100) if total_responses > 0 else 0

            col1, col2, col3 = st.columns(3)
            col1.metric("Total répondants", total_responses)
            col2.metric("Réponses positives", positive_responses)
            col3.metric("Taux de satisfaction", f"{satisfaction_rate:.1f}%")

def generate_geographic_analysis(df, variables):
    """Génère une analyse géographique."""
    # Vérifier d'abord s'il y a des variables géographiques dans la dataframe
    geographic_vars = [var for var in variables if var in df.columns and any(keyword in var.lower() for keyword in ['commune', 'ville', 'localité', 'département', 'province', 'région', 'zone', 'géographique'])]
    
    if not geographic_vars:
        st.info("ℹ️ Aucune variable géographique détectée dans cette collecte.")
        return
    
    st.markdown("**🗺️ Analyse géographique des données**")

    for var in geographic_vars:
            st.markdown(f"#### ")

            # Compter par région/commune
            geo_counts = df[var].value_counts()

            # Graphique en barres pour la répartition géographique
            # Créer un DataFrame pour le graphique
            geo_data = pd.DataFrame({
                'Zone géographique': geo_counts.index,
                'Nombre de répondants': geo_counts.values
            })

            fig = px.bar(
                data_frame=geo_data,
                x='Zone géographique',
                y='Nombre de répondants',
                title=f"Répartition géographique - "
            )
            fig.update_xaxes(tickangle=45)
            st.plotly_chart(fig, use_container_width=True)

            # Statistiques géographiques
            col1, col2, col3 = st.columns(3)
            col1.metric("Nombre de zones", geo_counts.nunique())
            col2.metric("Zone la plus représentée", geo_counts.index[0] if not geo_counts.empty else "N/A")
            col3.metric("Répondants dans la zone principale", geo_counts.iloc[0] if not geo_counts.empty else 0)

def generate_correlation_analysis(df, variables):
    """Génère une analyse des corrélations."""
    st.markdown("**🔗 Analyse des relations entre variables**")

    # Filtrer les variables numériques
    numerical_vars = [var for var in variables if var in df.columns and df[var].dtype in ['int64', 'float64']]

    if len(numerical_vars) >= 2:
        # Matrice de corrélation
        corr_matrix = df[numerical_vars].corr()

        fig = px.imshow(
            corr_matrix,
            title="Matrice de corrélation",
            color_continuous_scale='RdBu',
            aspect='auto'
        )
        st.plotly_chart(fig, use_container_width=True)

        # Interprétation des corrélations
        st.markdown("**💡 Interprétation :**")
        st.markdown("- Rouge : Corrélation positive forte")
        st.markdown("- Bleu : Corrélation négative forte")
        st.markdown("- Blanc : Pas de corrélation")

        # Corrélations significatives
        significant_correlations = []
        for i in range(len(corr_matrix.columns)):
            for j in range(i+1, len(corr_matrix.columns)):
                corr_value = corr_matrix.iloc[i, j]
                if abs(corr_value) > 0.3: # Seuil de corrélation significative
                    significant_correlations.append({
                        'Variable 1': corr_matrix.columns[i],
                        'Variable 2': corr_matrix.columns[j],
                        'Corrélation': f"{corr_value:.3f}"
                    })

        if significant_correlations:
            st.markdown("**🔍 Corrélations significatives :**")
            corr_df = pd.DataFrame(significant_correlations)
            st.dataframe(corr_df, use_container_width=True)
    else:
        st.warning("⚠️ Au moins 2 variables numériques sont nécessaires pour l'analyse des corrélations.")

def generate_comparative_analysis(df, variables):
    """Génère une analyse comparative entre groupes."""
    st.markdown("**⚖️ Analyse comparative entre groupes**")

    if len(variables) >= 2:
        # Variable de groupement (première variable)
        group_var = variables[0]
        # Variables à comparer (autres variables)
        compare_vars = variables[1:]

        if group_var in df.columns:
            st.markdown(f"#### Comparaison par : ")

            for compare_var in compare_vars:
                if compare_var in df.columns:
                    st.markdown(f"** par **")

                    # Statistiques par groupe
                    group_stats = df.groupby(group_var)[compare_var].agg(['count', 'mean', 'std']).round(2)
                    st.dataframe(group_stats, use_container_width=True)

                    # Graphique de comparaison
                    if df[compare_var].dtype in ['int64', 'float64']:
                        fig = px.box(
                            df,
                            x=group_var,
                            y=compare_var,
                            title=f"Comparaison de par "
                        )
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        # Pour les variables catégorielles
                        cross_tab = pd.crosstab(df[group_var], df[compare_var])
                        fig = px.imshow(
                            cross_tab,
                            title=f"Tableau croisé : vs ",
                            color_continuous_scale='Blues'
                        )
                        st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning("⚠️ Au moins 2 variables sont nécessaires pour l'analyse comparative.")

def generate_predictive_analysis(df, variables):
    """Génère une analyse prédictive basique."""
    st.markdown("**🔮 Analyse prédictive**")

    # Filtrer les variables numériques pour la prédiction
    numerical_vars = [var for var in variables if var in df.columns and df[var].dtype in ['int64', 'float64']]

    if len(numerical_vars) >= 2:
        st.markdown("**📈 Régression linéaire simple**")

        # Sélection de la variable cible
        target_var = st.selectbox("Variable à prédire :", numerical_vars)
        predictor_vars = [var for var in numerical_vars if var != target_var]

        if predictor_vars:
            predictor_var = st.selectbox("Variable prédictrice :", predictor_vars)

            # Régression linéaire
            from sklearn.linear_model import LinearRegression
            from sklearn.model_selection import train_test_split

            # Préparer les données
            X = df[[predictor_var]].dropna()
            y = df[target_var].dropna()

            # Aligner les indices
            common_index = X.index.intersection(y.index)
            X = X.loc[common_index]
            y = y.loc[common_index]

            if len(X) > 10: # Minimum de données
                # Diviser en train/test
                X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

                # Entraîner le modèle
                model = LinearRegression()
                model.fit(X_train, y_train)

                # Prédictions
                y_pred = model.predict(X_test)

                # Graphique de régression
                # Créer un DataFrame pour le graphique
                regression_data = pd.DataFrame({
                    predictor_var: X_test[predictor_var],
                    target_var: y_test
                })

                fig = px.scatter(
                    data_frame=regression_data,
                    x=predictor_var,
                    y=target_var,
                    title=f"Régression : vs "
                )

                # Ajouter la ligne de régression
                fig.add_scatter(
                    x=X_test[predictor_var],
                    y=y_pred,
                    mode='lines',
                    name='Prédiction',
                    line=dict(color='red')
                )

                st.plotly_chart(fig, use_container_width=True)

                # Métriques du modèle
                from sklearn.metrics import r2_score, mean_squared_error
                r2 = r2_score(y_test, y_pred)
                rmse = mean_squared_error(y_test, y_pred, squared=False)

                col1, col2 = st.columns(2)
                col1.metric("R² (Qualité du modèle)", f"{r2:.3f}")
                col2.metric("RMSE (Erreur)", f"{rmse:.3f}")

                st.markdown(f"**Équation :** = {model.coef_[0]:.3f} × + {model.intercept_:.3f}")
            else:
                st.warning("⚠️ Données insuffisantes pour la modélisation.")
    else:
        st.warning("⚠️ Au moins 2 variables numériques sont nécessaires pour l'analyse prédictive.")

def generate_general_analysis(df, variables, level):
    """Génère une analyse générale selon le niveau demandé."""
    st.markdown("**📈 Analyse générale des données**")

    if level == "Descriptif (statistiques de base)" or level == "Complet (tous les niveaux)":
        st.markdown("#### 📊 Statistiques descriptives")
        for var in variables:
            if var in df.columns:
                st.markdown(f"****")
                if df[var].dtype in ['int64', 'float64']:
                    stats = df[var].describe()
                    col1, col2, col3, col4 = st.columns(4)
                    col1.metric("Moyenne", f"{stats['mean']:.2f}")
                    col2.metric("Médiane", f"{stats['50%']:.2f}")
                    col3.metric("Écart-type", f"{stats['std']:.2f}")
                    col4.metric("Min-Max", f"{stats['min']:.0f} - {stats['max']:.0f}")

                    # Histogramme
                    fig = px.histogram(df, x=var, title=f"Distribution - ")
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    # Utiliser la nouvelle fonction d'analyse catégorielle améliorée
                    generate_enhanced_categorical_analysis(df, var)

    if level == "Comparatif (entre groupes)" or level == "Complet (tous les niveaux)":
        if len(variables) >= 2:
            generate_comparative_analysis(df, variables)

    if level == "Corrélationnel (relations entre variables)" or level == "Complet (tous les niveaux)":
        generate_correlation_analysis(df, variables)

    if level == "Prédictif (modélisation)" or level == "Complet (tous les niveaux)":
        generate_predictive_analysis(df, variables)

# --- FONCTION DE CHARGEMENT DES DONNÉES ---
@st.cache_data(ttl=600, max_entries=10) # Cache les données pendant 10 minutes, max 10 entrées
def load_data(url):
    """Charge les données depuis une URL KoboToolbox en utilisant requests et pandas (ultra-optimisé)."""
    try:
        # Récupérer un éventuel token Kobo depuis la session ou l'environnement
        kobo_token = st.session_state.get('kobo_token') or os.getenv('KOBOTOOLBOX_TOKEN')
        # Utiliser un User-Agent pour simuler un navigateur et éviter certains blocages
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
        if kobo_token:
            headers['Authorization'] = f'Token {kobo_token}'

        # Timeout plus court pour éviter les blocages
        response = requests.get(url, headers=headers, timeout=60) # Augmenter le timeout à 60 secondes
        try:
            response.raise_for_status() # Lève une exception pour les codes d'erreur HTTP (4xx ou 5xx)
        except requests.exceptions.HTTPError as http_err:
            # Si 404, tenter des URLs alternatives KoboToolbox
            status = getattr(http_err.response, 'status_code', None)
            if status == 404:
                alt_urls = []
                # 1) Remplacer export Excel par CSV si disponible
                if url.endswith('data.xlsx'):
                    alt_urls.append(url[:-5] + 'csv') # data.csv
                # 2) Essayer l'endpoint data avec format explicite (même domaine que l'URL originale)
                m = re.search(r"(https?://[^/]+)/api/v2/assets/([^/]+)/", url)
                if m:
                    original_domain = m.group(1)
                    asset_uid = m.group(2)
                    base = f"{original_domain}/api/v2/assets/{asset_uid}/data/"
                    alt_urls.extend([
                        base + "?format=xlsx",
                        base + "?format=csv"
                    ])

                last_error = http_err
                for alt in alt_urls:
                    try:
                        alt_resp = requests.get(alt, headers=headers, timeout=60)
                        alt_resp.raise_for_status()
                        # Décoder selon le format
                        if alt.endswith('csv') or 'format=csv' in alt:
                            df = pd.read_csv(io.BytesIO(alt_resp.content))
                        else:
                            df = pd.read_excel(io.BytesIO(alt_resp.content), engine='openpyxl')
                        url = alt  # garder trace de l'URL fonctionnelle
                        break
                    except Exception as e_alt:
                        last_error = e_alt
                        df = None
                if df is None:
                    raise last_error
            else:
                raise
        else:
            # Essayer de lire le fichier Excel avec des paramètres ultra-optimisés
            df = pd.read_excel(
                io.BytesIO(response.content),
                engine='openpyxl',
            )

        # Nettoyage initial des colonnes : supprimer les préfixes Kobo (ex: 'X / Y / Z') et les caractères non valides
        # Créer une nouvelle liste de noms de colonnes nettoyés et uniques
        new_columns = []
        seen_cleaned_names = set()
        for col in df.columns:
            # Garder seulement la dernière partie du nom si plusieurs '/' sont présents
            cleaned_col = col.split('/')[-1].strip()
            # Remplacer les caractères potentiellement problématiques pour l'IA/SQL
            cleaned_col = re.sub(r'[^\w\s-]', '', cleaned_col).strip() # Garder lettres, chiffres, espaces, tirets
            cleaned_col = re.sub(r'[-\s]+', '_', cleaned_col) # Remplacer espaces/tirets par underscore
            cleaned_col = cleaned_col.lower() # Tout mettre en minuscules

            # Gérer les doublons après nettoyage
            original_cleaned_col = cleaned_col
            counter = 1
            while cleaned_col in seen_cleaned_names:
                cleaned_col = f"{original_cleaned_col}_{counter}"
                counter += 1
            seen_cleaned_names.add(cleaned_col)
            new_columns.append(cleaned_col)

        df.columns = new_columns

        # Ajouter une colonne '_index' pour compter les formulaires soumis (utile pour les rapports)
        # Utiliser une colonne Kobo si elle existe, sinon un simple index
        potential_id_cols = ['_index', 'index', 'ID', 'id', 'formhub/uuid']
        id_col_found = None
        for col in potential_id_cols:
            if col in df.columns and df[col].nunique() > 0:
                id_col_found = col
                break
        if id_col_found:
            df['_index'] = df[id_col_found].astype(str) # Assurer que c'est une chaîne pour éviter problèmes
        else: # Si pas de colonne ID évidente, utiliser l'index de Pandas
            df['_index'] = df.index.astype(str)

        # Supprimer les colonnes potentiellement vides ou non pertinentes issues du nettoyage
        df = df.drop(columns=[col for col in df.columns if col.startswith('unnamed:')], errors='ignore')
        df = df.loc[:, ~df.columns.str.contains('^index_[0-9]*$')] # Supprimer les colonnes d'index générées par erreur
        df = df.loc[:, ~df.columns.str.contains('^col_[0-9]*$')] # Supprimer les colonnes nettoyées vides

        return df

    except requests.exceptions.Timeout:
        st.error(f"Erreur de timeout lors du chargement des données depuis {url}. Le serveur n'a pas répondu à temps.")
        return None
    except requests.exceptions.RequestException as e:
        st.error(f"Erreur lors de la requête vers {url}: {e}")
        return None
    except pd.errors.ParserError as e:
        st.error(f"Erreur lors de l'analyse du fichier Excel depuis {url}: {e}. Assurez-vous que le fichier est un .xlsx valide.")
        return None
    except Exception as e:
        st.error(f"Une erreur inattendue est survenue lors du chargement des données depuis {url}: {e}")
        import traceback
        st.error(f"Détails de l'erreur: {traceback.format_exc()}")
        return None

# --- FONCTION DE GÉNÉRATION DE REQUÊTE SQL AVEC IA ---
def generate_sql_with_gemini(nl_query, relevant_columns, numerical_cols, categorical_cols, current_data_columns, api_key):
    """
    Génère une requête SQL en utilisant l'API Google Gemini.
    """
    try:
        prompt = f"""
        Vous êtes un assistant IA expert en SQL, spécialisé dans la génération de requêtes précises et efficaces pour l'analyse de données.

        **OBJECTIF :** Traduire la demande en langage naturel en une requête SQL valide pour le DataFrame fourni.

        **CONTEXTE DES DONNÉES ACTUELLES :**
        - Table source : `data`
        - Colonnes disponibles (nettoyées et uniques) : {', '.join(f'"{c}"' for c in current_data_columns)}
        - Colonnes numériques : {', '.join(f'"{c}"' for c in numerical_cols) if numerical_cols else 'Aucune'}
        - Colonnes catégorielles : {', '.join(f'"{c}"' for c in categorical_cols) if categorical_cols else 'Aucune'}

        **DEMANDE UTILISATEUR :**
        {nl_query}

        **INSTRUCTIONS STRICTES POUR LA GÉNÉRATION SQL :**
        1. **Précision :** Analyser attentivement la demande et utiliser UNIQUEMENT les colonnes listées ci-dessus.
        2. **Formatage :** ENCADRER TOUS les noms de colonnes et le nom de la table (`data`) avec des guillemets doubles (`"`).
        3. **Calculs Explicites :** Si une agrégation ou un calcul est demandé, écrire la fonction SQL explicite.
        4. **Fonctions d'Agrégation :** Utiliser `COUNT(*)` pour les comptages globaux, `COUNT("colonne")` pour les comptages non nuls.
        5. **Groupement (`GROUP BY`) :** Utiliser `GROUP BY` avec les fonctions d'agrégation sur des colonnes catégorielles.
        6. **Tri (`ORDER BY`) :** Ajouter `ORDER BY` quand c'est logiquement nécessaire.
        7. **SORTIE :** UNIQUEMENT la requête SQL, sans commentaires ni explications.
        8. **INTERDICTIONS :** Ne jamais entourer avec ```sql ... ``` et ne jamais préfixer par `sql `.
        """

        genai.configure(api_key=api_key)
        model = genai.GenerativeModel(get_effective_gemini_model_name())
        response = model.generate_content(prompt)
        sql = (response.text or "").strip()
        sql = re.sub(r"^```(?:sql)?\\s*", "", sql, flags=re.IGNORECASE).strip()
        sql = re.sub(r"\\s*```\\s*$", "", sql).strip()
        sql = re.sub(r"^\\s*sql\\s+", "", sql, flags=re.IGNORECASE).strip()
        return sql, f"Requête : {nl_query[:40]}..."
        
    except Exception as e:
        raise Exception(f"Erreur avec Gemini: {str(e)}")

def generate_sql_with_openai(nl_query, relevant_columns, numerical_cols, categorical_cols, current_data_columns, api_key, model_name: str | None = None):
    """
    Génère une requête SQL en utilisant l'API OpenAI (fallback lorsque Gemini est en quota).
    """
    try:
        model_name = model_name or DEFAULT_OPENAI_MODEL
        prompt = f"""
Vous êtes un assistant expert en SQL. Traduisez la demande utilisateur en une requête SQL valide pour un DataFrame.

Contexte:
- Table: data
- Colonnes disponibles: {', '.join(f'"{c}"' for c in current_data_columns)}
- Colonnes numériques: {', '.join(f'"{c}"' for c in numerical_cols) if numerical_cols else 'Aucune'}
- Colonnes catégorielles: {', '.join(f'"{c}"' for c in categorical_cols) if categorical_cols else 'Aucune'}

Demande:
{nl_query}

Règles:
- N'utilisez QUE les colonnes listées
- Entourez TOUS les noms de colonnes et la table "data" avec des guillemets doubles
- Retournez UNIQUEMENT la requête SQL, sans explications, sans ``` ni commentaires, sans préfixe `sql `
""".strip()

        from openai import OpenAI  # type: ignore

        client = OpenAI(api_key=api_key)
        resp = client.chat.completions.create(
            model=model_name,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
            max_tokens=500,
        )
        sql = (resp.choices[0].message.content or "").strip()
        sql = re.sub(r"^```(?:sql)?\\s*", "", sql, flags=re.IGNORECASE).strip()
        sql = re.sub(r"\\s*```\\s*$", "", sql).strip()
        sql = re.sub(r"^\\s*sql\\s+", "", sql, flags=re.IGNORECASE).strip()
        return sql, f"Requête (OpenAI) : {nl_query[:40]}..."
    except Exception as e:
        raise Exception(f"Erreur avec OpenAI: {str(e)}")

def _generate_fallback_sql(nl_query: str, current_data_columns: list[str], numerical_cols: list[str], categorical_cols: list[str]) -> tuple[str, str]:
    """
    Fallback local (sans IA) pour générer une requête SQL simple.
    Objectif: éviter un crash quand Gemini/OpenAI sont en quota.
    """
    q = (nl_query or "").strip()
    ql = q.lower()

    # Utilitaires
    def first_present(cands: list[str]) -> str | None:
        for c in cands:
            if c in current_data_columns:
                return c
        return None

    def detect_col_from_text(cols: list[str]) -> str | None:
        for c in cols:
            if c and c.lower() in ql:
                return c
        return None

    cat = detect_col_from_text(categorical_cols) or (categorical_cols[0] if categorical_cols else None)
    num = detect_col_from_text(numerical_cols) or (numerical_cols[0] if numerical_cols else None)
    if not cat:
        cat = first_present([c for c in current_data_columns if "region" in c.lower() or "région" in c.lower() or "commune" in c.lower() or "depart" in c.lower() or "départ" in c.lower()])
    if not num:
        num = first_present([c for c in current_data_columns if "age" in c.lower() or "montant" in c.lower() or "prix" in c.lower() or "revenu" in c.lower() or "production" in c.lower() or "rendement" in c.lower()])

    # Colonnes utiles pour "point des soumissions"
    date_col = first_present(
        [
            c
            for c in current_data_columns
            if any(k in c.lower() for k in ["_submission_time", "submission", "submitted", "date", "time"])
        ]
    )
    collecte_col = first_present([c for c in current_data_columns if "collect" in c.lower() or "sous" in c.lower() or "enquete" in c.lower() or "enquête" in c.lower() or "form" in c.lower()])

    # Patterns
    if any(k in ql for k in ["soumis", "soumise", "soumissions", "formulaire", "formulaires", "enquête", "enquetes", "enquete", "point", "bilan", "derni"]) and (collecte_col or cat):
        group_col = collecte_col or cat
        if date_col:
            sql = (
                f'SELECT "{group_col}" AS "{group_col}", COUNT(*) AS "Nombre", MAX("{date_col}") AS "Dernière_soumission" '
                f'FROM "data" GROUP BY "{group_col}" ORDER BY "Nombre" DESC'
            )
            return sql, f"Requête (fallback) : point des soumissions par {group_col}"
        sql = (
            f'SELECT "{group_col}" AS "{group_col}", COUNT(*) AS "Nombre" '
            f'FROM "data" GROUP BY "{group_col}" ORDER BY "Nombre" DESC'
        )
        return sql, f"Requête (fallback) : point des soumissions par {group_col}"

    if any(k in ql for k in ["count", "compter", "nombre", "effectif", "répartition", "repartition", "par "]) and cat:
        sql = (
            f'SELECT "{cat}" AS "{cat}", COUNT(*) AS "Nombre" '
            f'FROM "data" GROUP BY "{cat}" ORDER BY "Nombre" DESC'
        )
        return sql, f"Requête (fallback) : comptage par {cat}"

    if any(k in ql for k in ["moyenne", "mean", "avg"]) and num and cat:
        sql = (
            f'SELECT "{cat}" AS "{cat}", AVG("{num}") AS "Moyenne_{num}" '
            f'FROM "data" GROUP BY "{cat}" ORDER BY "Moyenne_{num}" DESC'
        )
        return sql, f"Requête (fallback) : moyenne de {num} par {cat}"

    if any(k in ql for k in ["moyenne", "mean", "avg"]) and num:
        sql = f'SELECT AVG("{num}") AS "Moyenne_{num}" FROM "data"'
        return sql, f"Requête (fallback) : moyenne de {num}"

    if any(k in ql for k in ["max", "maximum"]) and num:
        sql = f'SELECT MAX("{num}") AS "Max_{num}" FROM "data"'
        return sql, f"Requête (fallback) : max de {num}"

    if any(k in ql for k in ["min", "minimum"]) and num:
        sql = f'SELECT MIN("{num}") AS "Min_{num}" FROM "data"'
        return sql, f"Requête (fallback) : min de {num}"

    # Par défaut: aperçu
    sql = 'SELECT * FROM "data" LIMIT 100'
    return sql, "Requête (fallback) : aperçu"

def generate_sql_with_local_gemma(nl_query, relevant_columns, numerical_cols, categorical_cols, current_data_columns):
    """
    Génère une requête SQL en utilisant le modèle Gemma local.
    """
    try:
        model, tokenizer = load_local_gemma()
        if not model or not tokenizer:
            raise Exception("Modèle Gemma non chargé")
            
        prompt = f"""<start_of_turn>user
        Vous êtes un expert SQL. Générez une requête SQL pour cette demande : {nl_query}
        
        Colonnes disponibles : {', '.join(current_data_columns)}
        Colonnes numériques : {', '.join(numerical_cols) if numerical_cols else 'Aucune'}
        Colonnes catégorielles : {', '.join(categorical_cols) if categorical_cols else 'Aucune'}
        
        Règles importantes :
        - Utilisez des guillemets doubles pour les noms de colonnes et la table
        - Ne mettez que la requête SQL, sans commentaires
        - La table s'appelle "data"
        - Exemple : SELECT "colonne" FROM "data" WHERE "age" > 30
        
        Requête SQL :<end_of_turn><start_of_turn>model
        """
        
        device = next(model.parameters()).device
        inputs = tokenizer(prompt, return_tensors="pt").to(device)
        outputs = model.generate(**inputs, max_new_tokens=200)
        sql = tokenizer.decode(outputs[0], skip_special_tokens=True)
        
        # Nettoyer la sortie
        sql = sql.replace(prompt, "").strip()
        sql = re.sub(r'<.*?>', '', sql)  # Enlever les balises HTML
        sql = re.sub(r'```sql', '', sql, flags=re.IGNORECASE)
        sql = re.sub(r'```', '', sql).strip()
        
        return sql, f"Requête (Gemma) : {nl_query[:40]}..."
        
    except Exception as e:
        raise Exception(f"Erreur avec Gemma local: {str(e)}")

def generate_sql_with_ai(nl_query, relevant_columns, numerical_cols, categorical_cols, current_data_columns):
    """
    Génère une requête SQL en langage naturel en utilisant soit Google Gemini soit le modèle Gemma local.
    Retourne la requête SQL générée et un titre descriptif.
    """
    use_local_model = st.session_state.get('use_local_model', False)
    allow_fallback = st.session_state.get("allow_fallback_ai", False)

    if use_local_model:
        if not os.path.exists(LOCAL_GEMMA_PATH):
            st.error(
                f"Gemma local activé mais introuvable sous `{LOCAL_GEMMA_PATH}`. "
                "Corrigez `LOCAL_GEMMA_PATH` ou placez les fichiers du modèle."
            )
            return "", ""
        elif _local_gemma_ready():
            try:
                return generate_sql_with_local_gemma(
                    nl_query, relevant_columns, numerical_cols, categorical_cols, current_data_columns
                )
            except Exception as e:
                st.error(f"Échec avec le modèle local Gemma : {e}")
                return "", ""
        else:
            details = ""
            try:
                global local_gemma_last_error
                if local_gemma_last_error:
                    details = f" (détail: {local_gemma_last_error})"
            except Exception:
                pass
            st.error(f"Gemma local activé mais non chargé.{details}")
            return "", ""

    current_api_key = st.session_state.get('user_api_key') or google_api_key
    if not current_api_key:
        st.error("La génération SQL par IA est désactivée car aucune clé API valide n'est configurée.")
        return "", ""
    try:
        return generate_sql_with_gemini(
            nl_query, relevant_columns, numerical_cols, categorical_cols, current_data_columns, current_api_key
        )
    except Exception as e:
        if _looks_like_gemini_quota_error(e) and _has_openai_key():
            st.warning("Quota Gemini atteint (429). Basculement vers OpenAI.")
            openai_key = os.environ.get("OPENAI_API_KEY") or getattr(openai, "api_key", None)
            try:
                return generate_sql_with_openai(
                    nl_query,
                    relevant_columns,
                    numerical_cols,
                    categorical_cols,
                    current_data_columns,
                    openai_key,
                    model_name=DEFAULT_OPENAI_MODEL,
                )
            except Exception as e2:
                st.error(f"OpenAI indisponible (quota / billing) : {e2}")
                if allow_fallback:
                    st.warning("Mode secours activé: génération SQL simple sans IA.")
                    return _generate_fallback_sql(nl_query, current_data_columns, numerical_cols, categorical_cols)
                st.error("Mode secours désactivé: impossible de générer la requête SQL sans IA.")
                return "", ""

        st.error(f"Gemini indisponible : {e}")
        if allow_fallback:
            st.warning("Mode secours activé: génération SQL simple sans IA.")
            return _generate_fallback_sql(nl_query, current_data_columns, numerical_cols, categorical_cols)
        st.error("Mode secours désactivé: impossible de générer la requête SQL sans IA.")
        return "", ""

# --- FONCTION DE GÉNÉRATION DE SCRIPT PYTHON ---
def execute_python_script(script, data):
    """
    Exécute un script Python généré par l'IA de manière sécurisée.
    Gère les résultats potentiels (graphiques, DataFrames) et les erreurs.
    Inclut des patchs pour les problèmes courants (MultiIndex, value_counts).
    """
    if not script or not script.strip():
        st.error("❌ Le script Python est vide.")
        return None

    # --- Préparation du script et des variables ---
    cleaned_script = script.replace('\r\n', '\n').replace('\r', '\n')
    # Supprimer BOM et espaces de tête
    cleaned_script = cleaned_script.lstrip('\ufeff\n\t\r ')
    # Supprimer les caractères de format/largeur nulle connus (peuvent casser la syntaxe en silence)
    try:
        import re as _re
        cleaned_script = _re.sub(r'[\u200B\u200C\u200D\u2060\uFEFF]', '', cleaned_script)
    except Exception:
        pass
    # Normaliser les espaces Unicode courants en espace ASCII
    cleaned_script = cleaned_script.replace('\u00A0', ' ').replace('\u202F', ' ').replace('\u2009', ' ')
    # Filtrer les caractères de contrôle invisibles (sauf \n et \t)
    cleaned_script = ''.join(ch for ch in cleaned_script if ch == '\n' or ch == '\t' or ord(ch) >= 32)
    # Retirer les nouvelles lignes superflues en fin
    cleaned_script = cleaned_script.strip('\n')

    # Supprimer les imports du script (l'environnement fournit déjà pd, np, etc.)
    try:
        cleaned_lines = []
        for _line in cleaned_script.split('\n'):
            stripped = _line.lstrip()
            if stripped.startswith('import ') or stripped.startswith('from '):
                continue
            cleaned_lines.append(_line)
        cleaned_script = '\n'.join(cleaned_lines).lstrip('\n')
    except Exception:
        pass

    # Harmoniser les vieux scripts basés sur "parts" pour retourner le dernier résultat utile
    try:
        if 'parts =' in cleaned_script and 'result = parts[0]' in cleaned_script:
            import re as _re2
            cleaned_script = _re2.sub(r"result\s*=\s*parts\[0\].*",
                                       "result = parts[-1] if parts else df.head(10)",
                                       cleaned_script)
    except Exception:
        pass

    # Patch Plotly Express pour gérer les MultiIndex et les noms de colonnes complexes
    safe_px_code = """
import pandas as pd
import re

try:
    import plotly.express as px
except Exception:
    px = None

import matplotlib.pyplot as plt
import seaborn as sns
import scipy.stats as stats


def safe_plotly_call(func_name, data_obj, **kwargs):
    # Récupérer la fonction plotly si px est disponible, sinon None
    try:
        plotly_func = getattr(px, func_name) if px is not None else None
    except Exception:
        plotly_func = None
    if not plotly_func:
        # Fallback Matplotlib si nécessaire
        if func_name == 'bar':
            try:
                if isinstance(data_obj, pd.DataFrame) and data_obj.shape[1] >= 2:
                    fig_mpl = plt.figure(figsize=(10, 6)); sns.barplot(data=data_obj, x=data_obj.columns[0], y=data_obj.columns[1])
                    plt.title(kwargs.get('title', f'Bar plot {data_obj.columns[1]} vs {data_obj.columns[0]}')); plt.xticks(rotation=45, ha='right'); plt.tight_layout()
                    return fig_mpl
                else: return "Erreur: Données invalides pour Matplotlib bar."
    except Exception as e_mpl:
        return f"Erreur Matplotlib: {e_mpl}"
        else:
            return f"Erreur: Fonction Plotly inconnue '{func_name}'"

    # Préparation des données pour Plotly
    processed_data = data_obj
    if isinstance(data_obj, pd.DataFrame):
        if isinstance(data_obj.index, pd.MultiIndex): processed_data = data_obj.reset_index()
        if isinstance(processed_data.columns, pd.MultiIndex):
            processed_data = processed_data.copy()
            processed_data.columns = ['_'.join(map(str, col)).strip() for col in processed_data.columns.values]

    # Adapter les kwargs si nécessaire (ex: colonnes)
    processed_kwargs = kwargs.copy()
    available_cols = processed_data.columns.tolist() if isinstance(processed_data, pd.DataFrame) else []
    for key, value in processed_kwargs.items():
        if isinstance(value, pd.Series):
            if isinstance(value.index, pd.MultiIndex): processed_kwargs[key] = value.reset_index(level=list(range(value.index.nlevels)))
        elif isinstance(value, pd.DataFrame) and value is not data_obj and isinstance(value.columns, pd.MultiIndex):
            value_processed = value.copy(); value_processed.columns = ['_'.join(map(str, col)).strip() for col in value_processed.columns.values]; processed_kwargs[key] = value_processed

    # Gérer spécifiquement les cas problématiques courants
    # 1. value_counts().reset_index() -> renommer les colonnes
    if func_name in ['bar', 'pie'] and isinstance(processed_data, pd.DataFrame):
        # Si le script utilise directement value_counts().reset_index() et assigne le résultat
        # il est préférable de le patcher avant l'exécution. Ici, on essaye de gérer les cas où
        # les colonnes sont mal nommées lors de l'appel direct à Plotly.
        # Les vrais patchs sont appliqués plus bas au niveau du script global.
        pass

    # Vérification finale des colonnes avant appel
    final_data_obj = processed_data
    final_kwargs = kwargs
    if isinstance(final_data_obj, pd.DataFrame):
        for arg_name, arg_value in final_kwargs.items():
            if isinstance(arg_value, str) and arg_value not in available_cols:
                potential_match = next((c for c in available_cols if c.lower() == arg_value.lower()), None)
                if potential_match: final_kwargs[arg_name] = potential_match

    try:
        # Si plotly_func est introuvable, fallback Matplotlib
        if plotly_func is None:
            if func_name == 'bar':
                try:
                    if isinstance(processed_data, pd.DataFrame) and processed_data.shape[1] >= 2:
                        fig_mpl = plt.figure(figsize=(10, 6)); sns.barplot(data=processed_data, x=processed_data.columns[0], y=processed_data.columns[1])
                        plt.title(final_kwargs.get('title', f"Bar plot {processed_data.columns[1]} vs {processed_data.columns[0]}")); plt.xticks(rotation=45, ha='right'); plt.tight_layout()
                        return fig_mpl
                except Exception:
                    return "Erreur: Fallback Matplotlib bar a échoué."
            if func_name == 'histogram':
                try:
                    fig_mpl = plt.figure(figsize=(10, 6))
                    x = final_kwargs.get('x')
                    if isinstance(processed_data, pd.DataFrame) and x in processed_data.columns:
                        sns.histplot(data=processed_data, x=x, bins=final_kwargs.get('nbins', 30))
                    elif isinstance(processed_data, pd.Series):
                        sns.histplot(processed_data, bins=final_kwargs.get('nbins', 30))
                    plt.tight_layout(); return fig_mpl
                except Exception:
                    return "Erreur: Fallback Matplotlib histogram a échoué."
            return f"Erreur: Fonction Plotly inconnue '{func_name}'"

        return plotly_func(final_data_obj, **final_kwargs)
    except Exception as e:
        if "MultiIndex" in str(e) or "Index" in str(e):
            if isinstance(processed_data, pd.DataFrame): processed_data.plotly_error = f"Problème d'index Plotly: {str(e)}"; return processed_data
            else: return f"Erreur Plotly Index (): {str(e)}"
        elif "KeyError" in str(e): return f"Erreur Plotly (): Colonne non trouvée - {str(e)}."
        else: return f"Erreur Plotly (): {str(e)}"
"""
    # Remplacer les appels directs px.func(...) par safe_plotly_call('func', ...)
    # Gérer les arguments multi-lignes en mode DOTALL
    for func_name in safe_px_funcs:
        pattern = rf'px\.{func_name}\((.*?)\)'
        replacer = lambda match: f"safe_plotly_call('{func_name}', {match.group(1)})"
        cleaned_script = re.sub(pattern, replacer, cleaned_script, flags=re.DOTALL)

    # --- Patch pour value_counts().reset_index() ---
    patch_vc_code = r"""
import re
# Fonction pour appliquer le patch sur une ligne de code
def apply_vc_patch(line):
    # Pattern: variable['col'].value_counts().reset_index()
    match = re.search(r'([\w\.\'"]+)\.value_counts\(\)\.reset_index\(\)', line)
    if match:
        expression_base = match.group(1) # ex: df['col']
        # Tentative d'extraction du nom de colonne original pour le titre
        original_col_name_match = re.search(r'[\'"]([^\'"]+)[\'"]', expression_base)
        original_col_name = original_col_name_match.group(1) if original_col_name_match else "Value"

        # Le patch doit créer le DataFrame temporaire, renommer les colonnes, puis assigner
        # Remplacer l'appel original par une assignation au résultat du patch
        # Le résultat sera une variable temporaire (__vc_result__) contenant le DataFrame correctement nommé
        patch_assignment = f"_vc_result_ = pd.DataFrame({expression_base}.value_counts().reset_index())\\n_vc_result_.columns = ['', 'Count']"
        # Si la ligne était une assignation, on la modifie pour utiliser la nouvelle structure
        # ex: df_counts = data['col'].value_counts().reset_index() devient:
        # _vc_result_ = data['col'].value_counts().reset_index()
        # _vc_result_.columns = ['', 'Count']
        # df_counts = _vc_result_

        # Pour simplifier, on remplace directement l'appel par une assignation au résultat du patch,
        # et on s'assure que le résultat final de l'analyse est bien ce DataFrame patché.
        # On va plutôt remplacer l'appel lui-même et laisser le script gérer l'assignation
        # L'idée est de générer le code qui sera substitué
        return f"({expression_base}.value_counts().reset_index(), ['', 'Count'])" # Retourne une paire (data, cols_names)
    return line

# Appliquer le patch ligne par ligne
lines = cleaned_script.split('\\n')
patched_lines = []
for line in lines:
    # Vérifier si la ligne contient un appel value_counts().reset_index() qui est assigné
    assign_match = re.search(r'(\w+)\s*=\s*([\w\.\'"]+\.value_counts\(\)\.reset_index\(\))', line)
    if assign_match:
        var_name = assign_match.group(1)
        expression = assign_match.group(2)
        # Créer le code patché qui retourne le tuple (df, colonnes)
        patched_expression_code = apply_vc_patch(expression)
        # Assigner le résultat du patch à la variable cible, puis renommer les colonnes
        patched_lines.append(f"{var_name}, {var_name}_cols = {patched_expression_code}")
        patched_lines.append(f"{var_name}.columns = {var_name}_cols")
    else:
        # Si ce n'est pas une assignation directe, juste appliquer le patch pour potentiellement capturer le résultat
        patched_lines.append(apply_vc_patch(line))

cleaned_script = '\\n'.join(patched_lines)
"""
    # Exécuter le patch de manière dynamique
    try:
        # Environnement minimal pour le patch, incluant les variables nécessaires
        local_vars_exec = {'__builtins__': __builtins__, 'pd': pd, 're': re, 'cleaned_script': cleaned_script}
        exec(patch_vc_code, globals(), local_vars_exec)
        cleaned_script = local_vars_exec.get('cleaned_script', cleaned_script) # Récupérer le script modifié
    except Exception as e_patch:
        st.error(f"Erreur lors de l'application du patch value_counts: {e_patch}")

    # --- Exécution ---

    try:
        # Pré-compilation pour mieux rapporter les erreurs de syntaxe
        compiled_obj = compile(cleaned_script, '<string>', 'exec')

        # S'assurer que px est disponible
        try:
            import plotly.express as _px
        except Exception:
            _px = None

        # Créer un environnement d'exécution sécurisé (utiliser le même dict pour globals et locals)
        env = {
            'data': data.copy(), # Copie pour éviter de modifier le DataFrame original
            'df': data.copy(),   # Alias attendu par certains scripts générés
            'pd': pd,
            'np': np,
            'px': (_px if _px is not None else px),
            'go': go,
            'plt': plt,
            'sns': sns if 'sns' in globals() else None,
            'stats': stats if 'stats' in globals() else None,
            'result': None # Variable pour stocker le résultat de l'exécution
        }

        # Fallback px si absent: fournir bar/histogram via Matplotlib/Seaborn
        if env.get('px') is None:
            class _PXFallback:
                @staticmethod
                def bar(data_obj, x=None, y=None, **kwargs):
                    fig = plt.figure(figsize=(10, 6))
                    try:
                        if isinstance(data_obj, pd.DataFrame):
                            x_col = x or (data_obj.columns[0] if data_obj.shape[1] >= 1 else None)
                            y_col = y or (data_obj.columns[1] if data_obj.shape[1] >= 2 else None)
                            if x_col is not None and y_col is not None:
                                sns.barplot(data=data_obj, x=x_col, y=y_col)
                                plt.xticks(rotation=45, ha='right')
                            else:
                                data_obj.plot(kind='bar')
                        else:
                            plt.plot([])
                    except Exception:
                        plt.plot([])
                    plt.tight_layout()
                    return fig

                @staticmethod
                def histogram(data_obj, x=None, **kwargs):
                    fig = plt.figure(figsize=(10, 6))
                    try:
                        if isinstance(data_obj, pd.DataFrame) and x in data_obj.columns:
                            sns.histplot(data=data_obj, x=x, bins=kwargs.get('nbins', 30))
                        elif isinstance(data_obj, pd.Series):
                            sns.histplot(data=data_obj, bins=kwargs.get('nbins', 30))
                        else:
                            plt.plot([])
                    except Exception:
                        plt.plot([])
                    plt.tight_layout()
                    return fig

            env['px'] = _PXFallback
        # Injecter la fonction safe_plotly_call dans l'environnement d'exécution
        # Utiliser exec (et non eval) car le bloc contient des imports/instructions
        try:
            _px_globals: dict = {}
            _px_locals: dict = {}
            _compiled_safe = compile(safe_px_code, '<safe_plotly_call>', 'exec')
            exec(_compiled_safe, _px_globals, _px_locals)
            env['safe_plotly_call'] = _px_locals.get('safe_plotly_call') or _px_globals.get('safe_plotly_call')
        except Exception as _px_err:
            st.warning(f"Chargement du wrapper Plotly échoué: {_px_err}. Utilisation directe de px.*")
            env['safe_plotly_call'] = None

        # Exécuter le script pré-compilé dans l'environnement préparé
        exec(compiled_obj, env, env)

        # --- RÉCUPÉRATION DU RÉSULTAT ---
        result = None
        # Chercher le résultat dans les variables connues (fig, result, df, table, stats, etc.)
        # ou dans la variable 'result' explicitement assignée par le script.
        if 'result' in env and env['result'] is not None:
            result = env['result']
        else:
            # Itérer sur les variables locales pour trouver le résultat le plus probable
            for var_name in ['fig', 'df', 'table', 'stats', 'data', 'res', 'analysis', 'figure', 'crosstab', '_vc_result_']:
                if var_name in env and env[var_name] is not None:
                    # Éviter de prendre le DataFrame original 'data' comme résultat s'il n'a pas été modifié
                    if var_name == 'data' and env[var_name] is data:
                        continue
                    # Éviter de prendre les variables temporaires de patch
                    if var_name.startswith('_vc_result_'):
                        continue

                    result = env[var_name]
                    break # Prendre le premier résultat trouvé

        # Si le résultat est un dictionnaire (plusieurs sorties)
        if isinstance(result, dict):
            # Vérifier si l'un des éléments est un graphique Plotly ou un DataFrame
            final_result_dict = {}
            for key, value in result.items():
                if isinstance(value, (pd.DataFrame, pd.Series)):
                    final_result_dict[key] = value
                elif hasattr(value, 'to_plotly_json'): # C'est un graphique Plotly
                    final_result_dict[key] = value
                elif isinstance(value, plt.Figure): # C'est un graphique Matplotlib
                    final_result_dict[key] = value
                elif isinstance(value, str) and "Erreur" in value: # Gérer les erreurs retournées par les wrappers
                    final_result_dict[key] = value # Conserver le message d'erreur
                # Ignorer les autres types de résultats pour le moment ou les gérer si nécessaire

            if final_result_dict: # Si on a trouvé au moins un résultat pertinent
                result = final_result_dict
            else:
                st.warning("Le script s'est exécuté mais aucun résultat graphique ou tabulaire n'a été trouvé.")
                result = None # Vider le résultat s'il n'est pas pertinent

        # Si le résultat est un graphique Plotly directement
        elif hasattr(result, 'to_plotly_json'):
            pass # C'est déjà le résultat attendu
        # Si le résultat est un DataFrame ou une Series
        elif isinstance(result, (pd.DataFrame, pd.Series)):
            pass # C'est déjà le résultat attendu
        # Si le résultat est une figure Matplotlib
        elif isinstance(result, plt.Figure):
            pass # C'est déjà le résultat attendu
        # Si le résultat est une chaîne contenant une erreur
        elif isinstance(result, str) and "Erreur" in result:
            st.error(f"Erreur détectée dans le script : ")
            return None # Retourner None en cas d'erreur
        # Si le résultat est None ou une chaîne vide
        elif result is None or (isinstance(result, str) and not result.strip()):
            st.warning("Le script s'est exécuté mais n'a pas produit de résultat identifiable (graphique, tableau, etc.).")
            return None
        # Si le résultat est un objet avec une erreur d'index Plotly
        elif hasattr(result, 'plotly_error'):
            st.error(f"Erreur lors de la génération du graphique : {result.plotly_error}")
            st.warning("Affichage du DataFrame brut à la place.")
            # Retourner le DataFrame pour qu'il soit affiché comme tableau
            if isinstance(result, pd.DataFrame):
                return result
            else:
                return None # Ne pas retourner si ce n'est pas un DataFrame

        # Si le résultat est toujours None à ce stade, c'est un problème
        if result is None:
            st.warning("Impossible de déterminer le résultat du script.")
            return None

        return result

    except SyntaxError as se:
        # Rapport détaillé pour les erreurs de syntaxe
        st.error("❌ Erreur de syntaxe dans le script Python:")
        bad_line = (se.text or '').rstrip('\n') if hasattr(se, 'text') else ''
        pointer = ''
        if getattr(se, 'offset', None):
            pointer = ' ' * (se.offset - 1) + '^'
        st.code(f"Ligne {se.lineno}: {bad_line}\n{pointer}\nMessage: {getattr(se, 'msg', '')}", language="text")
        # Afficher un contexte des 10 premières lignes avec caractères invisibles visibles
        try:
            def _visualize(s):
                out = []
                for ch in s:
                    o = ord(ch)
                    if ch in ('\n', '\t'):
                        out.append(ch)
                    elif 32 <= o <= 126:
                        out.append(ch)
                    else:
                        out.append(f"\\u{o:04X}")
                return ''.join(out)

            lines = cleaned_script.split('\n')
            preview = []
            for idx, line in enumerate(lines[:10], start=1):
                preview.append(f"{idx:02d}: " + _visualize(line))
            st.code("\n".join(preview), language="text")
        except Exception:
            pass
        return None
    except Exception as e:
        # Gérer les autres erreurs d'exécution du script
        st.error("❌ Erreur lors de l'exécution du script Python:")
        st.error(f"Détails: {str(e)}")
        st.info("💡 Conseils pour corriger le script :")
        st.info("- Vérifiez la syntaxe Python.")
        st.info("- Assurez-vous que les noms de colonnes sont corrects et correspondent à ceux de vos données.")
        st.info("- Vérifiez que toutes les bibliothèques nécessaires sont importées dans le script généré.")
        st.info("- Essayez de simplifier votre demande pour l'IA.")
        return None

# --- FONCTION D'EXÉCUTION DE REQUÊTE SQL ---
def execute_sql_query(query, data):
    """
    Exécute une requête SQL sur le DataFrame Pandas en utilisant pandasql.
    Gère le nettoyage des noms de colonnes pour la compatibilité SQL.
    """
    if not query or not query.strip():
        st.warning("Veuillez entrer ou générer une requête SQL.")
        return None

    try:
        # Nettoyer les fences markdown et autres wrappers IA
        q_raw = (query or "").strip()
        try:
            import re as _re_fence
            q_raw = _re_fence.sub(r"^```(?:sql)?\\s*", "", q_raw, flags=_re_fence.IGNORECASE)
            q_raw = _re_fence.sub(r"\\s*```\\s*$", "", q_raw)
            # Certains modèles renvoient un préfixe "sql " au lieu de ```sql
            q_raw = _re_fence.sub(r"^\\s*sql\\s+", "", q_raw, flags=_re_fence.IGNORECASE)
        except Exception:
            pass

        def _strip_to_sql_start(s: str) -> str:
            """Supprime les préfixes non-SQL (ex: 'Requête SQL générée:') et force le début sur SELECT/WITH."""
            try:
                import re as _re_start
                s2 = (s or "").strip()
                # Retirer un éventuel préfixe 'sql' collé au début ou après du texte
                s2 = _re_start.sub(r"^[^a-zA-Z]*sql\\s+", "", s2, flags=_re_start.IGNORECASE)
                m = _re_start.search(r"(?is)\\b(with|select)\\b", s2)
                if m:
                    s2 = s2[m.start():]
                # Re-retirer un éventuel 'sql' restant
                s2 = _re_start.sub(r"^[^a-zA-Z]*sql\\s+", "", s2, flags=_re_start.IGNORECASE)
                return s2.strip()
            except Exception:
                return (s or "").strip()

        q_raw = _strip_to_sql_start(q_raw)

        # Support multi-requêtes: pandasql/sqlite n'exécute qu'une instruction à la fois
        statements: list[str] = []
        try:
            for part in q_raw.split(";"):
                s = part.strip()
                if s:
                    # Nettoyer le préfixe "sql " sur chaque instruction (certaines réponses IA le répètent)
                    try:
                        import re as _re_stmt
                        s = _re_stmt.sub(r"^\s*sql\s+", "", s, flags=_re_stmt.IGNORECASE).strip()
                    except Exception:
                        pass
                    s = _strip_to_sql_start(s)
                    statements.append(s)
        except Exception:
            statements = [q_raw] if q_raw else []

        if not statements:
            st.warning("Requête SQL vide après nettoyage.")
            return None

        # Copier le DataFrame pour ne pas modifier l'original
        df_for_sql = data.copy()

        # Nettoyer les noms de colonnes pour SQL
        sanitized_columns = {}
        seen_sanitized_names = set()
        for col in df_for_sql.columns:
            sanitized_name = sanitize_column_name_for_sql(col, seen_sanitized_names)
            sanitized_columns[col] = sanitized_name

        df_for_sql.rename(columns=sanitized_columns, inplace=True)

        # Construire une table de correspondance "normalisée" (tolère accents/ponctuation)
        normalized_to_sanitized: dict[str, str] = {}
        try:
            import unicodedata as _ud
            import re as _re_norm

            def _norm_id(s: str) -> str:
                s = str(s or "").strip().strip('"').strip()
                s = _ud.normalize("NFKD", s)
                s = "".join(ch for ch in s if not _ud.combining(ch))
                s = s.lower()
                s = _re_norm.sub(r"[^a-z0-9_]+", "_", s)
                s = _re_norm.sub(r"_+", "_", s).strip("_")
                return s

            for original, sanitized in sanitized_columns.items():
                normalized_to_sanitized[_norm_id(original)] = sanitized
        except Exception:
            normalized_to_sanitized = {}

        # Réécrire la requête pour utiliser les noms de colonnes nettoyés lorsqu'ils sont entre guillemets
        try:
            q = statements[0]
            for original, sanitized in sanitized_columns.items():
                # Remplacer "original" par sanitized
                q = q.replace(f'"{original}"', sanitized)
                # Remplacer `original` par sanitized si l'utilisateur utilise des backticks
                q = q.replace(f'`{original}`', sanitized)
            # Remplacement tolérant pour les identifiants entre guillemets (ex: accents, apostrophes, etc.)
            if normalized_to_sanitized:
                def _replace_quoted_ident(m):
                    inner = m.group(1)
                    try:
                        import unicodedata as _ud2
                        import re as _re2
                        s = _ud2.normalize("NFKD", inner)
                        s = "".join(ch for ch in s if not _ud2.combining(ch))
                        s = s.lower()
                        s = _re2.sub(r"[^a-z0-9_]+", "_", s)
                        s = _re2.sub(r"_+", "_", s).strip("_")
                    except Exception:
                        s = inner
                    repl = normalized_to_sanitized.get(s)
                    return repl if repl else m.group(0)

                import re as _re_qi
                q = _re_qi.sub(r'"([^"]+)"', _replace_quoted_ident, q)
            # Normaliser la référence à la table data (pandasql attend le nom de variable sans guillemets)
            import re as _re_sql
            q = _re_sql.sub(r'(?i)["`\[]?data["`\]]?', 'data', q)
            query_sanitized = q
        except Exception:
            query_sanitized = statements[0]

        # Sécurité: s'assurer que la requête ne commence jamais par "sql "
        query_sanitized = _strip_to_sql_start(query_sanitized)

        # Vérifier si le DataFrame est valide après nettoyage
        if df_for_sql is None or df_for_sql.empty:
            st.error("Le DataFrame est vide ou invalide après le nettoyage des colonnes pour SQL.")
            return None

        # Exécuter la requête SQL (utiliser pandasql importé en tant que ps)
        try:
            _ps = globals().get('ps', None)
            if _ps is None:
                import pandasql as _ps  # type: ignore

            # Si plusieurs requêtes: exécuter toutes et retourner la dernière (et garder les autres en info)
            last_df = None
            # Exécuter la première déjà sanitizée
            last_df = _ps.sqldf(query_sanitized, {'data': df_for_sql})
            # Puis les suivantes (sanitization basique: table data + mapping colonnes)
            for stmt in statements[1:]:
                stmt2 = stmt
                for original, sanitized in sanitized_columns.items():
                    stmt2 = stmt2.replace(f'"{original}"', sanitized).replace(f'`{original}`', sanitized)
                stmt2 = _re_sql.sub(r'(?i)["`\[]?data["`\]]?', 'data', stmt2)
                stmt2 = _strip_to_sql_start(stmt2)
                last_df = _ps.sqldf(stmt2, {'data': df_for_sql})
            sql_result_df = last_df
        except ModuleNotFoundError:
            st.error("pandasql n'est pas installé. Exécutez: pip install pandasql")
            return None
        except Exception as e_px:
            st.error(f"Erreur d'exécution SQL: {e_px}")
            return None

        # Mapper les noms de colonnes SQL nettoyés vers les noms originaux si nécessaire pour l'affichage
        # (Pas strictement nécessaire si l'utilisateur s'attend à voir les noms nettoyés)
        # Pour l'instant, on retourne avec les noms nettoyés.

        return sql_result_df

    except Exception as e_sql_run:
        st.error(f"Erreur lors de l'exécution de la requête SQL: {e_sql_run}")
        return None

# --- INITIALISATION ET GESTION DE L'ÉTAT ---
def reset_analysis_state():
    """Réinitialise l'état des analyses et du chat IA."""
    st.session_state.analyses = []
    st.session_state.gemini_chat_history = []
    st.session_state.gemini_chat = None
    st.session_state.generated_python_script = None
    st.session_state.generated_python_title = None
    st.session_state.sql_query_manual = ''
    st.session_state.sql_query_title = '' # Ajouté pour le titre de la requête SQL IA
    st.session_state.show_complete_analysis = False
    st.session_state.include_complete_analysis_in_reports = False
    st.session_state.intelligent_analysis = None
    st.session_state.selected_analyses = []

# Variables pour gérer l'état de l'application (optimisé pour le chargement)
session_vars = {
    'data_loaded_id': None,
    'analyses': [],
    'dataframe_to_export': None,
    'df_current': None,
    'show_analysis_config': True,
    'gemini_chat_history': [],
    'gemini_chat': None,
    'show_chat_ia': True,
    'user_api_key': None,
    'pdf_report': b'',
    'pdf_filename': '',
    'generated_python_script': None,
    'generated_python_title': None,
    'sql_query_manual': '',
    'sql_query_title': '',
    'show_complete_analysis': False,
    'include_complete_analysis_in_reports': False,
    'include_intelligent_analysis_in_reports': True,
    'intelligent_analysis': None,
    'selected_analyses': [],
    'current_collection_header': None # Ajouté pour suivre la source de données
    ,'allow_fallback_ai': False  # Mode secours (sans IA) désactivé par défaut
    ,'interpretation_counter': 0
}

# Initialisation optimisée en une seule passe
for var_name, default_value in session_vars.items():
    if var_name not in st.session_state:
        st.session_state[var_name] = default_value

# --- DÉFINITION DES ONGLET ---
# Organisation des onglets principaux pour l'interface utilisateur
main_tabs = [
    "📈 Graphique Principal (Plein écran)",
    "📊 Application & Chat IA",
    "🔬 Analyse Avancée",
    "📖 Manuel d'utilisation",
    "📊 Tableau de Bord",
    "📅 Planification"
]
# Création des onglets
try:
    main_graph_tab, app_tab, advanced_tab, manual_tab, dashboard_tab, schedule_tab = st.tabs(main_tabs)
except Exception as e:
    st.error(f"Erreur lors de la création des onglets Streamlit: ")
    # En cas d'échec, afficher un message et arrêter l'exécution
    st.stop()

# --- CONTENU DE L'ONGLET "Graphique Principal (Plein écran)" ---
with main_graph_tab:
    st.header("📈 Graphique Principal (Plein écran)")
    st.info("Cet onglet est réservé pour l'affichage interactif de graphiques générés. Utilisez l'onglet 'Application & Chat IA' pour créer vos analyses.")

    # Afficher les graphiques Plotly existants s'il y en a
    analyses = st.session_state.get('analyses', [])
    if analyses:
        graph_analyses = [a for a in analyses if a.get('type') == 'graph' and a.get('result') is not None]
        if graph_analyses:
            st.subheader("Graphiques disponibles")
            for i, analysis in enumerate(graph_analyses):
                titre = analysis.get('executed_params', {}).get('title', f'Graphique {i+1}')
                st.markdown(f"### ")
                try:
                    st.plotly_chart(analysis['result'], use_container_width=True)
                except Exception as e:
                    # Gérer les erreurs d'affichage de graphique (ex: MultiIndex non géré)
                    st.warning(f"Erreur d'affichage du graphique '': ")
                    st.warning("Ce graphique peut contenir des données avec une structure complexe (ex: MultiIndex).")
                    # Afficher le DataFrame brut si le graphique échoue
                    if isinstance(analysis['result'], pd.DataFrame):
                        st.dataframe(analysis['result'])
                    elif hasattr(analysis['result'], 'plotly_error'): # Si notre wrapper a marqué une erreur
                        st.dataframe(analysis['result']) # Afficher le dataframe brut
                        st.code("Pour corriger, essayez de reformuler la demande ou d'utiliser .reset_index() sur les données avant le graphique.", language="python")
    else:
        st.info("Aucun graphique n'a encore été créé. Allez dans l'onglet 'Application & Chat IA' pour commencer.")
    # else:
    #     st.info("Aucune analyse n'a encore été effectuée. Allez dans l'onglet 'Application & Chat IA' pour commencer.")

# --- CONTENU DE L'ONGLET "Application & Chat IA" ---
with app_tab:
    st.title("📊 Application & Chat IA")
    st.sidebar.title("CCR-B & IFRIZ DataViz")

    # --- GESTION DES DONNÉES ET CONFIGURATION DE LA BARRE LATÉRALE ---

    # Déterminer le logo et le nom de l'organisation pour la barre latérale et l'en-tête
    selected_header_key = st.session_state.get('current_collection_header', '')
    logo_path, org_name, logo_width = get_logo_and_org_info(selected_header_key)

    # Afficher le logo dans la sidebar
    if logo_path and os.path.exists(logo_path):
        st.sidebar.image(logo_path, width=logo_width)
        st.sidebar.markdown(f"****")
    else:
        st.sidebar.warning(f"Logo '{os.path.basename(logo_path)}' non trouvé.")
        st.sidebar.markdown(f"****") # Afficher le nom même sans logo

    # Sélecteur de source de données
    data_source_mode = st.sidebar.radio("Sélectionnez la source de données :", ('Mode API KoboToolbox', 'Mode Fichier Local'), help="Choisissez de charger les données depuis une URL KoboToolbox ou un fichier Excel local.")

    # Variables pour stocker les informations sur les données chargées
    selected_collection_name = None
    uploaded_file = None
    data = None
    all_columns = []
    numerical_columns = []
    categorical_columns = []

    # --- CHARGEMENT DES DONNÉES VIA KOBOTOOLBOX ---
    if data_source_mode == 'Mode API KoboToolbox':
        st.sidebar.header("📊 Chargement des données KoboToolbox")
        # Champ optionnel pour le token API KoboToolbox
        kobo_token_input = st.sidebar.text_input(
            "Token API KoboToolbox (optionnel)",
            value=st.session_state.get('kobo_token', ''),
            type="password",
            help="Collez votre jeton API Kobo pour accéder aux formulaires privés."
        )
        if kobo_token_input != st.session_state.get('kobo_token'):
            st.session_state['kobo_token'] = kobo_token_input
        selected_collection_name = st.sidebar.selectbox("Sélectionnez une collecte:", list(urls.keys()), index=0, help="Choisissez une des collectes de données KoboToolbox prédéfinies.")

        if selected_collection_name:
            selected_url = urls[selected_collection_name]

            # Mettre à jour l'en-tête si la collecte sélectionnée a changé
            if st.session_state.get('current_collection_header') != selected_collection_name:
                st.session_state.current_collection_header = selected_collection_name
                reset_analysis_state() # Réinitialiser l'état

            # Boutons de contrôle du chargement
            col1, col2 = st.sidebar.columns(2)
            with col1:
                if st.button("📥 Charger les données", type="primary", help="Cliquez pour charger manuellement les données de la collecte sélectionnée"):
                    # Charger les données manuellement
                    current_data_id = f"url_{hash(selected_url)}"
                    if st.session_state.data_loaded_id != current_data_id:
                        # Indicateur de progression optimisé
                        progress_bar = st.progress(0)
                        status_text = st.empty()

                        status_text.text("🔄 Connexion au serveur KoboToolbox...")
                        progress_bar.progress(20)

                        try:
                            with st.spinner("Téléchargement des données..."):
                                status_text.text("📥 Téléchargement en cours...")
                                progress_bar.progress(40)

                                data = load_data(selected_url)

                                if data is not None and not data.empty:
                                    status_text.text("🔧 Traitement et nettoyage des données...")
                                    progress_bar.progress(80)

                                    st.session_state.dataframe_to_export = data
                                    st.session_state.df_current = data
                                    st.session_state.data_loaded_id = current_data_id

                                    progress_bar.progress(100)
                                    status_text.text("✅ Chargement terminé !")

                                    st.success(f"✅ Données chargées avec succès ! ({len(data)} lignes, {len(data.columns)} colonnes)")
                                else:
                                    st.error("❌ Erreur lors du chargement des données. Veuillez vérifier l'URL ou réessayer.")
                                    st.session_state.df_current = None
                                    st.session_state.current_collection_header = None

                        except Exception as e:
                            st.error(f"❌ Erreur de connexion ou de traitement : {str(e)}")
                            st.session_state.df_current = None
                            st.session_state.current_collection_header = None
                    else:
                        st.info("✅ Les données sont déjà chargées !")

            with col2:
                if st.button("🗑️ Effacer les données", type="secondary", help="Efface les données actuellement chargées"):
                    st.session_state.df_current = None
                    st.session_state.data_loaded_id = None
                    st.session_state.current_collection_header = None
                    st.success("✅ Données effacées avec succès !")
                    st.rerun()

            # Afficher le statut de chargement
            if st.session_state.df_current is not None:
                st.sidebar.success(f"✅ Données chargées : {len(st.session_state.df_current)} lignes")
            else:
                st.sidebar.info("📋 Aucune donnée chargée. Cliquez sur 'Charger les données' pour commencer.")

            # Si les données sont chargées, extraire les colonnes
            if st.session_state.df_current is not None:
                all_columns = st.session_state.df_current.columns.tolist()
                numerical_columns = [col for col in st.session_state.df_current.columns if pd.api.types.is_numeric_dtype(st.session_state.df_current[col])]
                categorical_columns = [col for col in st.session_state.df_current.columns if pd.api.types.is_object_dtype(st.session_state.df_current[col]) or isinstance(st.session_state.df_current[col].dtype, pd.CategoricalDtype)]

    # --- CHARGEMENT DES DONNÉES DEPUIS UN FICHIER LOCAL ---
    elif data_source_mode == 'Mode Fichier Local':
        st.header("Charger un fichier Excel local")
        uploaded_file = st.file_uploader("Déposez votre fichier Excel ici (.xlsx)", type=["xlsx"], help="Chargez votre propre fichier de données au format Excel.")
        if uploaded_file is not None:
            # Mettre à jour l'en-tête et réinitialiser l'état si un nouveau fichier est chargé
            current_file_header = f"Fichier local: {uploaded_file.name}"
            if st.session_state.get('current_collection_header') != current_file_header:
                st.session_state.current_collection_header = current_file_header
                reset_analysis_state() # Réinitialiser l'état

            # Boutons de contrôle du chargement
            col1, col2 = st.sidebar.columns(2)
            with col1:
                if st.button("📥 Charger le fichier", type="primary", help="Cliquez pour charger le fichier Excel sélectionné"):
                    try:
                        # Indicateur de progression pour les fichiers locaux
                        progress_bar = st.progress(0)
                        status_text = st.empty()

                        status_text.text("📁 Lecture du fichier...")
                        progress_bar.progress(30)

                        with st.spinner(f"Chargement du fichier '{uploaded_file.name}'..."):
                            status_text.text("🔧 Traitement et nettoyage des données...")
                            progress_bar.progress(60)

                            data = pd.read_excel(uploaded_file) # Suppression de la limitation de lignes

                            progress_bar.progress(100)
                            status_text.text("✅ Chargement terminé !")

                            st.success(f"✅ Fichier chargé avec succès ! ({len(data)} lignes, {len(data.columns)} colonnes)")
                            st.session_state.dataframe_to_export = data
                            st.session_state.df_current = data
                            st.session_state.data_loaded_id = f"file_{hash(uploaded_file.name)}" # Mettre à jour l'ID chargé

                    except Exception as e:
                        st.error(f"❌ Erreur lors du chargement du fichier : {str(e)}")
                        # Réinitialiser l'état si le chargement échoue
                        st.session_state.data_loaded_id = None
                        st.session_state.df_current = None
                        st.session_state.current_collection_header = None
                        data = None
            with col2:
                if st.button("🗑️ Effacer le fichier", type="secondary", help="Efface le fichier actuellement chargé"):
                    st.session_state.df_current = None
                    st.session_state.data_loaded_id = None
                    st.session_state.current_collection_header = None
                    st.success("✅ Fichier effacé avec succès !")
                    st.rerun()

            # Afficher le statut de chargement
            if st.session_state.df_current is not None:
                st.sidebar.success(f"✅ Fichier chargé : {len(st.session_state.df_current)} lignes")
            else:
                st.sidebar.info("📋 Aucun fichier chargé. Sélectionnez un fichier Excel pour commencer.")

            # Extraire les colonnes si les données sont chargées
            if st.session_state.df_current is not None:
                all_columns = st.session_state.df_current.columns.tolist()
                numerical_columns = [col for col in st.session_state.df_current.columns if pd.api.types.is_numeric_dtype(st.session_state.df_current[col])]
                categorical_columns = [col for col in st.session_state.df_current.columns if pd.api.types.is_object_dtype(st.session_state.df_current[col]) or isinstance(st.session_state.df_current[col].dtype, pd.CategoricalDtype)]

    # Afficher le titre de la collecte chargée ou du fichier
    if st.session_state.get('current_collection_header'):
        st.header(st.session_state.current_collection_header)
    else:
        st.header("Bienvenue sur l'outil d'Analyse de Données")
        st.info("Veuillez sélectionner une source de données dans la barre latérale pour commencer.")

    # --- BOUTONS DE CONTRÔLE ET CONFIGURATION ---

    # Bouton pour actualiser les données
    if st.sidebar.button("🔄 Actualiser les données", help="Recharge les données depuis la source sélectionnée."):
        st.cache_data.clear() # Efface le cache des données
        st.rerun() # Recharge l'application

    # Bouton pour réinitialiser l'application
    if st.sidebar.button("🗑️ Effacer et Recommencer", help="Réinitialise toutes les analyses, l'historique du chat et les données chargées."):
        keys_to_pop = ['analyses', 'gemini_chat_history', 'gemini_chat', 'show_analysis_config',
                'show_chat_ia', 'dataframe_to_export', 'data_loaded_id', 'df_current',
                'current_collection_header', 'generated_python_script', 'generated_python_title',
                'sql_query_manual', 'sql_query_title', 'show_complete_analysis', 'include_complete_analysis_in_reports',
                'pdf_report', 'pdf_filename', 'user_api_key', 'intelligent_analysis', 'selected_analyses']
        for key in keys_to_pop:
            if key in st.session_state:
                st.session_state.pop(key)
        # Réinitialiser les états par défaut
        st.session_state.show_analysis_config = True
        st.session_state.show_chat_ia = True
        st.cache_data.clear()
        st.rerun()

    # --- CONFIGURATION DE L'APPLICATION ---
    st.sidebar.header("⚙️ Configuration de l'application")
    
    # Toggle pour basculer entre le modèle local et Google Gemini
    use_local_model = st.sidebar.toggle(
        "Utiliser le modèle local Gemma",
        value=st.session_state.get('use_local_model', False),
        help="Activez pour utiliser le modèle Gemma local au lieu de Google Gemini"
    )
    st.session_state.use_local_model = use_local_model

    # Mode secours (fallback) : désactivé par défaut
    st.session_state.allow_fallback_ai = st.sidebar.toggle(
        "Autoriser le mode secours (sans IA)",
        value=st.session_state.get("allow_fallback_ai", False),
        help="Si Gemini/OpenAI sont en quota et Gemma indisponible, autorise une requête/script simple généré sans IA.",
    )
    
    # Afficher des informations sur le modèle sélectionné
    if use_local_model:
        st.sidebar.info("ℹ️ Utilisation du modèle Gemma local (plus lent mais sans limite de quota)")
        
        # Vérifier si le modèle Gemma est disponible
        if not os.path.exists(LOCAL_GEMMA_PATH):
            st.sidebar.warning(f"Le modèle Gemma n'a pas été trouvé dans {LOCAL_GEMMA_PATH}")
            if st.sidebar.button("Charger le modèle Gemma"):
                with st.sidebar.status("Chargement du modèle Gemma..."):
                    load_local_gemma()
    else:
        # Section pour configurer la clé API Google Gemini
        with st.sidebar.expander("🔑 Configuration de l'API Google Gemini", expanded=True):
            st.markdown("**Pour utiliser les fonctionnalités IA (Chat, Génération SQL/Python), vous devez configurer votre clé API Google Gemini.**")
            
            user_api_key = st.text_input(
                "Votre clé API Google Gemini",
                value=st.session_state.get('user_api_key', ''),
                type="password",
                help="Obtenez une clé API sur https://ai.google.dev/",
                placeholder="AIzaSyC..."
            )
            
            # Sauvegarder la clé API dans la session
            if user_api_key and user_api_key != st.session_state.get('user_api_key', ''):
                st.session_state.user_api_key = user_api_key
                st.success("Clé API mise à jour !")
            
            # Afficher un message si aucune clé n'est configurée
            if not user_api_key and not google_api_key:
                st.warning("⚠️ La génération par IA nécessite une clé API Google Gemini valide.")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("💾 Sauvegarder", key="save_api_key"):
                if api_key_input and api_key_input.strip():
                    st.session_state['user_api_key'] = api_key_input.strip()
                    # Reconfigurer l'API si la clé a changé
                    try:
                        genai.configure(api_key=st.session_state['user_api_key'])
                        st.success("✅ Clé API sauvegardée et API configurée !")
                    except Exception as e:
                        st.error(f"Erreur lors de la configuration de l'API avec la nouvelle clé: {e}")
                        st.session_state.pop('user_api_key') # Supprimer la clé si la config échoue
                    st.rerun()
                else:
                    st.error("❌ Veuillez entrer une clé API valide.")

        with col2:
            if st.button("🗑️ Effacer", key="clear_api_key"):
                if 'user_api_key' in st.session_state:
                    st.session_state.pop('user_api_key')
                    # Reconfigurer l'API avec la clé système si disponible
                    if google_api_key:
                        try: genai.configure(api_key=google_api_key)
                        except Exception: pass # Ignorer si la clé système échoue aussi
                    else: # Sinon, désactiver l'API si aucune clé n'est disponible
                        genai.configure(api_key=None)
                    st.success("✅ Clé API effacée !")
                    st.rerun()

        # Afficher le statut de la clé API
        if st.session_state.get('user_api_key'):
            st.success("✅ Clé API configurée")
            masked_key = st.session_state['user_api_key'][:4] + "..." + st.session_state['user_api_key'][-4:]
            st.info(f"Clé actuelle : {masked_key}")
        else:
            st.warning("⚠️ Aucune clé API configurée")
            st.info("Les fonctionnalités IA sont désactivées.")

    # --- TÉLÉCHARGEMENT DES COLLECTES ---
    st.sidebar.header("📥 Téléchargement des Collectes")
    
    # Ajouter un expander pour les options de téléchargement
    with st.sidebar.expander("📥 Télécharger les collectes", expanded=False):
        st.markdown("### Télécharger les collectes")
        
        # Case à cocher pour sélectionner/désélectionner toutes les collectes
        all_selected = st.checkbox("Sélectionner toutes les collectes", value=True, key="select_all_collections")
        
        # Cases à cocher pour chaque collecte
        selected_collections = []
        cols = st.columns(2)  # Créer 2 colonnes pour un meilleur affichage
        
        for i, collection_name in enumerate(urls.keys()):
            with cols[i % 2]:  # Alterner entre les colonnes
                if st.checkbox(collection_name, value=all_selected, key=f"collect_{i}"):
                    selected_collections.append(collection_name)
        
        # Boutons de téléchargement
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📥 Télécharger la sélection", type="primary", 
                        help="Télécharge uniquement les collectes sélectionnées dans un fichier Excel"):
                with st.spinner("🔄 Génération du fichier Excel avec les collectes sélectionnées..."):
                    excel_data, filename = download_selected_collections_excel(selected_collections)
                    
                    if excel_data and filename:
                        st.download_button(
                            label="💾 Télécharger le Fichier Excel",
                            data=excel_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_selected_collections"
                        )
        
        with col2:
            if st.button("📥 Télécharger tout", 
                        help="Télécharge toutes les collectes dans un fichier Excel"):
                with st.spinner("🔄 Génération du fichier Excel avec toutes les collectes..."):
                    excel_data, filename = download_all_collections_excel()
                    
                    if excel_data and filename:
                        st.download_button(
                            label="💾 Télécharger tout",
                            data=excel_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_all_collections",
                            help="Télécharge le fichier Excel contenant toutes les collectes"
                        )
                        st.sidebar.success("✅ Fichier Excel prêt ! Cliquez sur le bouton ci-dessus pour le télécharger.")
                    else:
                        st.sidebar.error("❌ Erreur lors de la génération du fichier Excel.")

    # Fermeture du bloc de téléchargement
    
    # --- OPTIONS DE RAPPORT ---
    st.sidebar.header("📄 Options de Rapport")

    # Générer le contenu HTML du rapport seulement si nécessaire
    report_html_content = None
    df_current = st.session_state.get('df_current')

    # Vérifier rapidement si on a du contenu à inclure dans le rapport
    has_content = (df_current is not None and
            (st.session_state.get('analyses') or
            st.session_state.get('gemini_chat_history') or
            st.session_state.get('include_complete_analysis_in_reports', False)))

    if has_content:
        # Configuration des rapports (lazy loading)
        with st.sidebar.expander("⚙️ Configuration des rapports", expanded=False):
            st.markdown("**📊 Contenu du rapport :**")
            include_analyses = st.checkbox("Inclure les analyses", value=True, help="Inclut toutes les analyses effectuées")
            include_chat = st.checkbox("Inclure l'historique du chat IA", value=True, help="Inclut les conversations avec l'IA")
            include_complete_analysis = st.checkbox("Inclure l'analyse complète", value=st.session_state.get('include_complete_analysis_in_reports', False), help="Inclut l'analyse complète personnalisée")
            include_intelligent_analysis = st.checkbox("Inclure l'analyse intelligente", value=st.session_state.get('include_intelligent_analysis_in_reports', True), help="Inclut les résultats de l'analyse intelligente basée sur la signification des variables")

            st.markdown("**🎨 Format du rapport :**")
            report_title = st.text_input("Titre du rapport", value=f"Rapport - {st.session_state.get('current_collection_header', 'Analyse')}", help="Titre personnalisé pour le rapport")
            include_logo = st.checkbox("Inclure le logo", value=True, help="Affiche le logo de l'organisation")
            include_timestamp = st.checkbox("Inclure la date/heure", value=True, help="Affiche la date et heure de génération")

        # Générer le rapport seulement quand nécessaire
        report_html_content = generate_report_html_content()

        # Section d'export HTML
        st.sidebar.markdown("**📥 Export HTML :**")
        st.sidebar.download_button(
            label="📄 Télécharger Rapport HTML",
            data=report_html_content,
            file_name=f"rapport_{st.session_state.get('current_collection_header', 'analyse').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html",
            mime="text/html",
            help="Génère et télécharge un rapport HTML consolidant les analyses, le chat IA et l'analyse complète.",
            key="download_html_report"
        )

        # Section d'export PDF
        st.sidebar.markdown("**📥 Export PDF :**")

        # Bouton de génération PDF
        if st.sidebar.button("🔄 Générer PDF", help="Génère le rapport PDF avant le téléchargement", key="generate_pdf_btn"):
            with st.spinner("Génération du PDF en cours..."):
                success = _generate_and_store_pdf(report_html=report_html_content)
                if success:
                    st.sidebar.success("✅ PDF généré avec succès!")
                else:
                    st.sidebar.error("❌ Erreur lors de la génération du PDF")

        # Afficher le bouton de téléchargement PDF seulement si le PDF a été généré
        pdf_data = st.session_state.get('pdf_report', b'')
        if pdf_data and len(pdf_data) > 0:
            st.sidebar.download_button(
                label="📄 Télécharger Rapport PDF",
                data=pdf_data,
                file_name=st.session_state.get('pdf_filename', f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"),
                mime="application/pdf",
                help="Télécharge le rapport PDF généré.",
                key="download_pdf_report"
            )
        else:
            st.sidebar.info("💡 Cliquez sur '🔄 Générer PDF' pour créer le rapport PDF")

        # Informations sur le rapport (optimisé)
        analyses_count = len(st.session_state.get('analyses', []))
        chat_messages = len(st.session_state.get('gemini_chat_history', []))
        st.sidebar.info(f"📈 {analyses_count} analyses | 💬 {chat_messages} messages")

    else: # Si pas de contenu
        st.sidebar.info("💡 Chargez des données et effectuez des analyses pour générer un rapport.")
        st.sidebar.markdown("**📋 Prérequis :**")
        st.sidebar.markdown("• Données chargées")
        st.sidebar.markdown("• Analyses effectuées")
        st.sidebar.markdown("• Ou conversation IA")

    # --- AFFICHAGE PRINCIPAL DU CONTENU DE L'APPLICATION ---
    # Utiliser le DataFrame chargé pour les analyses
    df_current = st.session_state.get('df_current')

    if df_current is not None:
        # --- AFFICHAGE DES INFORMATIONS SUR LES DONNÉES ---
        st.subheader("📊 Informations sur les données chargées")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Nombre de lignes", len(df_current))
        col2.metric("Nombre de colonnes", len(df_current.columns))

        # Calculer et afficher les données manquantes (optimisé)
        missing_data_count = df_current.isnull().sum().sum()
        col3.metric("Données manquantes", missing_data_count)

        # Compter les colonnes numériques et catégorielles (optimisé)
        numerical_cols_count = len([col for col in df_current.columns if pd.api.types.is_numeric_dtype(df_current[col])])
        categorical_cols_count = len([col for col in df_current.columns if pd.api.types.is_object_dtype(df_current[col]) or isinstance(df_current[col].dtype, pd.CategoricalDtype)])
        col4.metric("Colonnes Catégorielles", categorical_cols_count)

        # Afficher un aperçu des données
        with st.expander("👀 Aperçu des données (5 premières lignes)"):
            st.dataframe(df_current.head(), use_container_width=True)

        # Afficher les colonnes disponibles avec suggestions
        with st.expander("🔍 Voir les colonnes disponibles et suggestions d'analyses"):
            st.write("**Colonnes disponibles :**")
            for i, col in enumerate(df_current.columns):
                col_type = df_current[col].dtype
                unique_count = df_current[col].nunique()
                suggestion = ""
                # Générer des suggestions basées sur le nom de la colonne et son type/cardinalité
                if "commune" in col.lower() or "département" in col.lower():
                    suggestion = "✅ Parfait pour les analyses géographiques"
                elif "satisfait" in col.lower():
                    suggestion = "✅ Idéal pour analyser la satisfaction"
                elif "age" in col.lower() or "an" in col.lower(): # Colonnes numériques potentiellement liées à l'âge
                    suggestion = "📈 Potentiellement numérique (âge) - statistiques/histogrammes"
                elif col_type in ['int64', 'float64']:
                    suggestion = "📊 Colonne numérique - statistiques descriptives"
                elif unique_count <= 15: # Si peu de valeurs uniques, suggérer graphiques catégorie
                    suggestion = "📈 Faible cardinalité - graphiques en barres/circulaires"
                else:
                    suggestion = "📋 Colonne catégorielle"

                st.write(f"• **** ({col_type}) - {unique_count} uniques - {suggestion}")

            st.write("") # Ligne vide pour l'espacement
            st.write("**💡 Suggestions d'analyses populaires :**")
            # Suggestions basées sur les colonnes détectées
            if categorical_cols_count > 0:
                st.write("- Analyse de la répartition des variables catégorielles (graphiques).")
            if numerical_cols_count > 0:
                st.write("- Statistiques descriptives pour les variables numériques.")
            if numerical_cols_count > 1:
                st.write("- Matrice de corrélation entre variables numériques.")
            if find_first_matching_column(df_current, ['commune', 'ville', 'localité']):
                st.write("- Analyse géographique par zone (commune/ville).")
            if find_first_matching_column(df_current, ['sexe', 'genre']) and find_first_matching_column(df_current, ['âge', 'age']):
                st.write("- Analyse démographique (sexe, âge).")
            if find_first_matching_column(df_current, ['satisfaction', 'satisfait']):
                st.write("- Analyse de satisfaction.")
                if find_first_matching_column(df_current, ['département', 'région', 'province']):
                    st.write(" - Satisfaction par département/région.")

    else: # Si pas de données chargées
        st.info("Veuillez charger des données via la barre latérale pour commencer l'analyse.")

    # --- SECTION DE CONFIGURATION DES ANALYSES ---
    st.subheader("🔧 Configuration des Analyses")

    # Boutons pour afficher/masquer la configuration
    if st.session_state.show_analysis_config:
        if st.button("❌ Masquer la configuration"):
            st.session_state.show_analysis_config = False
            st.rerun()
    else:
        if st.button("⚙️ Afficher la configuration"):
            st.session_state.show_analysis_config = True
            st.rerun()

    if st.session_state.show_analysis_config and df_current is not None:
        st.markdown("---") # Séparateur visuel

        # --- AJOUT D'UN TABLEAU AGRÉGÉ ---
        st.markdown("### 📋 Tableau agrégé")
        agg_title_input = st.text_input("Titre du tableau agrégé:", value="Tableau agrégé", key="agg_title_input", help="Entrez un titre personnalisé.")
        group_by_cols = st.multiselect("Colonnes de groupement:", all_columns, help="Sélectionnez les colonnes pour grouper les données.")
        agg_cols_selection = st.multiselect("Colonnes à agréger:", numerical_columns, help="Sélectionnez les colonnes numériques à agréger.")
        agg_functions_selection = st.multiselect("Fonctions d'agrégation:", ['sum', 'mean', 'count', 'min', 'max', 'std'], default=['sum'], help="Choisissez les fonctions d'agrégation.")

        if st.button("➕ Ajouter tableau agrégé", key="add_agg_table"):
            if group_by_cols and agg_cols_selection and agg_functions_selection:
                try:
                    # Création du dictionnaire d'agrégation
                    agg_dict = {col: agg_functions_selection for col in agg_cols_selection}
                    # Exécution de l'agrégation
                    result_df = df_current.groupby(group_by_cols).agg(agg_dict)

                    # Aplatir le MultiIndex des colonnes si nécessaire
                    if isinstance(result_df.columns, pd.MultiIndex):
                        result_df.columns = ['_'.join(map(str, col)).strip() for col in result_df.columns.values]

                    result_df = result_df.reset_index()

                    # Ajouter l'analyse à la liste
                    analysis_id = f"agg_{len(st.session_state.analyses)}"
                    st.session_state.analyses.append({
                        'type': 'aggregated_table',
                        'result': result_df,
                        'id': analysis_id,
                        'executed_params': {
                            'group_by_columns': group_by_cols,
                            'aggregation_columns': agg_cols_selection,
                            'aggregation_functions': agg_functions_selection,
                            'title': agg_title_input if agg_title_input.strip() else f"Tableau agrégé par {', '.join(group_by_cols)}"
                        }
                    })
                    st.success("Tableau agrégé ajouté!")
                    st.rerun() # Rafraîchir pour montrer la nouvelle analyse
                except Exception as e:
                    st.error(f"Erreur lors de la création du tableau agrégé: {e}")
            else:
                st.warning("Veuillez sélectionner au moins une colonne de groupement, une colonne à agréger et une fonction d'agrégation.")

        # --- AJOUT DE STATISTIQUES DESCRIPTIVES ---
        st.markdown("### 📈 Statistiques descriptives")
        desc_title_input = st.text_input("Titre des statistiques:", value="Statistiques descriptives", key="desc_title_input", help="Entrez un titre personnalisé.")
        desc_cols_selection = st.multiselect("Colonnes pour statistiques:", numerical_columns, help="Sélectionnez les colonnes numériques.")

        if st.button("➕ Ajouter statistiques descriptives", key="add_desc_stats"):
            if desc_cols_selection:
                try:
                    result_stats = df_current[desc_cols_selection].describe()
                    analysis_id = f"desc_{len(analyses)}"
                    st.session_state.analyses.append({
                        'type': 'descriptive_stats',
                        'result': result_stats,
                        'id': analysis_id,
                        'executed_params': {
                            'selected_columns': desc_cols_selection,
                            'title': desc_title_input if desc_title_input.strip() else f"Statistiques descriptives de {', '.join(desc_cols_selection)}"
                        }
                    })
                    st.success("Statistiques descriptives ajoutées!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erreur lors de la création des statistiques descriptives: {e}")
            else:
                st.warning("Veuillez sélectionner au moins une colonne numérique.")

        # --- AJOUT DE GRAPHIQUE ---
        st.markdown("### 📊 Graphique")
        chart_title_input = st.text_input("Titre du graphique:", value="Graphique", key="chart_title_input", help="Entrez un titre personnalisé.")
        chart_type_selection = st.selectbox("Type de graphique:", ['bar', 'line', 'scatter', 'histogram', 'box', 'pie'], help="Choisissez le type de visualisation.")

        # Options d'agrégation et de personnalisation
        agg_functions = {
            'Aucune': None,
            'Somme': 'sum',
            'Moyenne': 'mean',
            'Médiane': 'median',
            'Minimum': 'min',
            'Maximum': 'max',
            'Décompte': 'count',
            'Écart-type': 'std'
        }
        
        # Palettes de couleurs prédéfinies
        color_palettes = {
            'Défaut': 'Plotly',
            'Vibrant': 'Vivid',
            'Pastel': 'Pastel',
            'Bleus': 'Blues',
            'Rouges': 'Reds',
            'Verts': 'Greens',
            'Arc-en-ciel': 'Rainbow'
        }

        # Configuration spécifique selon le type de graphique
        x_col_selected, y_col_selected, hist_col_selected, box_col_selected, pie_col_selected = None, None, None, None, None

        if chart_type_selection in ['bar', 'line', 'scatter']:
            col1, col2, col3 = st.columns(3)
            
            with col1:
                x_col_selected = st.selectbox("Colonne X:", all_columns, help="Sélectionnez la colonne pour l'axe X.")
                y_col_selected = st.selectbox("Colonne Y:", numerical_columns, help="Sélectionnez la colonne numérique pour l'axe Y.")
                
            with col2:
                agg_func = st.selectbox("Fonction d'agrégation:", list(agg_functions.keys()), 
                                     help="Sélectionnez une fonction d'agrégation pour les données")
                color_by = st.selectbox("Couleur par:", ['Aucun'] + categorical_columns, 
                                     help="Choisissez une colonne pour le dégradé de couleurs")
                
            with col3:
                palette = st.selectbox("Palette de couleurs:", list(color_palettes.keys()))
                
                # Options avancées
                with st.expander("Options avancées"):
                    opacity = st.slider("Opacité:", 0.1, 1.0, 0.8, 0.1, 
                                     help="Ajustez l'opacité des éléments du graphique")
                    barmode = st.selectbox("Mode des barres:", ['group', 'stack', 'relative', 'overlay']) \
                        if chart_type_selection == 'bar' else None

            if st.button(f"➕ Ajouter graphique {chart_type_selection.capitalize()}", key=f"add__chart"):
                if x_col_selected and y_col_selected:
                    try:
                        # Préparation des données avec agrégation si nécessaire
                        df_to_plot = df_current.copy()
                        
                        # Appliquer l'agrégation si nécessaire
                        if agg_func and agg_func != 'Aucune':
                            agg_dict = {y_col_selected: agg_functions[agg_func]}
                            if x_col_selected in categorical_columns:
                                df_to_plot = df_to_plot.groupby(x_col_selected).agg(agg_dict).reset_index()
                            else:
                                st.warning("L'agrégation nécessite une colonne X catégorielle")
                                df_to_plot = df_to_plot[[x_col_selected, y_col_selected]]
                        
                        # Préparation des arguments du graphique
                        chart_kwargs = {
                            'data_frame': df_to_plot,
                            'x': x_col_selected,
                            'y': y_col_selected,
                            'title': chart_title_input if chart_title_input.strip() else f"{chart_type_selection.capitalize()} : {y_col_selected} par {x_col_selected}",
                            'template': 'plotly_white',
                            'color_discrete_sequence': getattr(px.colors.qualitative, color_palettes[palette], None) if color_palettes[palette] != 'Plotly' else None,
                            'opacity': opacity,
                        }
                        
                        # Ajouter la couleur si spécifiée
                        if color_by != 'Aucun':
                            chart_kwargs['color'] = color_by
                            chart_kwargs['color_continuous_scale'] = color_palettes[palette] if color_palettes[palette] != 'Plotly' else None
                            
                        # Ajouter le mode des barres si pertinent
                        if chart_type_selection == 'bar' and barmode:
                            chart_kwargs['barmode'] = barmode
                        # Adapter les arguments selon le type de graphique
                        if chart_type_selection == 'bar' and 'color' not in chart_kwargs:
                            chart_kwargs['color'] = y_col_selected  # Colorer par défaut par Y pour les barres
                        if chart_type_selection == 'scatter' and 'color' not in chart_kwargs:
                            chart_kwargs['color'] = x_col_selected  # Colorer par défaut par X pour les nuages de points

                        # Appel de la fonction wrapper sécurisée - passer d'abord le type de graphique, puis les données
                        # Créer une copie des kwargs et retirer data_frame pour l'utiliser comme premier argument
                        call_kwargs = chart_kwargs.copy()
                        data = call_kwargs.pop('data_frame', None)
                        fig = safe_plotly_call(chart_type_selection, data, **call_kwargs)

                        if isinstance(fig, str) and "Erreur" in fig: # Si la fonction wrapper a retourné une erreur
                            st.error(f"Erreur lors de la génération du graphique: {fig}")
                        elif fig is not None:
                            analysis_id = f"chart_{len(analyses)}"
                            analyses.append({
                                'type': 'graph',
                                'result': fig,
                                'id': analysis_id,
                                'executed_params': {
                                    'chart_type': chart_type_selection,
                                    'x_column': x_col_selected,
                                    'y_column': y_col_selected,
                                    'title': chart_title_input if chart_title_input.strip() else f"Graphique : {x_col_selected} par {y_col_selected}"
                                }
                            })
                            st.success("Graphique ajouté!")
                            st.rerun()
                        else: # Si fig est None sans erreur spécifique
                            st.warning("Le graphique n'a pas pu être généré.")

                    except Exception as e:
                        st.error(f"Erreur lors de l'ajout du graphique: {e}")
                else:
                    st.warning("Veuillez sélectionner les colonnes X et Y.")

        elif chart_type_selection == 'histogram':
            hist_col_selected = st.selectbox("Colonne pour histogramme:", numerical_columns, help="Sélectionnez la colonne numérique.")
            if st.button("➕ Ajouter histogramme", key="add_histogram_chart"):
                if hist_col_selected:
                    try:
                        fig = safe_plotly_call('histogram', df_current, x=hist_col_selected, title=chart_title_input if chart_title_input.strip() else f"Histogramme de ", nbins=30, template='plotly_white')
                        if isinstance(fig, str) and "Erreur" in fig: st.error(f"Erreur lors de la génération de l'histogramme: {fig}")
                        elif fig is not None:
                            analysis_id = f"hist_{len(analyses)}"
                            analyses.append({
                                'type': 'graph',
                                'result': fig,
                                'id': analysis_id,
                                'executed_params': {
                                    'chart_type': 'histogram',
                                    'x_column': hist_col_selected,
                                    'title': chart_title_input if chart_title_input.strip() else f"Histogramme de "
                                }
                            })
                            st.success("Histogramme ajouté!")
                            st.rerun()
                        else: st.warning("L'histogramme n'a pas pu être généré.")
                    except Exception as e: st.error(f"Erreur lors de l'ajout de l'histogramme: {e}")
                else: st.warning("Veuillez sélectionner une colonne numérique.")

        elif chart_type_selection == 'box':
            box_col_selected = st.selectbox("Colonne pour boîte à moustaches:", numerical_columns, help="Sélectionnez la colonne numérique.")
            if st.button("➕ Ajouter boîte à moustaches", key="add_box_chart"):
                if box_col_selected:
                    try:
                        fig = safe_plotly_call('box', df_current, y=box_col_selected, title=chart_title_input if chart_title_input.strip() else f"Boîte à moustaches de ", template='plotly_white')
                        if isinstance(fig, str) and "Erreur" in fig: st.error(f"Erreur lors de la génération de la boîte à moustaches: {fig}")
                        elif fig is not None:
                            analysis_id = f"box_{len(analyses)}"
                            analyses.append({
                                'type': 'graph',
                                'result': fig,
                                'id': analysis_id,
                                'executed_params': {
                                    'chart_type': 'box',
                                    'y_column': box_col_selected,
                                    'title': chart_title_input if chart_title_input.strip() else f"Boîte à moustaches de "
                                }
                            })
                            st.success("Boîte à moustaches ajoutée!")
                            st.rerun()
                        else: st.warning("La boîte à moustaches n'a pas pu être générée.")
                    except Exception as e: st.error(f"Erreur lors de l'ajout de la boîte à moustaches: {e}")
                else: st.warning("Veuillez sélectionner une colonne numérique.")

        elif chart_type_selection == 'pie':
            pie_col_selected = st.selectbox("Colonne pour graphique circulaire:", categorical_columns, help="Sélectionnez la colonne catégorielle.")
            if st.button("➕ Ajouter graphique circulaire", key="add_pie_chart"):
                if pie_col_selected:
                    try:
                        # value_counts() peut retourner un MultiIndex si la colonne est déjà catégorielle avec MultiIndex
                        pie_data = df_current[pie_col_selected].value_counts()
                        if isinstance(pie_data.index, pd.MultiIndex): # Gérer MultiIndex
                            try: pie_data = pie_data.reset_index(level=list(range(pie_data.index.nlevels))).groupby(pie_data.index.names[0]).sum().iloc[:,0]
                            except Exception: st.warning(f"Impossible de simplifier l'index MultiIndex pour la colonne '{pie_col_selected}'."); pie_data = None
                        if pie_data is not None:
                            fig = safe_plotly_call('pie', pie_data, names=pie_data.index, values=pie_data.values, title=chart_title_input if chart_title_input.strip() else f"Répartition de ", template='plotly_white')
                        if isinstance(fig, str) and "Erreur" in fig: st.error(f"Erreur lors de la génération du graphique circulaire: {fig}")
                        elif fig is not None:
                            analysis_id = f"pie_{len(analyses)}"
                            analyses.append({
                                'type': 'graph',
                                'result': fig,
                                'id': analysis_id,
                                'executed_params': {
                                    'chart_type': 'pie',
                                    'names_column': pie_col_selected,
                                    'title': chart_title_input if chart_title_input.strip() else f"Répartition de "
                                }
                            })
                            st.success("Graphique circulaire ajouté!")
                            st.rerun()
                        else: st.warning("Le graphique circulaire n'a pas pu être généré.")
                    except Exception as e: st.error(f"Erreur lors de l'ajout du graphique circulaire: {e}")
                else: st.warning("Veuillez sélectionner une colonne catégorielle.")

        # --- SECTION PYTHON ET SQL VIA IA ---
        st.markdown("---")
        st.subheader("🤖 Génération d'analyses par IA")

        python_tab, sql_tab_ai, sql_tab_manual = st.tabs(["🐍 Génération Python", "🤖 Génération IA SQL", "✍️ SQL Manuel"])

        # --- ONGLET GÉNÉRATION PYTHON ---
        with python_tab:
            st.markdown("**Décrivez l'analyse que vous souhaitez réaliser en Python.**")
            with st.expander("💡 Exemples de demandes"):
                st.markdown("""
                - "Créer un graphique en barres du nombre de répondants par commune."
                - "Faire un tableau croisé entre sexe et niveau d'éducation."
                - "Calculer les statistiques descriptives de l'âge par département."
                - "Créer un nuage de points entre âge et revenu."
                - "Faire une analyse de corrélation entre toutes les variables numériques."
                - "Créer un histogramme de la distribution des âges."
                - "Générer un rapport de qualité des données avec les valeurs manquantes."
                """)

            python_request_input = st.text_area(
                "Décrivez votre demande ici :", height=100, key="python_request_input",
                placeholder="Exemple: Créez un graphique en barres montrant le nombre de participants par département, coloré par sexe."
            )

            # Étape 1: Proposer les colonnes à utiliser selon la requête
            cols = df_current.columns.tolist()
            req_lower = (python_request_input or "").lower()
            if st.button("🔎 Proposer les colonnes utilisées", key="python_suggest_cols"):
                suggested = []
                for c in cols:
                    try:
                        if c and c.strip() and c.lower() in req_lower:
                            suggested.append(c)
                    except Exception:
                        continue
                # Si rien trouvé, suggérer 1 cat + 1 num par défaut
                default_cat = (categorical_columns[0] if categorical_columns else (cols[0] if cols else None))
                default_num = (numerical_columns[0] if numerical_columns else None)
                if not suggested:
                    for v in [default_cat, default_num]:
                        if v and v not in suggested:
                            suggested.append(v)
                st.session_state.python_detected_columns = {
                    'all': suggested,
                    'categorical': [c for c in suggested if c in (categorical_columns or [])],
                    'numerical': [c for c in suggested if c in (numerical_columns or [])]
                }

            # Interface de validation/ajustement des colonnes proposées
            if st.session_state.get('python_detected_columns'):
                st.markdown("**Colonnes proposées par l'IA (modifiable avant génération):**")
                with st.form(key="python_columns_validation_form", clear_on_submit=False):
                    sel_cat = st.multiselect(
                        "Colonnes catégorielles à utiliser",
                        options=categorical_columns,
                        default=st.session_state.python_detected_columns.get('categorical', [])
                    )
                    sel_num = st.multiselect(
                        "Colonnes numériques à utiliser",
                        options=numerical_columns,
                        default=st.session_state.python_detected_columns.get('numerical', [])
                    )
                    submitted_cols = st.form_submit_button("✅ Valider la sélection de colonnes")
                    if submitted_cols:
                        st.session_state.python_selected_columns = {
                            'categorical': sel_cat,
                            'numerical': sel_num
                        }
                        # Nettoyer tout script précédemment généré pour éviter l'ambiguïté
                        st.session_state.generated_python_script = None
                        st.session_state.generated_python_title = None
                        st.success("Sélection de colonnes validée. Vous pouvez maintenant générer le script.")
                        st.rerun()

            if st.button("🐍 Générer le script Python", key="generate_python_action", type="primary"):
                if python_request_input.strip():
                    # Générer le script Python
                    # Utiliser les colonnes validées par l'utilisateur si disponibles
                    sel = st.session_state.get('python_selected_columns', {})
                    use_num = sel.get('numerical', numerical_columns)
                    use_cat = sel.get('categorical', categorical_columns)
                    generated_script, generated_title = generate_python_script_with_ai(
                        python_request_input, df_current, all_columns, use_num, use_cat
                    )
                    if generated_script:
                        # Ajouter un en-tête informatif sur les colonnes réellement utilisées
                        header_comment = (
                            f"# Colonnes catégorielles sélectionnées: {use_cat}\n"
                            f"# Colonnes numériques sélectionnées: {use_num}\n"
                        )
                        st.session_state.generated_python_script = header_comment + ("\n" if not generated_script.startswith("#") else "") + generated_script
                        st.session_state.generated_python_title = generated_title
                        st.success("Script Python généré !")
                        st.rerun() # Rafraîchir pour afficher le script généré
                    else:
                        st.error("Échec de la génération du script Python. Veuillez reformuler votre demande.")
                else:
                    st.warning("Veuillez décrire votre demande.")

            # Afficher le script généré et le bouton d'exécution
            if st.session_state.generated_python_script:
                st.markdown("**Script Python généré :**")
                edited_python_script = st.text_area(
                    "Vous pouvez modifier le script avant exécution :",
                    value=st.session_state.generated_python_script,
                    height=300,
                    key="python_script_editor",
                )

                # Afficher le champ pour le titre de l'analyse
                analysis_title_input = st.text_input(
                    "Titre de l'analyse Python:",
                    value=st.session_state.generated_python_title if st.session_state.generated_python_title else "Résultat du script Python",
                    key="python_analysis_title_input"
                )

                if st.button("▶️ Exécuter ce script", key="execute_python_action"):
                    # Exécuter le script généré
                    python_result = execute_python_script(edited_python_script, df_current)
                    if python_result is not None:
                        # Afficher immédiatement le résultat dans l'onglet
                        display_title = analysis_title_input if analysis_title_input.strip() else (st.session_state.generated_python_title or "Résultat du script Python")
                        st.subheader(display_title)
                        try:
                            if hasattr(python_result, 'to_plotly_json'):
                                st.plotly_chart(python_result, use_container_width=True)
                            elif isinstance(python_result, (pd.DataFrame, pd.Series)):
                                show_df(python_result, use_container_width=True)
                            elif isinstance(python_result, dict):
                                for k, v in python_result.items():
                                    st.markdown(f"**{k}**")
                                    if hasattr(v, 'to_plotly_json'):
                                        st.plotly_chart(v, use_container_width=True)
                                    elif isinstance(v, (pd.DataFrame, pd.Series)):
                                        show_df(v, use_container_width=True)
                                    else:
                                        st.write(v)
                            else:
                                st.write(python_result)
                        except Exception as _e_disp:
                            st.warning(f"Résultat exécuté mais affichage non standard: {_e_disp}")

                        # Ajouter aussi aux analyses pour le reporting
                        analysis_id = f"py_exec_{len(analyses)}"
                        analyses.append({
                            'type': 'python_script',
                            'result': python_result,
                            'id': analysis_id,
                            'executed_params': {
                                'script': edited_python_script,
                                'title': display_title
                            }
                        })
                        st.success("Script Python exécuté. Résultat affiché et ajouté aux analyses.")
                    # else: L'erreur est déjà affichée dans execute_python_script

                # Bouton pour effacer le script généré
                if st.button("🗑️ Effacer le script", key="clear_python_script_action"):
                    st.session_state.generated_python_script = None
                    st.session_state.generated_python_title = None
                    st.rerun()

        # --- ONGLET GÉNÉRATION SQL VIA IA ---
        with sql_tab_ai:
            st.markdown("**Décrivez la requête SQL que vous souhaitez exécuter.**")
            with st.expander("💡 Exemples de requêtes"):
                st.markdown("""
                - "Compter le nombre de répondants par commune."
                - "Calculer la moyenne d'âge par sexe."
                - "Afficher les 10 communes avec le plus d'habitants."
                - "Filtrer les données pour les personnes satisfaites et de plus de 30 ans."
                - "Calculer la somme des 'revenus' par département."
                """)

            nl_query_input = st.text_area(
                "Décrivez votre requête SQL ici :", height=100, key="nl_query_input_ai",
                placeholder="Exemple: Compte le nombre de répondants par département, puis affiche les 3 départements avec le plus de répondants."
            )

            # Aide: proposer/sélectionner les colonnes à utiliser pour guider l'IA
            sql_cols = df_current.columns.tolist()

            # Appliquer une mise à jour "pending" AVANT l'instanciation du widget (contrainte Streamlit)
            if st.session_state.get("sql_selected_columns_pending") is not None:
                pending = st.session_state.pop("sql_selected_columns_pending", None) or []
                st.session_state.sql_detected_columns = pending
                st.session_state.sql_selected_columns = pending
                st.session_state["sql_selected_columns_widget"] = pending

            if st.button("🔎 Proposer les colonnes utilisées (SQL)", key="sql_suggest_cols"):
                req_lower_sql = (nl_query_input or "").lower()
                suggested_sql = []
                for c in sql_cols:
                    try:
                        if c and c.strip() and c.lower() in req_lower_sql:
                            suggested_sql.append(c)
                    except Exception:
                        continue
                # Fallback: proposer 1-2 colonnes fréquentes
                if not suggested_sql:
                    for fallback_key in ["commune", "département", "departement", "région", "region", "_submission_time"]:
                        for c in sql_cols:
                            if fallback_key in (c or "").lower():
                                suggested_sql.append(c)
                                break
                        if suggested_sql:
                            break
                st.session_state.sql_detected_columns = suggested_sql

            if st.session_state.get("sql_detected_columns") is not None:
                st.markdown("**Colonnes à utiliser pour la génération SQL (modifiable):**")
                st.session_state.sql_selected_columns = st.multiselect(
                    "Colonnes autorisées",
                    options=sql_cols,
                    default=st.session_state.get("sql_selected_columns", st.session_state.get("sql_detected_columns", [])),
                    key="sql_selected_columns_widget",
                )

            if st.button("🤖 Générer la requête SQL", key="generate_sql_ai_action", type="primary"):
                if nl_query_input.strip():
                    # Utiliser les colonnes nettoyées disponibles pour l'IA
                    selected_cols = st.session_state.get("sql_selected_columns") or st.session_state.get("sql_selected_columns_widget") or []
                    current_data_cols = selected_cols if selected_cols else df_current.columns.tolist()
                    generated_sql, generated_title = generate_sql_with_ai(
                        nl_query_input, all_columns, numerical_columns, categorical_columns, current_data_cols
                    )
                    if generated_sql:
                        # Colonnes réellement utilisées dans la requête => refléter dans la sélection
                        used_cols = extract_columns_from_sql(generated_sql, df_current.columns.tolist())
                        if used_cols:
                            # Ne pas modifier le state du widget après instanciation; appliquer au prochain rerun
                            st.session_state.sql_selected_columns_pending = used_cols

                        # Si l'utilisateur a restreint les colonnes, refuser une requête hors-scope
                        if selected_cols:
                            bad = [c for c in used_cols if c not in selected_cols]
                            if bad:
                                st.error(
                                    "La requête générée utilise des colonnes hors sélection : "
                                    + ", ".join(f"`{c}`" for c in bad)
                                )
                                st.error("Ajustez la sélection de colonnes ou régénérez la requête.")
                                st.session_state.sql_query_manual = ""
                                st.session_state["sql_query_title"] = ""
                                st.stop()

                        st.session_state.sql_query_manual = generated_sql # Stocker la requête générée
                        st.session_state["sql_query_title"] = generated_title # Stocker le titre généré
                        st.success("Requête SQL générée !")
                        st.rerun() # Rafraîchir pour afficher la requête générée
                    else:
                        st.error("Échec de la génération de la requête SQL. Veuillez reformuler votre demande.")
                else:
                    st.warning("Veuillez décrire votre requête SQL.")

            # Afficher la requête SQL générée et le bouton d'exécution
            if st.session_state.sql_query_manual:
                st.markdown("**Requête SQL générée :**")
                edited_sql_query = st.text_area(
                    "Vous pouvez modifier la requête avant exécution :",
                    value=st.session_state.sql_query_manual,
                    height=220,
                    key="sql_ai_query_editor",
                )

                # Champ pour le titre de l'analyse
                sql_title_input = st.text_input(
                    "Titre de l'analyse SQL:",
                    value=st.session_state.get("sql_query_title", "Résultat de la requête SQL IA"),
                    key="sql_analysis_title_input_ai"
                )

                if st.button("▶️ Exécuter cette requête SQL", key="execute_sql_ai_action"):
                    # Exécuter la requête SQL générée
                    sql_result = execute_sql_query(edited_sql_query, df_current)
                    if sql_result is not None:
                        # Ajouter le résultat aux analyses
                        analysis_id = f"sql_ai_{len(analyses)}"
                        analyses.append({
                            'type': 'sql_query',
                            'result': sql_result,
                            'id': analysis_id,
                            'executed_params': {
                                'query': edited_sql_query,
                                'title': sql_title_input if sql_title_input.strip() else "Résultat de la requête SQL IA"
                            }
                        })
                        st.success("Requête SQL exécutée et ajoutée aux analyses !")
                        # Nettoyer les variables de requête générée
                        st.session_state.sql_query_manual = ''
                        st.session_state["sql_query_title"] = ''
                        st.rerun()

                # Bouton pour effacer la requête générée
                if st.button("🗑️ Effacer la requête générée", key="clear_sql_ai_action"):
                    st.session_state.sql_query_manual = ''
                    st.session_state["sql_query_title"] = ''
                    st.rerun()

        # --- ONGLET SQL MANUEL ---
        with sql_tab_manual:
            st.markdown("**Écrivez votre requête SQL manuellement.**")
            with st.expander("💡 Aide SQL"):
                st.markdown("""
                - Utilisez `data` comme nom de table.
                - Encadrez les noms de colonnes avec des guillemets doubles : `"nom_colonne"`.
                - Exemples : `SELECT * FROM data LIMIT 10`, `SELECT "q3_commune", COUNT(*) FROM data GROUP BY "q3_commune"`
                """)

            sql_query_manual_input = st.text_area(
                "Votre requête SQL :", height=100, key="sql_query_manual_input",
                placeholder="SELECT * FROM data LIMIT 10"
            )

            sql_manual_title_input = st.text_input(
                "Titre de l'analyse SQL:", value="Résultat de la requête SQL manuelle", key="sql_manual_title_input"
            )

            if st.button("➕ Exécuter ma requête SQL", key="execute_sql_manual_action"):
                if sql_query_manual_input.strip():
                    # Exécuter la requête SQL manuelle
                    sql_result = execute_sql_query(sql_query_manual_input, df_current)
                    if sql_result is not None:
                        # Ajouter le résultat aux analyses
                        analysis_id = f"sql_manual_{len(analyses)}"
                        analyses.append({
                            'type': 'sql_query',
                            'result': sql_result,
                            'id': analysis_id,
                            'executed_params': {
                                'query': sql_query_manual_input,
                                'title': sql_manual_title_input if sql_manual_title_input.strip() else "Résultat de la requête SQL manuelle"
                            }
                        })
                        st.success("Requête SQL exécutée et ajoutée aux analyses !")
                        st.rerun()
                else:
                    st.warning("Veuillez entrer une requête SQL.")

    # --- AFFICHAGE DES ANALYSES EFFECTUÉES ---
    if analyses:
        st.subheader("📊 Résultats des analyses")

        # Afficher les cartes d'insights pour un aperçu rapide
        insight_cards_list = []
        for analysis in analyses:
            if analysis.get('result') is not None:
                icon, color, titre, resume = get_insight_card(analysis)
                insight_cards_list.append({'icon': icon, 'color': color, 'titre': titre, 'resume': resume, 'analysis': analysis})

        if insight_cards_list:
            # Afficher les cartes en colonnes
            num_cols = min(3, len(insight_cards_list))
            cols = st.columns(num_cols)
            for i, card in enumerate(insight_cards_list):
                with cols[i % num_cols]:
                    st.markdown(f"""
                    <div style="background: white; padding: 1rem; border-radius: 10px; border-left: 5px solid {card['color']}; margin-bottom: 1rem; height: 150px; display: flex; flex-direction: column; justify-content: space-between;">
                        <div style="font-size: 2rem; margin-bottom: 0.5rem;">{card['icon']}</div>
                        <div style="font-weight: bold; font-size: 1rem; margin-bottom: 0.5rem; color: #2c3e50; overflow: hidden; text-overflow: ellipsis; display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical;">{card['titre']}</div>
                        <div style="font-size: 0.9rem; color: #666; overflow: hidden; text-overflow: ellipsis; display: -webkit-box; -webkit-line-clamp: 3; -webkit-box-orient: vertical;">{card['resume']}</div>
                    </div>
                    """, unsafe_allow_html=True)

        st.markdown("---") # Séparateur
        st.subheader("📋 Détails des analyses")

        # Afficher chaque analyse dans un expander
        for i, analysis in enumerate(analyses):
            if analysis.get('result') is not None:
                params = analysis.get('executed_params', {})
                # Réutiliser le titre déjà généré ou formaté
                _, _, titre, _ = get_insight_card(analysis)

                with st.expander(f"{i+1}. {titre} ({analysis['type'].replace('_', ' ').title()})"):
                    # Section pour modifier l'interprétation IA
                    if analysis['type'] in ['aggregated_table', 'descriptive_stats']:
                        st.markdown("### 📝 Modification de l'interprétation IA")
                        
                        # Récupérer l'interprétation actuelle
                        current_interpretation = analysis.get('modified_interpretation', '')
                        has_modified_interpretation = bool(current_interpretation)
                        
                        if not current_interpretation:
                            # Essayer de générer une interprétation si elle n'existe pas
                            try:
                                if current_api_key and analysis['result'] is not None:
                                    model_chat_report = genai.GenerativeModel('gemini-1.5-flash-latest')
                                    if analysis['type'] == 'aggregated_table' and isinstance(analysis['result'], pd.DataFrame):
                                        df_summary_head = analysis['result'].head(5).to_html(index=False, classes='dataframe', border=0)
                                        interpretation_prompt = f"""Je dispose d'un tableau de données agrégées. Voici les premières lignes:

{df_summary_head}

Paramètres utilisés : {params}.
Interprète brièvement ce tableau de manière concise et professionnelle. Commence directement par l'interprétation."""
                                    elif analysis['type'] == 'descriptive_stats' and isinstance(analysis['result'], pd.DataFrame):
                                        df_summary_stats = analysis['result'].to_html(index=True, classes='dataframe', border=0)
                                        interpretation_prompt = f"""J'ai les statistiques descriptives suivantes pour une ou plusieurs colonnes :

{df_summary_stats}

Analyse des colonnes: {params.get('selected_columns', [])}.
Interprète ces statistiques de manière concise. Commence directement par l'interprétation."""
                                    
                                    if interpretation_prompt:
                                        genai.configure(api_key=current_api_key)
                                        response_report_ia = model_chat_report.generate_content(interpretation_prompt)
                                        current_interpretation = response_report_ia.text.strip()
                            except Exception as e:
                                current_interpretation = f"Erreur lors de la génération de l'interprétation IA: {e}"
                        
                        # Afficher le statut de l'interprétation
                        if has_modified_interpretation:
                            st.info("📝 **Interprétation modifiée** - Cette interprétation sera utilisée dans les rapports")
                        else:
                            st.info("🤖 **Interprétation IA générée automatiquement**")
                        
                        # Zone de texte pour modifier l'interprétation
                        # Créer une clé unique basée sur l'ID de l'analyse, le type, le titre et un compteur
                        if 'interpretation_counter' not in st.session_state:
                            st.session_state.interpretation_counter = 0
                        st.session_state.interpretation_counter += 1
                        
                        # Créer une clé unique avec un préfixe, l'index de l'analyse et un compteur
                        unique_key = f"interpretation_edit_{analysis['id']}_{analysis.get('type', '')}_{st.session_state.interpretation_counter}"
                        # Nettoyer la clé pour éviter les caractères spéciaux
                        unique_key = re.sub(r'[^a-zA-Z0-9_]', '_', unique_key)
                        
                        # Créer un conteneur avec une clé unique pour isoler ce widget
                        with st.container():
                            modified_interpretation = st.text_area(
                                "Modifiez l'interprétation IA si nécessaire :",
                                value=current_interpretation,
                                height=150,
                                key=unique_key,
                                help="Vous pouvez modifier l'interprétation générée par l'IA avant de générer le rapport final."
                            )
                        
                        # Boutons pour gérer l'interprétation
                        col1, col2 = st.columns(2)
                        with col1:
                            # Utiliser le même compteur que pour la zone de texte pour garantir l'unicité
                            save_button_key = f"save_interpretation_{analysis['id']}_{st.session_state.interpretation_counter}"
                            if st.button("💾 Sauvegarder l'interprétation", key=save_button_key):
                                # Mettre à jour l'analyse dans st.session_state
                                analyses = st.session_state.get('analyses', [])
                                for idx, session_analysis in enumerate(analyses):
                                    if session_analysis.get('id') == analysis['id']:
                                        analyses[idx]['modified_interpretation'] = modified_interpretation
                                        st.success("Interprétation sauvegardée avec succès!")
                                        st.rerun()
                                st.success("✅ Interprétation sauvegardée !")
                                st.rerun()
                        
                        with col2:
                            # Utiliser le même compteur que pour les autres widgets pour garantir l'unicité
                            regenerate_button_key = f"regenerate_interpretation_{analysis['id']}_{st.session_state.interpretation_counter}"
                            if st.button("🔄 Régénérer l'interprétation", key=regenerate_button_key):
                                try:
                                    if current_api_key and analysis['result'] is not None:
                                        model_chat_report = genai.GenerativeModel('gemini-1.5-flash-latest')
                                        if analysis['type'] == 'aggregated_table' and isinstance(analysis['result'], pd.DataFrame):
                                            df_summary_head = analysis['result'].head(5).to_html(index=False, classes='dataframe', border=0)
                                            interpretation_prompt = f"""Je dispose d'un tableau de données agrégées. Voici les premières lignes:

{df_summary_head}

Paramètres utilisés : {params}.
Interprète brièvement ce tableau de manière concise et professionnelle. Commence directement par l'interprétation."""
                                        elif analysis['type'] == 'descriptive_stats' and isinstance(analysis['result'], pd.DataFrame):
                                            df_summary_stats = analysis['result'].to_html(index=True, classes='dataframe', border=0)
                                            interpretation_prompt = f"""J'ai les statistiques descriptives suivantes pour une ou plusieurs colonnes :

{df_summary_stats}

Analyse des colonnes: {params.get('selected_columns', [])}.
Interprète ces statistiques de manière concise. Commence directement par l'interprétation."""
                                        
                                        if interpretation_prompt:
                                            genai.configure(api_key=current_api_key)
                                            response_report_ia = model_chat_report.generate_content(interpretation_prompt)
                                            # Mettre à jour l'analyse dans st.session_state
                                            analyses = st.session_state.get('analyses', [])
                                            for idx, session_analysis in enumerate(analyses):
                                                if session_analysis.get('id') == analysis['id']:
                                                    analyses[idx]['modified_interpretation'] = response_report_ia.text.strip()
                                                    st.session_state.analyses = analyses
                                                    break
                                            st.success("✅ Interprétation régénérée !")
                                            st.rerun()
                                except Exception as e:
                                    st.error(f"❌ Erreur lors de la régénération : {e}")
                        
                        st.markdown("---")
                    
                    # Afficher le contenu de l'analyse
                    if analysis['type'] == 'graph' and hasattr(analysis['result'], 'to_plotly_json'):
                        st.plotly_chart(analysis['result'], use_container_width=True)
                    elif analysis['type'] == 'python_script':
                        # Afficher le script généré
                        if 'script' in params:
                            st.markdown("**📝 Script Python généré :**")
                            st.code(params['script'], language="python")
                        # Afficher le résultat
                        if analysis['result'] is not None:
                            if hasattr(analysis['result'], 'to_plotly_json'):
                                st.plotly_chart(analysis['result'], use_container_width=True)
                            elif isinstance(analysis['result'], pd.DataFrame):
                                st.dataframe(analysis['result'], use_container_width=True, hide_index=True)
                            elif isinstance(analysis['result'], plt.Figure):
                                st.pyplot(analysis['result'])
                            else:
                                st.write(analysis['result'])
                        else:
                            st.warning("Aucun résultat à afficher pour ce script Python.")
                    elif analysis['type'] == 'sql_query':
                        if 'query' in params:
                            st.markdown("**📝 Requête SQL utilisée :**")
                            st.code(params['query'], language="sql")
                        if analysis['result'] is not None and isinstance(analysis['result'], pd.DataFrame):
                            st.dataframe(analysis['result'], use_container_width=True, hide_index=True)
                            # Bouton d'export pour les DataFrames avec clé unique
                            df_csv = analysis['result'].copy()
                            try:
                                df_csv.columns = [str(c) for c in df_csv.columns]
                            except Exception:
                                pass
                            try:
                                csv_data = df_csv.to_csv(index=False)
                            except Exception:
                                # fallback: forcer DataFrame simple
                                csv_data = pd.DataFrame(df_csv).to_csv(index=False)
                            st.download_button(
                                label=f"📥 Télécharger (CSV)",
                                data=csv_data,
                                file_name=f"{titre.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                                mime="text/csv",
                                key=f"download_csv_{analysis['id']}"
                            )
                        else:
                            st.warning("La requête SQL n'a retourné aucun résultat ou un format inattendu.")
                    elif isinstance(analysis['result'], pd.DataFrame):
                        # Améliorer l'affichage des tableaux avec un style personnalisé
                        st.markdown(f"""
                        <div style="background: #f8f9fa; padding: 10px; border-radius: 6px; border-left: 4px solid #007bff; margin-bottom: 15px;">
                            <h4 style="margin: 0 0 10px 0; color: #2c3e50;">{titre}</h4>
                        </div>
                        """, unsafe_allow_html=True)

                        # Afficher le tableau avec pagination pour éviter la troncature
                        df_to_display = analysis['result']
                        total_rows = len(df_to_display)
                        
                        # Pagination
                        rows_per_page = 100
                        total_pages = (total_rows + rows_per_page - 1) // rows_per_page
                        
                        if total_pages > 1:
                            col1, col2, col3 = st.columns([1, 2, 1])
                            with col2:
                                current_page = st.selectbox(
                                    f"Page (sur {total_pages})",
                                    options=range(1, total_pages + 1),
                                    key=f"page_{analysis['id']}"
                                )
                            
                            start_idx = (current_page - 1) * rows_per_page
                            end_idx = min(start_idx + rows_per_page, total_rows)
                            df_to_display = df_to_display.iloc[start_idx:end_idx]
                            
                            st.info(f"Affichage des lignes {start_idx + 1} à {end_idx} sur {total_rows} lignes totales")
                        
                        st.dataframe(
                            df_to_display,
                            use_container_width=True,
                            hide_index=True
                        )

                        # Bouton d'export pour les DataFrames avec clé unique
                        df_csv = analysis['result'].copy()
                        try:
                            df_csv.columns = [str(c) for c in df_csv.columns]
                        except Exception:
                            pass
                        try:
                            csv_data = df_csv.to_csv(index=False)
                        except Exception:
                            csv_data = pd.DataFrame(df_csv).to_csv(index=False)
                        st.download_button(
                            label=f"📥 Télécharger (CSV)",
                            data=csv_data,
                            file_name=f"{titre.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                            mime="text/csv",
                            key=f"download_analysis_result_{analysis['id']}_{st.session_state.interpretation_counter}"
                        )
                    else:
                        st.write(analysis['result'])

                    # Bouton de suppression avec clé unique
                    delete_button_key = f"delete_{analysis['id']}_{st.session_state.interpretation_counter}"
                    if st.button(f"🗑️ Supprimer cette analyse", key=delete_button_key):
                        analyses.pop(i) # Supprimer de la liste
                        st.rerun() # Rafraîchir pour mettre à jour l'affichage
    else:
        st.info("Aucune analyse n'a encore été effectuée. Utilisez les outils ci-dessus pour créer des analyses.")

    # Bouton pour effacer toutes les analyses
    if analyses: # Afficher seulement s'il y a des analyses à effacer
        if st.button("🗑️ Effacer toutes les analyses", type="secondary", help="Supprime toutes les analyses et réinitialise la section"):
            st.session_state.analyses = []
            st.success("✅ Toutes les analyses ont été supprimées.")
            st.rerun()

    # --- SECTION ANALYSE INTELLIGENTE ET CONTEXTUELLE ---
    st.subheader("🧠 Analyse Intelligente et Contextuelle")

    # Interface pour recueillir le contexte et les objectifs de l'utilisateur
    st.markdown("### 🎯 Décrivez vos objectifs d'analyse")

    # Zone de texte pour les objectifs de l'utilisateur
    analysis_objective = st.text_area(
        "**Quel est votre objectif principal ?**",
        placeholder="Exemple: Analyser la répartition géographique des répondants et la satisfaction par département.",
        help="Décrivez ce que vous souhaitez analyser. L'IA vous proposera un plan d'analyse.",
        height=150
    )

    # Bouton pour lancer l'analyse intelligente
    if st.button("🧠 Proposer un plan d'analyse par IA", type="primary"):
        if analysis_objective.strip():
            st.session_state.intelligent_analysis_objective = analysis_objective
            # Générer le plan d'analyse intelligent
            if df_current is not None:
                analysis_plan = generate_intelligent_analysis_plan(df_current, analysis_objective)
                st.session_state.intelligent_analysis = analysis_plan
                st.success("✅ Plan d'analyse intelligent généré !")
                st.rerun()
            else:
                st.warning("Veuillez d'abord charger des données.")
        else:
            st.warning("Veuillez décrire votre objectif d'analyse.")

        # Afficher les résultats de l'analyse intelligente
        if st.session_state.get('intelligent_analysis'):
            st.markdown("---")
            st.markdown("### 📊 Résultats de l'analyse intelligente")

            analysis_config = st.session_state.intelligent_analysis
            context = analysis_config['context']
            variables = analysis_config['variables']
            level = analysis_config['level']
        var_analysis = analysis_config.get('variable_analysis', {})

        # Afficher le plan d'analyse généré
        st.markdown(f"**🎯 Objectif :** {analysis_config.get('objective', 'Non spécifié')}")
        st.markdown(f"**📋 Contexte détecté :** {context}")
        st.markdown(f"**📊 Niveau d'analyse :** {level}")
        
        # Permettre à l'utilisateur de modifier les variables sélectionnées
        st.markdown("#### 🔧 Personnalisation des variables")
        st.info("💡 Vous pouvez modifier la sélection des variables avant d'exécuter les analyses")
        
        # Afficher toutes les variables disponibles avec leur signification
        all_variables = list(df_current.columns)
        available_vars_with_significance = []
        
        for var in all_variables:
            if var in var_analysis:
                sig = var_analysis[var]['significance']
                available_vars_with_significance.append({
                    'variable': var,
                    'significance': sig,
                    'selected': var in variables
                })
            else:
                available_vars_with_significance.append({
                    'variable': var,
                    'significance': 'unknown',
                    'selected': var in variables
                })
        
        # Créer un DataFrame pour l'affichage
        var_df = pd.DataFrame(available_vars_with_significance)
        
        # Afficher le tableau des variables avec cases à cocher
        st.markdown("**📋 Variables disponibles :**")
        
        # Créer des colonnes pour organiser l'affichage
        col1, col2 = st.columns([2, 1])
        
        with col1:
            # Cases à cocher pour chaque variable
            selected_vars = []
            for _, row in var_df.iterrows():
                var_name = row['variable']
                var_sig = row['significance']
                is_selected = st.checkbox(
                    f"**{var_name}** ({var_sig.replace('_', ' ').title()})",
                    value=row['selected'],
                    key=f"var_select_{var_name}",
                    help=f"Type: {var_sig.replace('_', ' ').title()}"
                )
                if is_selected:
                    selected_vars.append(var_name)
        
        with col2:
            # Résumé de la sélection
            st.markdown("**📊 Résumé de la sélection :**")
            st.metric("Variables sélectionnées", len(selected_vars))
            st.metric("Total disponibles", len(all_variables))
            
            # Bouton pour mettre à jour la sélection
            if st.button("🔄 Mettre à jour la sélection", key="update_var_selection"):
                # Mettre à jour les variables dans l'analyse intelligente
                st.session_state.intelligent_analysis['variables'] = selected_vars
                st.success(f"✅ Sélection mise à jour : {len(selected_vars)} variables sélectionnées")
                st.rerun()
        
        # Afficher les variables actuellement sélectionnées
        current_vars = st.session_state.intelligent_analysis.get('variables', variables)
        st.markdown(f"**🔍 Variables actuellement sélectionnées :** {', '.join(current_vars)}")

        # Afficher l'analyse de signification des variables
        if var_analysis:
            st.markdown("#### 🔍 Analyse de la signification des variables")
            significance_summary = {}
            for col, analysis in var_analysis.items():
                if col in current_vars:
                    sig = analysis['significance']
                    if sig not in significance_summary:
                        significance_summary[sig] = []
                    significance_summary[sig].append(col)
            
            for sig, cols in significance_summary.items():
                st.markdown(f"**{sig.replace('_', ' ').title()} :** {', '.join(cols)}")

        # Bouton pour exécuter automatiquement les analyses intelligentes
        if st.button("🚀 Exécuter les analyses intelligentes", type="primary", key="execute_intelligent_analyses"):
            with st.spinner("Exécution des analyses intelligentes..."):
                analyses = st.session_state.get('analyses', [])
                
                # Utiliser les variables mises à jour
                current_vars = st.session_state.intelligent_analysis.get('variables', variables)

                # Générer les analyses appropriées selon le contexte
                if context == "Analyse démographique":
                    st.markdown("#### 👥 Analyse Démographique")
                    generate_demographic_analysis(df_current, current_vars)
                    
                    # Ajouter à la liste des analyses
                    for var in current_vars:
                        if var in df_current.columns:
                            analysis_id = f"intelligent_demographic_{var}_{len(analyses)}"
                            analyses.append({
                                'type': 'intelligent_demographic',
                                'result': df_current[var].value_counts().to_frame(),
                                'id': analysis_id,
                                'executed_params': {
                                    'variable': var,
                                    'title': f"Analyse démographique - {var}",
                                    'context': context
                                }
                            })

                elif context == "Étude de satisfaction":
                    st.markdown("#### 😊 Analyse de Satisfaction")
                    generate_satisfaction_analysis(df_current, current_vars)
                    
                    # Ajouter à la liste des analyses
                    for var in current_vars:
                        if var in df_current.columns:
                            analysis_id = f"intelligent_satisfaction_{var}_{len(analyses)}"
                            analyses.append({
                                'type': 'intelligent_satisfaction',
                                'result': df_current[var].value_counts().to_frame(),
                                'id': analysis_id,
                                'executed_params': {
                                    'variable': var,
                                    'title': f"Analyse de satisfaction - {var}",
                                    'context': context
                                }
                            })

                elif context == "Analyse géographique":
                    st.markdown("#### 🗺️ Analyse Géographique")
                    generate_geographic_analysis(df_current, current_vars)
                    
                    # Ajouter à la liste des analyses
                    for var in current_vars:
                        if var in df_current.columns:
                            analysis_id = f"intelligent_geographic_{var}_{len(analyses)}"
                            analyses.append({
                                'type': 'intelligent_geographic',
                                'result': df_current[var].value_counts().to_frame(),
                                'id': analysis_id,
                                'executed_params': {
                                    'variable': var,
                                    'title': f"Analyse géographique - {var}",
                                    'context': context
                                }
                            })

                elif context == "Étude des corrélations":
                    st.markdown("#### 🔗 Analyse des Corrélations")
                    generate_correlation_analysis(df_current, current_vars)
                    
                    # Ajouter à la liste des analyses
                    if len(current_vars) >= 2:
                        analysis_id = f"intelligent_correlation_{len(analyses)}"
                        corr_matrix = df_current[current_vars].corr()
                        analyses.append({
                            'type': 'intelligent_correlation',
                            'result': corr_matrix,
                            'id': analysis_id,
                            'executed_params': {
                                'variables': current_vars,
                                'title': f"Analyse des corrélations - {', '.join(current_vars)}",
                                'context': context
                            }
                        })

                elif context == "Analyse comparative":
                    st.markdown("#### ⚖️ Analyse Comparative")
                    generate_comparative_analysis(df_current, current_vars)

                elif context == "Prédiction ou modélisation":
                    st.markdown("#### 🔮 Analyse Prédictive")
                    generate_predictive_analysis(df_current, current_vars)

                elif context == "Vue d'ensemble des données":
                    st.markdown("#### 📊 Vue d'Ensemble des Données")
                    # Générer des analyses descriptives pour toutes les variables importantes
                    for var in current_vars:
                        if var in df_current.columns:
                            st.markdown(f"**📈 Analyse de {var}**")
                            if df_current[var].dtype in ['int64', 'float64']:
                                # Statistiques descriptives pour les variables numériques
                                stats = df_current[var].describe()
                                col1, col2, col3, col4 = st.columns(4)
                                col1.metric("Moyenne", f"{stats['mean']:.2f}")
                                col2.metric("Médiane", f"{stats['50%']:.2f}")
                                col3.metric("Écart-type", f"{stats['std']:.2f}")
                                col4.metric("Min-Max", f"{stats['min']:.0f} - {stats['max']:.0f}")
                            else:
                                # Analyse catégorielle pour les variables non-numériques
                                generate_enhanced_categorical_analysis(df_current, var)
                    
                    # Ajouter à la liste des analyses
                    for var in current_vars:
                        if var in df_current.columns:
                            analysis_id = f"intelligent_overview_{var}_{len(analyses)}"
                            if df_current[var].dtype in ['int64', 'float64']:
                                result_data = df_current[var].describe().to_frame()
                            else:
                                result_data = df_current[var].value_counts().to_frame()
                            
                            analyses.append({
                                'type': 'intelligent_overview',
                                'result': result_data,
                                'id': analysis_id,
                                'executed_params': {
                                    'variable': var,
                                    'title': f"Vue d'ensemble - {var}",
                                    'context': context
                                }
                            })

                else:
                    st.markdown("#### 📈 Analyse Générale")
                    generate_general_analysis(df_current, current_vars, level)
                    
                    # Ajouter à la liste des analyses
                    for var in current_vars:
                        if var in df_current.columns:
                            analysis_id = f"intelligent_general_{var}_{len(analyses)}"
                            analyses.append({
                                'type': 'intelligent_general',
                                'result': df_current[var].value_counts().to_frame(),
                                'id': analysis_id,
                                'executed_params': {
                                    'variable': var,
                                    'title': f"Analyse générale - {var}",
                                    'context': context
                                }
                            })

                st.session_state.analyses = analyses
                st.success(f"✅ {len(current_vars)} analyses intelligentes ajoutées à la liste des analyses !")
                # Ne pas faire de rerun pour éviter la disparition des résultats
                # st.rerun()

        # Afficher les analyses sans les exécuter automatiquement
        st.markdown("#### 📋 Aperçu des analyses proposées")
        if context == "Analyse démographique":
            st.markdown("**👥 Analyse Démographique**")
            st.info(f"Variables à analyser : {', '.join(current_vars)}")
            
        elif context == "Étude de satisfaction":
            st.markdown("**😊 Analyse de Satisfaction**")
            st.info(f"Variables à analyser : {', '.join(current_vars)}")
            
        elif context == "Analyse géographique":
            st.markdown("**🗺️ Analyse Géographique**")
            st.info(f"Variables à analyser : {', '.join(current_vars)}")
                
        elif context == "Étude des corrélations":
            st.markdown("**🔗 Analyse des Corrélations**")
            st.info(f"Variables à analyser : {', '.join(current_vars)}")
                
        elif context == "Analyse comparative":
            st.markdown("**⚖️ Analyse Comparative**")
            st.info(f"Variables à analyser : {', '.join(current_vars)}")
                
        elif context == "Prédiction ou modélisation":
            st.markdown("**🔮 Analyse Prédictive**")
            st.info(f"Variables à analyser : {', '.join(current_vars)}")
                
        elif context == "Vue d'ensemble des données":
            st.markdown("**📊 Vue d'Ensemble des Données**")
            st.info(f"Variables principales à analyser : {', '.join(current_vars)}")
            st.markdown("**🎯 Objectif :** Fournir une vue globale et synthétique des données collectées")
        else:
            st.markdown("**📈 Analyse Générale**")
            st.info(f"Variables à analyser : {', '.join(current_vars)}")
        
        # Afficher les résultats des analyses intelligentes exécutées
        st.markdown("#### 📊 Résultats des analyses intelligentes")
        
        # Vérifier s'il y a des analyses intelligentes dans la session
        intelligent_analyses = [a for a in st.session_state.get('analyses', []) if a['type'].startswith('intelligent_')]
        
        if intelligent_analyses:
            st.success(f"✅ {len(intelligent_analyses)} analyses intelligentes ont été exécutées")
            
            # Afficher un résumé de chaque analyse
            for analysis in intelligent_analyses:
                with st.expander(f"📋 {analysis.get('executed_params', {}).get('title', 'Analyse intelligente')}", expanded=False):
                    if analysis['type'] == 'intelligent_overview':
                        st.markdown("**📊 Vue d'ensemble**")
                        if isinstance(analysis['result'], pd.DataFrame):
                            st.dataframe(analysis['result'], use_container_width=True)
                        else:
                            st.write(analysis['result'])
                    elif analysis['type'] == 'intelligent_demographic':
                        st.markdown("**👥 Analyse démographique**")
                        if isinstance(analysis['result'], pd.DataFrame):
                            st.dataframe(analysis['result'], use_container_width=True)
                        else:
                            st.write(analysis['result'])
                    elif analysis['type'] == 'intelligent_satisfaction':
                        st.markdown("**😊 Analyse de satisfaction**")
                        if isinstance(analysis['result'], pd.DataFrame):
                            st.dataframe(analysis['result'], use_container_width=True)
                        else:
                            st.write(analysis['result'])
                    elif analysis['type'] == 'intelligent_geographic':
                        st.markdown("**🗺️ Analyse géographique**")
                        if isinstance(analysis['result'], pd.DataFrame):
                            st.dataframe(analysis['result'], use_container_width=True)
                        else:
                            st.write(analysis['result'])
                    elif analysis['type'] == 'intelligent_correlation':
                        st.markdown("**🔗 Analyse des corrélations**")
                        if isinstance(analysis['result'], pd.DataFrame):
                            st.dataframe(analysis['result'], use_container_width=True)
                        else:
                            st.write(analysis['result'])
                    else:
                        st.markdown("**📈 Analyse générale**")
                        if isinstance(analysis['result'], pd.DataFrame):
                            st.dataframe(analysis['result'], use_container_width=True)
                        else:
                            st.write(analysis['result'])
                
                # Bouton pour voir les insights complets pour tous les tableaux
                if isinstance(analysis['result'], pd.DataFrame) and not analysis['result'].empty:
                    col1, col2 = st.columns([1, 4])
                    with col1:
                        if st.button(f"💡 Voir les insights", key=f"insights_{analysis.get('id', 'default')}_{analysis['type']}"):
                            st.session_state[f"show_insights_{analysis.get('id', 'default')}_{analysis['type']}"] = True
                    
                    # Afficher les insights si demandé
                    if st.session_state.get(f"show_insights_{analysis.get('id', 'default')}_{analysis['type']}", False):
                        with col2:
                            st.markdown("### 💡 Insights automatiques du tableau")
                            insights = generate_table_insights(analysis['result'], analysis.get('executed_params', {}))
                            st.markdown(insights)
        else:
            st.info("💡 Aucune analyse intelligente n'a encore été exécutée. Utilisez le bouton '🚀 Exécuter les analyses intelligentes' ci-dessus.")

    # --- SECTION ANALYSE COMPLÈTE PERSONNALISÉE (ancienne) ---
    st.subheader("📊 Analyse Complète Personnalisée (Méthode classique)")

    if df_current is not None:
        # Section de détection automatique supprimée - Plus nécessaire
        st.info("💡 Utilisez l'analyse intelligente ci-dessus pour obtenir des suggestions d'analyses basées sur la signification de vos variables.")

        # Section supprimée - Plus de propositions automatiques d'analyses
        st.info("💡 Utilisez l'analyse intelligente ci-dessus pour obtenir des suggestions d'analyses basées sur la signification de vos variables.")

# --- CONTENU DE L'ONGLET "Analyse Avancée" ---
with advanced_tab:
    st.header("🧠 Analyse Intelligente des Données")
    st.caption("Ce module analyse automatiquement le jeu de données actuellement chargé : détection des types de "
               "variables, statistiques pertinentes, corrélations, tendances et répartitions géographiques — "
               "avec des visualisations interactives et un rapport Word exportable prêt à partager.")

    df_current = st.session_state.get('df_current')

    if df_current is None or df_current.empty:
        st.info("👆 Aucun jeu de données chargé. Rendez-vous dans l'onglet **« Application & Chat IA »** pour "
                "charger une collecte Kobo (ou un fichier), puis revenez ici.")
    else:
        insights = render_smart_dashboard(df_current)

        # ================================================================
        # INTERPRÉTATION APPROFONDIE PAR IA (fournisseur au choix)
        # ================================================================
        st.markdown("---")
        st.subheader("🤖 Interprétation approfondie par IA")
        st.caption("Envoie la synthèse des constats ci-dessus (pas les données brutes) à un modèle d'IA de "
                   "ton choix pour obtenir une lecture stratégique : hypothèses d'explication, disparités "
                   "territoriales, recommandations opérationnelles. Ta clé API n'est jamais enregistrée : "
                   "elle reste en mémoire pour cette session uniquement.")

        with st.expander("⚙️ Configurer le fournisseur IA", expanded=not st.session_state.get("_ai_provider_configured")):
            col_ai1, col_ai2 = st.columns(2)
            with col_ai1:
                ai_provider = st.selectbox(
                    "Fournisseur",
                    options=list(ai_interpreter.PROVIDERS.keys()),
                    format_func=lambda k: ai_interpreter.PROVIDERS[k],
                    key="_ai_provider_select",
                )
            with col_ai2:
                ai_model = st.text_input(
                    "Modèle (optionnel)",
                    value="",
                    placeholder=ai_interpreter.DEFAULT_MODELS.get(ai_provider, ""),
                    key="_ai_model_input",
                )
            ai_base_url = ""
            if ai_provider == "openai_compatible":
                ai_base_url = st.text_input(
                    "Endpoint (base URL, ex: http://localhost:11434/v1 pour Ollama)",
                    key="_ai_base_url_input",
                )
            ai_api_key = st.text_input("Clé API", type="password", key="_ai_api_key_input")
            ai_extra_context = st.text_area(
                "Contexte additionnel pour l'IA (optionnel)",
                placeholder="Ex : objectifs du programme PRIMA, zones prioritaires, éléments déjà connus du terrain...",
                key="_ai_extra_context_input",
            )
            st.session_state["_ai_provider_configured"] = bool(ai_api_key)

        gen_ai = st.button("🤖 Générer l'interprétation IA", type="secondary", use_container_width=False)

        if gen_ai:
            if not insights:
                st.warning("Aucun insight disponible pour l'interprétation.")
            elif not st.session_state.get("_ai_api_key_input"):
                st.warning("Merci de renseigner une clé API dans la configuration ci-dessus.")
            else:
                with st.spinner("Interprétation en cours par l'IA..."):
                    try:
                        profile = st.session_state.get("_smart_profile_cache") or profile_dataframe(df_current)
                        ctx = build_survey_context(df_current, profile)
                        ai_text = ai_interpreter.generate_ai_interpretation(
                            profile, insights, ctx.theme_label,
                            provider=st.session_state["_ai_provider_select"],
                            api_key=st.session_state["_ai_api_key_input"],
                            model=st.session_state.get("_ai_model_input") or None,
                            base_url=st.session_state.get("_ai_base_url_input") or None,
                            extra_context=st.session_state.get("_ai_extra_context_input", ""),
                        )
                        st.session_state["_smart_ai_interpretation"] = ai_text
                        st.success("✅ Interprétation générée !")
                    except ai_interpreter.AIInterpreterError as e:
                        st.error(f"Erreur IA : {e}")
                    except Exception as e:
                        st.error(f"Erreur inattendue : {e}")

        if st.session_state.get("_smart_ai_interpretation"):
            with st.container(border=True):
                st.markdown(st.session_state["_smart_ai_interpretation"])

        # ================================================================
        # RAPPORT WORD
        # ================================================================
        st.markdown("---")
        st.subheader("📄 Rapport Word")
        st.caption("Génère un rapport Word soigné (page de garde, sommaire, résumé exécutif, graphiques et "
                   "tableaux) à partir de l'analyse ci-dessus, prêt à être partagé ou imprimé.")

        col_r1, col_r2 = st.columns([1, 2])
        with col_r1:
            gen_report = st.button("📝 Générer le rapport Word", type="primary", use_container_width=True)
        with col_r2:
            report_title = st.text_input(
                "Titre du rapport",
                value=f"Rapport d'analyse — {st.session_state.get('selected_collection', 'Jeu de données')}"[:120],
                key="smart_report_title",
            )

        include_ai_in_report = False
        if st.session_state.get("_smart_ai_interpretation"):
            include_ai_in_report = st.checkbox(
                "Inclure l'interprétation IA générée ci-dessus dans le rapport Word", value=True,
                key="_include_ai_in_report",
            )

        if gen_report:
            if not insights:
                st.warning("Aucun insight disponible pour générer le rapport.")
            else:
                with st.spinner("Génération du rapport Word en cours..."):
                    try:
                        profile = st.session_state.get("_smart_profile_cache") or profile_dataframe(df_current)
                        selected_header_key = st.session_state.get('current_collection_header', '')
                        logo_path, org_name, _ = get_logo_and_org_info(selected_header_key)
                        summary_text = build_executive_summary(df_current, profile, insights)
                        report_meta = {
                            "title": report_title or "Rapport d'analyse de données",
                            "subtitle": "Analyse intelligente automatisée",
                            "org_name": org_name,
                            "logo_path": logo_path,
                            "author": f"{org_name} — Suivi-Évaluation",
                        }
                        ai_text_for_report = (
                            st.session_state.get("_smart_ai_interpretation") if include_ai_in_report else None
                        )
                        buffer = build_word_report(
                            df_current, profile, insights,
                            meta=report_meta, executive_summary=summary_text,
                            ai_interpretation=ai_text_for_report,
                        )
                        st.session_state["_smart_report_buffer"] = buffer.getvalue()
                        st.success("✅ Rapport généré avec succès !")
                    except Exception as e:
                        st.error(f"Erreur lors de la génération du rapport : {e}")
                        import traceback
                        st.code(traceback.format_exc())

        if st.session_state.get("_smart_report_buffer"):
            st.download_button(
                "⬇️ Télécharger le rapport (.docx)",
                data=st.session_state["_smart_report_buffer"],
                file_name=f"rapport_analyse_{datetime.now().strftime('%Y%m%d_%H%M')}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                use_container_width=True,
            )
            st.caption("💡 À l'ouverture dans Word, faites un clic droit sur le sommaire puis "
                       "**« Mettre à jour les champs »** pour qu'il se remplisse automatiquement.")

# --- CONTENU DE L'ONGLET "Manuel d'utilisation" ---
with manual_tab:
    st.header("Manuel d'utilisation")
    # Correction de la délimitation de la chaîne multiligne
    st.markdown("""
Bienvenue dans le manuel d'utilisation de l'application CCR-B & IFRIZ DataViz.

**1. Charger des Données :**
* Utilisez la barre latérale pour sélectionner une **collecte KoboToolbox prédéfinie** ou pour **charger votre propre fichier Excel (.xlsx)**.
* Après le chargement, les données s'affichent dans l'onglet 'Application & Chat IA'.

**2. Effectuer des Analyses :**
* Dans l'onglet **'Application & Chat IA'**, vous trouverez plusieurs options :
    * **Configuration des Analyses :** Ajoutez des tableaux agrégés, des statistiques descriptives, des graphiques (barres, lignes, etc.), des histogrammes, des boîtes à moustaches, des camemberts, des requêtes SQL et des scripts Python.
    * **Génération par IA :** Décrivez votre besoin en langage naturel pour générer automatiquement des scripts Python ou des requêtes SQL.
    * **Chat IA :** Posez des questions à l'IA Gemini sur vos données. Elle peut vous aider à interpréter les résultats ou à suggérer des analyses. N'oubliez pas de configurer votre clé API Google Gemini dans la barre latérale pour activer ces fonctionnalités.
* **Analyse Complète Automatique :** Cliquez sur le bouton '🚀 Lancer l'analyse complète' pour obtenir une vue d'ensemble rapide de vos données (analyses démographiques, géographiques, distributions, corrélations).

**3. Exporter des Rapports :**
* Une fois vos analyses terminées, utilisez les boutons **'📥 Télécharger Rapport HTML'** et **'📥 Télécharger Rapport PDF'** dans la barre latérale pour exporter un résumé complet de votre travail.
* Vous pouvez choisir d'inclure ou d'exclure l'analyse complète automatique dans vos rapports via les boutons de la barre latérale.

**4. Gérer les Analyses :**
* Chaque analyse effectuée apparaît dans l'onglet **'Application & Chat IA'** sous 'Résultats des analyses'.
* Vous pouvez voir un aperçu rapide (insights) et les détails complets de chaque analyse.
* Utilisez le bouton **'🗑️ Supprimer cette analyse'** pour retirer une analyse spécifique.
* Utilisez **'🗑️ Effacer toutes les analyses'** pour recommencer.

**5. Navigation :**
* Utilisez les onglets en haut de la page pour naviguer entre les différentes sections de l'application.
* La **barre latérale** contient les contrôles principaux : chargement des données, configuration IA, options de rapport, et boutons de réinitialisation.

**Conseils :**
* Pour les analyses IA, soyez aussi précis que possible dans vos descriptions.
* Si un script généré par l'IA échoue, essayez de simplifier votre demande ou de régénérer le script. Les messages d'erreur vous aideront à identifier le problème.
* Assurez-vous que votre clé API Google Gemini est correctement configurée pour profiter pleinement des fonctionnalités IA.
""")

# --- CONTENU DE L'ONGLET "Tableau de Bord" ---
with dashboard_tab:
    st.header("📊 Tableau de Bord - Vue Globale")

    # --- SECTION KPIs GLOBAUX ---
    st.subheader("📈 Indicateurs Clés de Performance (KPIs)")

    # Calculer les KPIs globaux
    total_collectes = len(urls)
    analyses_list = st.session_state.get('analyses', [])
    total_analyses = len(analyses_list)
    df_current = st.session_state.get('df_current')

    # KPIs en colonnes
    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        st.metric("📋 Total Collectes", total_collectes, help="Nombre total de collectes configurées")
    with col2:
        st.metric("📊 Analyses Réalisées", total_analyses, help="Nombre d'analyses effectuées dans la session")
    with col3:
        if df_current is not None:
            st.metric("📝 Lignes de Données", f"{len(df_current):,}", help="Volume de données chargées")
        else:
            st.metric("📝 Lignes de Données", "N/A", help="Aucune donnée chargée")
    with col4:
        if df_current is not None:
            st.metric("🔢 Colonnes", len(df_current.columns), help="Nombre de variables disponibles")
        else:
            st.metric("🔢 Colonnes", "N/A")
    with col5:
        if df_current is not None:
            missing_pct = (df_current.isnull().sum().sum() / (df_current.shape[0] * df_current.shape[1]) * 100)
            st.metric("⚠️ Données Manquantes", f"{missing_pct:.1f}%", help="Pourcentage de valeurs manquantes")
        else:
            st.metric("⚠️ Données Manquantes", "N/A")

    st.markdown("---")

    # --- SECTION ÉVOLUTION DES COLLECTES ---
    st.subheader("📅 Évolution et Suivi des Collectes")

    # Tableau récapitulatif de toutes les collectes
    collections_summary = []
    for name, url in urls.items():
        # Extraire le numéro de collecte
        num = name.split(":")[0].replace("Collecte", "").strip() if "Collecte" in name else "N/A"
        collections_summary.append({
            "N°": num,
            "Nom": name.split(":")[1].strip() if ":" in name else name,
            "URL Disponible": "✅" if url else "❌",
        })

    summary_df = pd.DataFrame(collections_summary)

    # Afficher avec pagination
    st.markdown("**📋 Liste des collectes configurées :**")
    show_df = ensure_arrow_compatible(summary_df)
    st.dataframe(show_df, use_container_width=True, height=400)

    st.markdown("---")

    # --- SECTION DERNIÈRES ANALYSES ---
    st.subheader("🔍 Dernières Analyses Effectuées")

    if analyses_list:
        # Afficher les 5 dernières analyses
        recent_analyses = analyses_list[-5:][::-1]
        for i, analysis in enumerate(recent_analyses):
            if analysis.get('result') is not None:
                params = analysis.get('executed_params', {})
                titre = params.get('title', f"Analyse {analysis.get('id', '')}")
                analysis_type = analysis.get('type', 'unknown')

                with st.expander(f"📊 {titre} ({analysis_type.replace('_', ' ').title()})", expanded=(i == 0)):
                    if isinstance(analysis['result'], pd.DataFrame):
                        st.dataframe(analysis['result'].head(10), use_container_width=True, hide_index=True)
                    elif hasattr(analysis['result'], 'to_plotly_json'):
                        st.plotly_chart(analysis['result'], height=300, use_container_width=True)
                    elif isinstance(analysis['result'], plt.Figure):
                        st.pyplot(analysis['result'])
                    else:
                        st.write(f"Résultat : {str(analysis['result'])[:200]}")
    else:
        st.info("Aucune analyse n'a encore été effectuée. Utilisez l'onglet 'Application & Chat IA' pour commencer.")

    st.markdown("---")

    # --- SECTION ANALYSES CROISÉES AVANCÉES ---
    st.subheader("🔗 Analyses Croisées et Statistiques Avancées")

    if df_current is not None and not df_current.empty:
        # Sélection des variables pour l'analyse croisée
        st.markdown("**Configurer une analyse croisée :**")

        col1, col2 = st.columns(2)
        with col1:
            var_x = st.selectbox("Variable X (catégorielle) :", options=[""] + [c for c in df_current.columns if df_current[c].dtype == 'object' or df_current[c].nunique() <= 15], key="dashboard_var_x")
        with col2:
            var_y = st.selectbox("Variable Y (numérique ou catégorielle) :", options=[""] + list(df_current.columns), key="dashboard_var_y")

        if var_x and var_y:
            if st.button("🚀 Lancer l'analyse croisée", type="primary", key="run_cross_analysis"):
                with st.spinner("Analyse en cours..."):
                    # Statistiques descriptives croisées
                    if df_current[var_y].dtype in ['int64', 'float64']:
                        st.markdown(f"**📊 Statistiques croisées : {var_y} par {var_x}**")
                        cross_stats = df_current.groupby(var_x)[var_y].agg(['count', 'mean', 'median', 'std', 'min', 'max']).round(2)
                        st.dataframe(cross_stats, use_container_width=True)

                        # Graphique de comparaison
                        fig_cross = px.box(df_current, x=var_x, y=var_y, title=f"Distribution de {var_y} par {var_x}")
                        fig_cross.update_layout(xaxis_tickangle=45)
                        st.plotly_chart(fig_cross, use_container_width=True)

                        # Test statistique (ANOVA si applicable)
                        groups = [group[var_y].dropna().values for name, group in df_current.groupby(var_x) if len(group[var_y].dropna()) > 0]
                        if len(groups) >= 2 and all(len(g) > 1 for g in groups):
                            try:
                                f_stat, p_value = stats.f_oneway(*groups)
                                st.markdown(f"**🧪 Test ANOVA :** F = {f_stat:.3f}, p-value = {p_value:.4f}")
                                if p_value < 0.05:
                                    st.success("✅ Différence statistiquement significative entre les groupes (p < 0.05)")
                                else:
                                    st.info("ℹ️ Pas de différence significative entre les groupes (p >= 0.05)")
                            except Exception:
                                pass
                    else:
                        # Tableau croisé pour deux variables catégorielles
                        st.markdown(f"**📊 Tableau croisé : {var_x} vs {var_y}**")
                        cross_tab = pd.crosstab(df_current[var_x], df_current[var_y], margins=True, margins_name="Total")
                        st.dataframe(cross_tab, use_container_width=True)

                        # Graphique heatmap
                        cross_pct = pd.crosstab(df_current[var_x], df_current[var_y], normalize='index') * 100
                        fig_heatmap = px.imshow(cross_pct, text_auto='.1f', title=f"Pourcentage croisé : {var_x} vs {var_y}",
                                                labels=dict(x=var_y, y=var_x, color="%"), color_continuous_scale="Blues")
                        st.plotly_chart(fig_heatmap, use_container_width=True)

                        # Test du Chi-2
                        try:
                            chi2, p_val_chi2, dof, expected = stats.chi2_contingency(pd.crosstab(df_current[var_x], df_current[var_y]))
                            st.markdown(f"**🧪 Test du Chi-2 :** χ² = {chi2:.3f}, p-value = {p_val_chi2:.4f}, ddl = {dof}")
                            if p_val_chi2 < 0.05:
                                st.success("✅ Association statistiquement significative (p < 0.05)")
                            else:
                                st.info("ℹ️ Pas d'association significative (p >= 0.05)")
                        except Exception:
                            pass

        # --- Matrice de corrélation pour variables numériques ---
        st.markdown("---")
        st.markdown("**🔗 Matrice de Corrélation (Variables Numériques)**")
        numerical_cols_df = [col for col in df_current.columns if pd.api.types.is_numeric_dtype(df_current[col]) and df_current[col].nunique() > 5]

        if len(numerical_cols_df) >= 2:
            selected_num_cols = st.multiselect("Sélectionner les variables numériques :", numerical_cols_df, default=numerical_cols_df[:min(5, len(numerical_cols_df))], key="dashboard_corr_cols")
            if len(selected_num_cols) >= 2 and st.button("📊 Générer la matrice de corrélation", key="dashboard_corr_btn"):
                corr_matrix = df_current[selected_num_cols].corr()
                fig_corr = px.imshow(corr_matrix, text_auto='.2f', title="Matrice de Corrélation",
                                     color_continuous_scale='RdBu_r', zmin=-1, zmax=1)
                st.plotly_chart(fig_corr, use_container_width=True)

                # Identifier les corrélations fortes
                strong_corrs = []
                for i in range(len(corr_matrix.columns)):
                    for j in range(i+1, len(corr_matrix.columns)):
                        corr_val = corr_matrix.iloc[i, j]
                        if abs(corr_val) > 0.5:
                            strong_corrs.append({
                                "Variable 1": corr_matrix.columns[i],
                                "Variable 2": corr_matrix.columns[j],
                                "Corrélation": f"{corr_val:.3f}",
                                "Force": "🟢 Forte (+)" if corr_val > 0.7 else ("🔴 Forte (-)" if corr_val < -0.7 else "🟡 Modérée")
                            })
                if strong_corrs:
                    st.markdown("**🔍 Corrélations significatives détectées :**")
                    st.dataframe(pd.DataFrame(strong_corrs), use_container_width=True)
        else:
            st.info("ℹ️ Au moins 2 variables numériques nécessaires pour la matrice de corrélation.")

        # --- Analyse de normalité ---
        st.markdown("---")
        st.markdown("**🧪 Tests de Normalité**")
        normality_var = st.selectbox("Variable pour test de normalité :", options=[""] + numerical_cols_df, key="dashboard_normality")
        if normality_var and st.button("Tester la normalité", key="dashboard_normality_btn"):
            data_normality = df_current[normality_var].dropna()
            if len(data_normality) >= 8:
                stat_sw, p_sw = stats.shapiro(data_normality[:5000])  # Shapiro limité à 5000
                stat_ks, p_ks = stats.kstest(data_normality, 'norm', args=(data_normality.mean(), data_normality.std()))
                st.markdown(f"**Test de Shapiro-Wilk :** Stat = {stat_sw:.4f}, p-value = {p_sw:.4f}")
                st.markdown(f"**Test de Kolmogorov-Smirnov :** Stat = {stat_ks:.4f}, p-value = {p_ks:.4f}")
                if p_sw > 0.05:
                    st.success(f"✅ Distribution normale (Shapiro p={p_sw:.4f} > 0.05)")
                else:
                    st.warning(f"⚠️ Distribution non normale (Shapiro p={p_sw:.4f} < 0.05)")

                # Histogramme avec courbe de densité
                fig_dist = px.histogram(data_normality, x=normality_var, title=f"Distribution de {normality_var}",
                                       marginal="box", opacity=0.7, color_discrete_sequence=['#3498db'])
                fig_dist.update_layout(bargap=0.1)
                st.plotly_chart(fig_dist, use_container_width=True)

    else:
        st.info("💡 Chargez des données dans l'onglet 'Application & Chat IA' pour accéder aux analyses avancées.")

    # --- SECTION POINTS IMPORTANTS PAR QUESTION ---
    st.markdown("---")
    st.subheader("💡 Points Importants par Question")

    if df_current is not None and not df_current.empty:
        # Identifier les questions (colonnes) avec des points importants
        if st.button("🔍 Analyser les points clés automatiquement", key="dashboard_key_points_btn"):
            with st.spinner("Analyse des points importants en cours..."):
                key_findings = []

                # 1. Colonnes catégorielles avec une valeur dominante forte
                categorical_cols_dash = [col for col in df_current.columns
                                         if df_current[col].dtype == 'object' or df_current[col].nunique() <= 10]

                for col in categorical_cols_dash[:10]:  # Limiter à 10 pour la performance
                    if df_current[col].nunique() > 1:
                        value_counts = df_current[col].value_counts(normalize=True, dropna=True)
                        if len(value_counts) > 0:
                            top_val = value_counts.index[0]
                            top_pct = value_counts.iloc[0] * 100
                            if top_pct >= 40:  # Seuil de dominance
                                key_findings.append({
                                    "Question": col,
                                    "Type": "🎯 Valeur dominante",
                                    "Insight": f"'{top_val}' représente {top_pct:.1f}% des réponses",
                                    "Sévérité": "🔴 Forte" if top_pct >= 60 else "🟡 Modérée"
                                })

                # 2. Colonnes numériques avec valeurs extremes
                numerical_cols_dash = [col for col in df_current.columns
                                      if pd.api.types.is_numeric_dtype(df_current[col]) and df_current[col].nunique() > 5]

                for col in numerical_cols_dash[:10]:
                    if not df_current[col].dropna().empty:
                        mean_val = df_current[col].mean()
                        std_val = df_current[col].std()
                        max_val = df_current[col].max()
                        min_val = df_current[col].min()

                        # Détecter les valeurs aberrantes (outliers)
                        q1 = df_current[col].quantile(0.25)
                        q3 = df_current[col].quantile(0.75)
                        iqr = q3 - q1
                        outliers = df_current[(df_current[col] < q1 - 1.5 * iqr) | (df_current[col] > q3 + 1.5 * iqr)]

                        if len(outliers) > 0:
                            outlier_pct = (len(outliers) / len(df_current)) * 100
                            key_findings.append({
                                "Question": col,
                                "Type": "⚠️ Valeurs aberrantes",
                                "Insight": f"{len(outliers)} valeurs aberrantes ({outlier_pct:.1f}%) - Min: {min_val:.1f}, Max: {max_val:.1f}",
                                "Sévérité": "🟡 Modérée" if outlier_pct < 5 else "🔴 Forte"
                            })

                        # Détecter les colonnes avec forte variance
                        if std_val > 0 and (max_val - min_val) / (std_val + 1e-10) > 10:
                            key_findings.append({
                                "Question": col,
                                "Type": "📊 Grande dispersion",
                                "Insight": f"Étendue: {min_val:.1f} à {max_val:.1f} (Moyenne: {mean_val:.1f})",
                                "Sévérité": "🟢 Info"
                            })

                # 3. Colonnes avec données manquantes importantes
                missing_cols = df_current.isnull().sum()
                missing_cols = missing_cols[missing_cols > 0].sort_values(ascending=False)
                for col, missing_count in missing_cols.head(5).items():
                    missing_pct = (missing_count / len(df_current)) * 100
                    if missing_pct >= 10:
                        key_findings.append({
                            "Question": col,
                            "Type": "❓ Données manquantes",
                            "Insight": f"{missing_count} valeurs manquantes ({missing_pct:.1f}%)",
                            "Sévérité": "🔴 Critique" if missing_pct >= 30 else "🟡 Modérée"
                        })

                # Afficher les résultats
                if key_findings:
                    findings_df = pd.DataFrame(key_findings)
                    st.markdown(f"**🎯 {len(key_findings)} points importants identifiés :**")
                    st.dataframe(findings_df, use_container_width=True)

                    # Graphique de répartition des types de points
                    type_counts = findings_df['Type'].value_counts()
                    fig_findings = px.bar(x=type_counts.index, y=type_counts.values,
                                         title="Répartition des types de points importants",
                                         labels={'x': 'Type de point', 'y': 'Nombre'})
                    fig_findings.update_layout(xaxis_tickangle=45)
                    st.plotly_chart(fig_findings, use_container_width=True)
                else:
                    st.success("✅ Aucun point critique détecté dans les 10 premières colonnes analysées.")
        else:
            st.info("👆 Cliquez pour analyser automatiquement les points importants des données chargées.")
    else:
        st.info("💡 Chargez des données pour accéder à l'analyse des points importants.")

    # --- SECTION RÉSUMÉ ET ACTIONS RAPIDES ---
    st.markdown("---")
    st.subheader("⚡ Actions Rapides")

    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("📊 Charger les données", key="dashboard_quick_load"):
            st.info("👉 Allez dans l'onglet 'Application & Chat IA' pour charger les données.")
    with col2:
        if st.button("📈 Voir les analyses", key="dashboard_quick_analysis"):
            st.info("👉 Utilisez l'onglet 'Application & Chat IA' pour créer des analyses.")
    with col3:
        if st.button("📄 Générer un rapport", key="dashboard_quick_report"):
            st.info("👉 Configurez les options de rapport dans la barre latérale.")

# --- CONTENU DE L'ONGLET "Planification" ---
with schedule_tab:
    st.header("Planification & Tâches")
    st.markdown("## Prochaines étapes et tâches de développement")

    # Tableau de planification (exemple)
    planning_data = {
        "Tâche": [
            "Améliorer la gestion des erreurs dans l'exécution des scripts IA",
            "Ajouter la validation des entrées utilisateur (ex: types de colonnes)",
            "Optimiser la génération de rapports PDF (performance et fidélité)",
            "Intégrer la gestion des fichiers CSV pour l'upload et l'export",
            "Développer des fonctionnalités de filtrage interactif des données",
            "Mettre en place des tests unitaires pour les fonctions clés",
            "Explorer l'utilisation d'autres modèles IA pour l'analyse (ex: classification, clustering)",
            "Ajouter la possibilité de sauvegarder/charger des sessions d'analyse",
            "Implémenter une gestion plus fine des versions des données",
            "Améliorer le nettoyage des noms de colonnes générés par KoboToolbox",
            "Intégrer un sélecteur de thème pour l'interface Streamlit"
        ],
        "Responsable": [
            "Back-end / IA", "Front-end",
            "Back-end / PDF", "Full-stack",
            "Front-end",
            "Back-end / Test",
            "IA / Back-end",
            "Full-stack",
            "Back-end",
            "Data Engineering",
            "Front-end"
        ],
        "Statut": [
            "À faire", "À faire",
            "À faire", "À faire",
            "À faire",
            "À faire",
            "À faire",
            "À faire",
            "À faire",
            "En cours",
            "À faire"
        ],
        "Priorité": [
            "Haute", "Moyenne",
            "Haute", "Moyenne",
            "Moyenne",
            "Moyenne",
            "Basse",
            "Moyenne",
            "Basse",
            "Moyenne",
            "Basse"
        ]
    }

    planning_df = pd.DataFrame(planning_data)
    st.dataframe(planning_df, use_container_width=True)

    st.markdown("""
    ---
    **Note :** Ce plan de tâches est indicatif. Les priorités peuvent évoluer en fonction des besoins et des retours utilisateurs.
    Utilisez cet espace pour noter vos idées d'amélioration ou pour suivre l'avancement du projet.
    """)

# --- FONCTION DE GÉNÉRATION DE REQUÊTE SQL AVEC IA ---
def generate_python_script_with_ai(nl_request: str, df: pd.DataFrame, all_columns: list[str], numerical_cols: list[str], categorical_cols: list[str]):
    """
    Génère un script Python simple et exécutable sur le DataFrame courant à partir
    d'une demande en langage naturel. Utilise le gestionnaire IA pour la génération.
    
    Retourne (script_str, title_str).
    """
    try:
        ai = get_ai_handler()
        use_local = st.session_state.get('use_local_model', False)
        allow_fallback = st.session_state.get("allow_fallback_ai", False)
        gemini_model = get_effective_gemini_model_name()

        # Préparer le contexte pour l'IA
        context = f"""
Contexte des données:
- DataFrame avec {len(df)} lignes et {len(df.columns)} colonnes
- Colonnes disponibles: {df.columns.tolist()}
- Colonnes numériques: {numerical_cols}
- Colonnes catégorielles: {categorical_cols}
- Types de données: {df.dtypes.to_dict()}

Demande utilisateur: {nl_request}

Génère un script Python simple et exécutable qui répond à cette demande.
Le script doit:
1. Utiliser uniquement les colonnes mentionnées dans le contexte
2. Stocker le résultat dans une variable appelée 'result'
3. Gérer les erreurs potentielles
4. Être compatible avec pandas
5. Retourner un DataFrame ou une série pandas

Ne génère que le code Python, sans explications.
"""

        if use_local:
            if _local_gemma_ready():
                ai_response = ai.generate_text(context, provider="local")
            else:
                st.error("Gemma local activé mais non chargé: impossible de générer le script Python.")
                return "", ""
        else:
            if use_local and not _local_gemma_ready():
                st.info(
                    f"Gemma local indisponible (`{LOCAL_GEMMA_PATH}`). "
                    "Génération avec Gemini (modèle flash)."
                )
            ai_response = ai.generate_text(context, provider="gemini", model=gemini_model)

            # Fallback OpenAI si quota Gemini
            if (
                isinstance(ai_response, str)
                and ai_response.startswith("Erreur Gemini:")
                and _looks_like_gemini_quota_error(ai_response)
                and _has_openai_key()
            ):
                st.warning("Quota Gemini atteint (429). Basculement vers OpenAI.")
                ai_response = ai.generate_text(context, provider="openai", model=DEFAULT_OPENAI_MODEL)
        
        # Nettoyer et valider la réponse
        if "Erreur" not in ai_response and ai_response.strip():
            # Extraire le code Python de la réponse
            script = ai_response.strip()
            
            # S'assurer que le script commence par une déclaration de variable result
            if not script.strip().startswith('result'):
                script = f"# Script Python généré par IA\n{script}"
            
            title = f"Analyse IA: {nl_request[:50]}..."
            return script, title
        else:
            if allow_fallback:
                return _generate_fallback_script(nl_request, df, all_columns, numerical_cols, categorical_cols)
            st.error("Impossible de générer le script via IA (et mode secours désactivé).")
            return "", ""
            
    except Exception as e:
        if allow_fallback:
            return _generate_fallback_script(nl_request, df, all_columns, numerical_cols, categorical_cols)
        st.error(f"Erreur IA: {e}")
        return "", ""

def _generate_fallback_script(nl_request: str, df: pd.DataFrame, all_columns: list[str], numerical_cols: list[str], categorical_cols: list[str]):
    """
    Méthode de secours pour générer des scripts Python sans IA externe.
    """
    try:
        request_lower = (nl_request or "").lower()
        title = "Analyse Python générée"

        # Choix de colonnes par défaut (strictement parmi les colonnes retenues)
        cat_col = categorical_cols[0] if categorical_cols else None
        num_col = numerical_cols[0] if numerical_cols else None

        # Cas 1: Groupby / comptage (mots-clés forts)
        if any(k in request_lower for k in ["group", "grouper", "par ", "répartition", "count par", "compter par", "value_counts"]):
            # Essayer de détecter une colonne citée dans la requête
            group_col = None
            for c in (categorical_cols or []):
                if c and c.lower() in request_lower:
                    group_col = c
                    break
            if not group_col:
                group_col = cat_col
            title = f"Comptage par {group_col}" if group_col else "Comptage des valeurs"
            script = f"""
# Script Python généré (comptage)
result = None
if '{group_col}' in df.columns:
    result = (
        df['{group_col}']
          .value_counts(dropna=False)
          .reset_index()
          .rename(columns={{'index': '{group_col}', '{group_col}': 'Nombre'}})
    )
else:
    # Fallback: aperçu si la colonne n'existe pas
    result = df.head(20)
            """
            return script, title

        # Cas 2: Statistiques descriptives (mots-clés forts)
        if any(k in request_lower for k in ["stat", "statistique", "describe", "moyenne", "écart-type", "distribution"]):
            target_num = None
            for c in (numerical_cols or []):
                if c and c.lower() in request_lower:
                    target_num = c
                    break
            if not target_num:
                target_num = num_col
            title = f"Statistiques descriptives de {target_num}" if target_num else "Statistiques descriptives"
            script = f"""
# Script Python généré (statistiques descriptives)
result = None
num_cols = [c for c in df.select_dtypes(include=['number']).columns]
if '{target_num}' in df.columns:
    result = df['{target_num}'].describe().to_frame()
elif num_cols:
    result = df[num_cols].describe()
else:
    # Fallback: aperçu
    result = df.head(20)
            """
            return script, title

        # Cas 3: Tableau croisé simple (cat vs num) (mots-clés forts)
        if any(k in request_lower for k in ["moyenne par", "sum par", "agrég", "aggregate", "pivot", "tableau croisé"]):
            gcol = cat_col
            vcol = num_col
            title = f"Moyenne de {vcol} par {gcol}" if (gcol and vcol) else "Tableau croisé"
            agg_line = f"df.groupby('{gcol}')['{vcol}'].mean().reset_index(name='Moyenne')" if (gcol and vcol) else "df.head(20)"
            script = f"""
# Script Python généré (tableau croisé)
result = {agg_line}
            """
            return script, title

        # Heuristique par colonnes sélectionnées (si pas de mots-clés forts)
        if cat_col and not num_col:
            title = f"Comptage par {cat_col}"
            script = f"""
result = (
    df['{cat_col}']
      .value_counts(dropna=False)
      .reset_index()
      .rename(columns={{'index': '{cat_col}', '{cat_col}': 'Nombre'}})
)
            """
            return script, title

        if num_col and not cat_col:
            title = f"Statistiques descriptives de {num_col}"
            script = f"""
num_cols = [c for c in df.select_dtypes(include=['number']).columns]
result = df['{num_col}'].describe().to_frame() if '{num_col}' in df.columns else (df[num_cols].describe() if num_cols else df.head(20))
            """
            return script, title

        if cat_col and num_col:
            title = f"Moyenne de {num_col} par {cat_col}"
            script = f"""
result = df.groupby('{cat_col}')['{num_col}'].mean().reset_index(name='Moyenne')
            """
            return script, title

        # Fallback: aperçu minimal
        title = "Aperçu des données"
        script = f"""
# Script Python généré (aperçu et stats)
parts = []
parts.append(df.head(20))

# Répartition (première colonne catégorielle)
if '{cat_col}' in df.columns and '{cat_col}' != 'None':
    vc = df['{cat_col}'].value_counts(dropna=False).reset_index()
    vc.columns = ['{cat_col}', 'Nombre']
    parts.append(vc)

# Statistiques (première colonne numérique)
if '{num_col}' in df.columns and '{num_col}' != 'None':
    desc = df['{num_col}'].describe().to_frame()
    parts.append(desc)

# Fusion d'affichage (le consommateur concatène ou affiche séquentiellement)
result = parts[0] if parts else df.head(10)
            """
        return script, title
    except Exception as e:
        # En cas d'erreur, renvoyer un script d'aperçu minimal
        fallback = """
# Fallback: aperçu
result = df.head(20)
        """
        return fallback, f"Analyse Python - Erreur gérée ({e})"
